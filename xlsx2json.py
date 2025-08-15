"""
xlsx2json - Excel の名前付き範囲を JSON に変換するツール
"""

from __future__ import annotations
import re
import logging
import time
import json
import argparse
import datetime
import importlib
import importlib.util
import subprocess
import io
import sys
import shlex
import yaml
from contextlib import redirect_stdout, redirect_stderr
from dataclasses import dataclass, field
from pathlib import Path
from types import TracebackType
from typing import Any, Dict, List, Optional, Tuple, TypeGuard, Union, cast, Callable, Sequence, Iterable, Mapping

# モジュール全体で使用する外部ライブラリ
from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from jsonschema import Draft7Validator, FormatChecker

# ロガー
logger = logging.getLogger(__name__)

# 可読性のための型エイリアス
JSONScalar = Union[str, int, float, bool, None]
JSONValue = Union[JSONScalar, "JSONDict", "JSONList"]
JSONDict = Dict[str, JSONValue]
JSONList = List[JSONValue]
RectTuple = Tuple[int, int, int, int]
TransformRulesMap = Dict[str, List["ArrayTransformRule"]]

@dataclass
class ProcessingStats:
    """処理全体の統計情報を収集するシンプルなデータクラス。

    - containers_processed: コンテナ処理数
    - cells_generated: JSONセル（項目）生成数
    - cells_read: Excelセル読取数
    - empty_cells_skipped: 空セルをスキップした数
    - errors: 発生したエラーメッセージの一覧
    - start_time/end_time: 処理の開始/終了時刻（秒）
    """

    containers_processed: int = 0
    cells_generated: int = 0
    cells_read: int = 0
    empty_cells_skipped: int = 0
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    start_time: Optional[float] = None
    end_time: Optional[float] = None

    def start_processing(self) -> None:
        self.start_time = time.time()

    def end_processing(self) -> None:
        self.end_time = time.time()

    def add_error(self, message: str) -> None:
        self.errors.append(message)

    def log_summary(self) -> None:
        """収集した統計のサマリをINFOログに出力する。"""
        duration = None
        if self.start_time is not None and self.end_time is not None:
            duration = max(0.0, self.end_time - self.start_time)
        logger.info(
            "処理統計サマリ: containers=%d, cells_generated=%d, cells_read=%d, empty_skipped=%d, errors=%d, duration=%.3fs",
            self.containers_processed,
            self.cells_generated,
            self.cells_read,
            self.empty_cells_skipped,
            len(self.errors),
            duration if duration is not None else -1.0,
        )
        # テスト互換: 各項目を日本語で個別にも出力
        logger.info("処理されたコンテナ数: %d", self.containers_processed)
        logger.info("エラー数: %d", len(self.errors))
        logger.info("警告数: %d", len(self.warnings))

    # 以下はテスト互換のための補助API
    def reset(self) -> None:
        self.containers_processed = 0
        self.cells_generated = 0
        self.cells_read = 0
        self.empty_cells_skipped = 0
        self.errors.clear()
        self.start_time = None
        self.end_time = None

    def add_warning(self, message: str) -> None:
        # 現状は警告メッセージはエラーリストに含めずログのみ
        logger.warning(message)
        self.warnings.append(message)

    def get_duration(self) -> float:
        if self.start_time is None or self.end_time is None:
            return 0.0
        return max(0.0, self.end_time - self.start_time)


@dataclass(frozen=True)
class CLIConfig:
    """起動時に使用する CLI/派生設定をまとめたコンテナ。

    - args: argparse で解析されたコマンドライン引数
    - raw_config: 設定ファイルから読み込んだマッピング（存在する場合）
    - merged: CLI 上書き適用後のマージ済み設定
    """

    args: Any
    raw_config: Dict[str, Any]
    merged: Dict[str, Any]


@dataclass(frozen=True)
class RectChain:
    """軽量な矩形ユーティリティ: (top, left, bottom, right)

    提供する機能:
    - 幅/高さの取得
    - 他矩形との交差判定
    - 指定座標の包含判定
    - タプル化
    既存コードの差し替えは行わず、段階的リファクタ用の補助ユーティリティです。
    """

    top: int
    left: int
    bottom: int
    right: int

    def width(self) -> int:
        return max(0, self.right - self.left + 1)

    def height(self) -> int:
        return max(0, self.bottom - self.top + 1)

    def as_tuple(self) -> RectTuple:
        """モジュール互換の順序で矩形をタプルとして返す。

        注意: モジュールの残りの部分は矩形タプルを
        (sc, sr, ec, er) == (left, top, right, bottom) として期待している。
        RectChain のフィールドは (top, left, bottom, right) として格納されているが、
        このメソッドは既存のタプルベースのヘルパーとの互換性を維持するために意図的に (left, top, right, bottom) を返す。
        """
        return (self.left, self.top, self.right, self.bottom)

    def intersects(self, other: "RectChain") -> bool:
        return not (self.right < other.left or self.left > other.right or self.bottom < other.top or self.top > other.bottom)

    def contains(self, row: int, col: int) -> bool:
        return self.top <= row <= self.bottom and self.left <= col <= self.right


@dataclass(frozen=True)
class TransformContext:
    """変換ルール適用用のコンテキストホルダー。

    将来のリファクタリングで多くの位置引数ではなく単一のパラメータを受け取れるように、
    一般的に渡される値を保持します。
    """
    workbook: Any
    prefix: str
    transform_rules_map: TransformRulesMap
    insert_keys: List[str]
    root_result: JSONDict

    @staticmethod
    def from_processing_config(cfg: ProcessingConfig) -> "TransformContext":
        # 初期段階では空の root_result / transform_rules_map を仮で保持し後段で差し替える
        return TransformContext(
            workbook=None,
            prefix=cfg.prefix,
            transform_rules_map={},
            insert_keys=[],
            root_result={},
        )


@dataclass(frozen=True)
class NestedScanParams:
    """ネスト走査用のパラメータコンテナ。

    ワークブック関連パラメータを統合。
    """
    workbook: Any
    container_name: str
    direction: str
    current_positions: Dict[str, Tuple[int, int]]
    labels: List[str]
    policy: Any  # ExtractionPolicy
    parent_anchor: str
    parent_rects: List[Tuple[int, int, int, int]]
    ancestors_rects_chain: List[List[Tuple[int, int, int, int]]]
    generated_names: Dict[str, Any]
    num_positions: List[int]
    ends_with_numeric: bool
    target_sheet: Optional[str] = None
    ws0: Any = None



def is_json_dict(x: Any) -> TypeGuard[JSONDict]:
    """x が JSON オブジェクト (dict) の場合に True を返します。"""
    return isinstance(x, dict)


def is_json_list(x: Any) -> TypeGuard[JSONList]:
    """x が JSON 配列 (list) の場合に True を返します。"""
    return isinstance(x, list)


_NUMERIC_TOKEN_RE = re.compile(r"^\d+(?:-\d+)*$")


def is_numeric_token_string(value: Any) -> bool:
    """"1-2-3" のような数値トークン列かを判定する。"""
    if not isinstance(value, str):
        return False
    try:
        # ここに本来の数値トークン判定ロジックを記述（仮実装: 数値トークン正規表現）
        return bool(_NUMERIC_TOKEN_RE.match(value))
    except Exception:
        return False




def pick_effective_bounds(pt: int, pb: int, ancestors_rects_chain: List[List[RectTuple]]) -> Tuple[int, int]:
    """ネスト境界選択。

    契約:
    - 入力: 親の上端`pt`/下端`pb`、先祖矩形群`ancestors_rects_chain`
    - 出力: 実際に走査に使用する (eff_pt, eff_pb)
    - 例外: なし（失敗時も入力を返す）
    - 副作用: なし
    祖先がある場合はトップ祖先グループ境界を優先し、無ければ親境界を返す。
    """
    # ancestors_rects_chain が空でなければ、先頭（トップ祖先）だけを見て親矩形の上端に最も近い境界を選ぶ
    if ancestors_rects_chain:
        top_level = ancestors_rects_chain[0] or []
        # デフォルトは親境界
        eff_pt, eff_pb = pt, pb
        # 親上端 pt を含む（または近い）祖先境界にクリップ
        for (_al, at, _ar, ab) in top_level:
            if at <= pt <= ab:
                eff_pt, eff_pb = at, ab
                break
        return eff_pt, eff_pb
    return pt, pb


def trim_trailing_empty(seq: Any) -> Any:
    """1D/2D 配列の末尾の空(None/"")をトリムする共通ロジック。"""
    def _trim_1d(lst: list[Any]) -> list[Any]:
        while lst and (lst[-1] in (None, "")):
            lst.pop()
        return lst

    if isinstance(seq, list):
        if seq and all(not isinstance(x, list) for x in seq):
            return _trim_1d(seq)
        if seq and all(isinstance(x, list) for x in seq):
            return [_trim_1d(list(row)) for row in seq]
    return seq


# -----------------------------------------------------------------------------
# ネスト走査のヘルパー抽象
# -----------------------------------------------------------------------------

def align_row_phase(eff_pt: int, anchor_row: int, step: int) -> int:
    """開始行 `eff_pt` を、`anchor_row` と同じ位相になるように最小の行へ合わせる。

    要件: 最小の r >= eff_pt かつ (r - anchor_row) % step == 0 を返す。
    """
    if step <= 0:
        step = 1
    # 既に同位相ならそのまま
    if (eff_pt - anchor_row) % step == 0:
        return eff_pt
    # 次の位相へ切り上げ
    delta = (step - ((eff_pt - anchor_row) % step)) % step
    return eff_pt + (delta if delta != 0 else 0)

def derive_eff_step_local(
    *,
    labels_present: bool,
    ends_with_numeric: bool,
    workbook,
    container_name: str,
    target_sheet: Optional[str],
    eff_pt: int,
    eff_pb: int,
    direction: str,
    policy: NestedScanPolicy | None = None,
) -> int:
    """親矩形ごとの実効ステップを決定する。

    挙動:
    - labels が存在する場合は常に step=1。
    - それ以外: コンテナ名が数値で終わる（'.1'）場合は子アンカー矩形を取得し、その高さを [eff_pt, eff_pb] 内で使用。
      見つからない場合は step=1。列方向（column）はここでは未対応（既存実装は行方向の境界のみを使用）。
    """
    # labels → step=1
    if labels_present:
        return 1
    try:
        child_step_local: Optional[int] = None
        if ends_with_numeric and target_sheet is not None:
            rects_child_all = _get_anchor_rects_naive(
                workbook, container_name, target_sheet, col_tolerance=0
            )
            for _cl, _ct, _cr, _cb in rects_child_all or []:
                if _ct >= eff_pt and _cb <= eff_pb:
                    child_step_local = max(1, _cb - _ct + 1)
                    break
        return int(child_step_local) if child_step_local is not None else 1
    except Exception:
        return 1


def select_probe_fields(
    *,
    current_positions: Dict[str, Tuple[int, int]],
    labels: List[str],
    numeric_token_fields: List[str],
) -> List[str]:
    """優先順位でプローブ用フィールドを選択する: ラベルかつ数値トークン > ラベル > 数値トークン > 全フィールド（ソート）。"""
    label_fields = [lf for lf in (labels or []) if lf in current_positions]
    label_numeric = [lf for lf in label_fields if lf in numeric_token_fields]
    if label_numeric:
        return label_numeric
    if label_fields:
        return label_fields
    if numeric_token_fields:
        try:
            return sorted(numeric_token_fields)
        except Exception:
            return list(numeric_token_fields)
    try:
        return sorted(current_positions.keys())
    except Exception:
        return list(current_positions.keys())


def prepare_probe_fields(
    *,
    current_positions: Dict[str, Tuple[int, int]],
    labels: List[str],
    numeric_token_fields: Sequence[str],
) -> Tuple[List[str], List[str]]:
    """プローブ準備: 行ループ外で `resolved_labels` と `probe_fields` を決定する。

    入力:
    - `current_positions`: フィールド→セル位置
    - `labels`: 指定ラベル一覧（未解決）
    - `numeric_token_fields`: 数値トークン扱いのフィールド候補

    出力:
    - `(resolved_labels, probe_fields)` のタプル

    例外/副作用:
    - なし（安全に派生値を返すのみ）
    """
    resolved_labels = [lf for lf in (labels or []) if lf in current_positions]
    try:
        nt_fields: List[str] = list(numeric_token_fields)
    except Exception as e:
        logger.debug("numeric_token_fields fallback to list copy due to: %s", e)
        nt_fields = [x for x in numeric_token_fields]
    probe_fields = select_probe_fields(
        current_positions=current_positions,
        labels=labels,
        numeric_token_fields=nt_fields,
    )
    return resolved_labels, probe_fields


def find_local_anchor_row(
    *,
    ws,
    current_positions: Mapping[str, Tuple[int, int]],
    probe_fields: Sequence[str],
    numeric_probe_cols: Sequence[int],
    local_aligned_row: int,
    eff_pb: int,
    step: int,
    expected_len: int,
    expected_prefix: Sequence[str],
) -> Optional[int]:
    """指定ステップで [local_aligned_row..eff_pb] の範囲を走査し、
    数値トークンの期待に合致する最初の行を探す。見つからない場合は、
    任意のプローブ項目が非空となる行をフォールバックとして採用する。
    """
    SAFE_SCAN = 5000
    # 構造化ログ（探索条件）
    try:
        _sheet_name = getattr(ws, "title", "")
    except Exception:
        _sheet_name = ""
    logger.debug(
        "ANCHOR-FIND sheet=%s step=%s expected_len=%s range=[%s..%s] probe_cols=%s probe_fields=%s",
        _sheet_name,
        step,
        expected_len,
        local_aligned_row,
        eff_pb,
        list(numeric_probe_cols),
        list(probe_fields),
    )
    # 数値トークンにもとづき探索
    if len(numeric_probe_cols) > 0:
        r2 = local_aligned_row
        cnt2 = 0
        while r2 <= eff_pb and cnt2 < SAFE_SCAN:
            sval = ""
            for sc_col in numeric_probe_cols:
                sval = read_cell_value((sc_col, r2), ws)
                if is_numeric_token_string(sval):
                    break
            if is_numeric_token_string(sval):
                toks = [t for t in str(sval).split("-") if t]
                prefix_ok = len(toks) >= len(expected_prefix) and all(
                    toks[i] == expected_prefix[i] for i in range(len(expected_prefix))
                )
                if prefix_ok and len(toks) == expected_len:
                    logger.debug(
                        "ANCHOR-HIT sheet=%s row=%s prefix=%s tokens=%s",
                        _sheet_name,
                        r2,
                        "-".join(expected_prefix),
                        toks,
                    )
                    return r2
            r2 += step
            cnt2 += 1
    # フォールバック: プローブ項目のいずれかが非空となる行を採用
    r = local_aligned_row
    scan_cnt = 0
    while r <= eff_pb and scan_cnt < SAFE_SCAN:
        non_empty_probe = False
        for pf in probe_fields:
            pc, _pr0 = current_positions.get(pf, (None, None))
            if pc is None:
                continue
            val_probe = read_cell_value((pc, r), ws)
            if val_probe not in (None, ""):
                non_empty_probe = True
                break
        if non_empty_probe:
            logger.debug(
                "ANCHOR-FALLBACK sheet=%s row=%s",
                _sheet_name,
                r,
            )
            return r
        r += step
        scan_cnt += 1
    return None


def make_seq_spec_for_level(
    *, expected_len: int, group_indexes: List[int], parent_local_index: int
) -> SeqIndexSpec:
    """SeqIndexSpec を構築する。
    レベル2（<=2）の場合は parent_local を省略し、実行時のみ緩和（祖先プレフィックス一致のみ）とする。
    """
    if expected_len <= 2:
        return SeqIndexSpec(
            ancestor_prefix=tuple(str(x) for x in group_indexes),
            parent_local=None,
            expected_length=expected_len,
        )
    return SeqIndexSpec(
        ancestor_prefix=tuple(str(x) for x in group_indexes),
        parent_local=int(parent_local_index),
        expected_length=int(expected_len),
    )


def check_seq_accept_and_dedup(
    *,
    policy: NumericTokenPolicy,
    expected_len: int,
    has_numeric_series_field: bool,
    seq_like_val: Optional[str],
    group_indexes: List[int],
    parent_local_index: int,
    group_key_for_dedup: Tuple[int, ...],
    seen_tokens: Dict[Tuple[int, ...], set[str]],
) -> bool:
    """数値トークン仕様によるフィルタリングと、(group, parent) 単位での重複排除を適用する。
    受理した場合は seen_tokens を更新する。

    要素を採用する場合は True、スキップする場合は False を返す。
    """
    if expected_len < 2:
        return True
    if has_numeric_series_field:
        if seq_like_val is None:
            return False
        spec = make_seq_spec_for_level(
            expected_len=expected_len, group_indexes=list(group_indexes), parent_local_index=parent_local_index
        )
        if policy.strict_spec_match and not spec.matches(seq_like_val):
            return False
        seen_set = seen_tokens.setdefault(group_key_for_dedup, set())
        if seq_like_val in seen_set:
            return False
        seen_set.add(seq_like_val)
        return True
    # 数値トークン系列フィールドがない場合: 値が存在して仕様に反するなら不採用。値があるときは重複排除を適用
    if seq_like_val is not None:
        spec = make_seq_spec_for_level(
            expected_len=expected_len, group_indexes=list(group_indexes), parent_local_index=parent_local_index
        )
        if policy.strict_spec_match and not spec.matches(seq_like_val):
            return False
        seen_set = seen_tokens.setdefault(group_key_for_dedup, set())
        if seq_like_val in seen_set:
            return False
        seen_set.add(seq_like_val)
    return True


def should_skip_by_row_ownership(
    *,
    policy: NestedScanPolicy,
    expected_len: int,
    numeric_token_fields: List[str],
    used_positions: Dict[str, Tuple[int, int]],
    non_empty: bool,
    group_key: Tuple[int, ...],
    claims_by_group: Dict[Tuple[int, ...], set[int]],
) -> bool:
    """数値トークンフィールドが存在しない場合の行オーナーシップ抑止（lv2 以上のみ）。"""
    if not (policy.row_ownership_without_tokens and (not numeric_token_fields) and expected_len >= 2):
        return False
    try:
        row_key_candidates = [tr for (_tc, tr) in used_positions.values() if isinstance(tr, int)]
        if not row_key_candidates:
            return False
        row_key = min(row_key_candidates)
        claims = claims_by_group.setdefault(group_key, set())
        if row_key in claims:
            return True
        if non_empty:
            claims.add(row_key)
        return False
    except Exception:
        return False


# -----------------------------------------------------------------------------
# 階層的なシーケンスインデックス仕様
# -----------------------------------------------------------------------------

def parse_seq_tokens(s: str) -> List[str]:
    """数値トークン列を配列に分解（非数値や空は空配列）。

    環境変数の影響を受けない純粋パーサ。数値でないトークンは除外する。
    """
    if not isinstance(s, str) or not s:
        return []
    return [t for t in s.split("-") if t and t.isdigit()]


@dataclass(frozen=True)
class SeqIndexSpec:
    """階層的列挙インデックス仕様。

    - ancestor_prefix: 祖先のローカルインデックス列（例: ["1"], ["1","1"]）
    - parent_local: 直近の親のローカルインデックス（例: 2）。None なら未確定。
    - expected_length: この階層のトークン数（コンテナ名の数値トークン数）
    """

    ancestor_prefix: Tuple[str, ...]
    parent_local: Optional[int]
    expected_length: int

    def prefix(self) -> Tuple[str, ...]:
        if self.parent_local is None:
            return self.ancestor_prefix
        return (*self.ancestor_prefix, str(self.parent_local))

    def matches(self, value: Any) -> bool:
        """セル値が期待プレフィックス・期待長に一致するか。"""
        if not is_numeric_token_string(value):
            return False
        toks = parse_seq_tokens(str(value))
        if len(toks) != self.expected_length:
            return False
        # 厳格適用: parent_local が与えられていれば、祖先に続いて親ローカルも
        # 必須プレフィックスとして要求する。None の場合は祖先のみを要求。
        required_prefix: Tuple[str, ...]
        if self.parent_local is None:
            required_prefix = self.ancestor_prefix
        else:
            required_prefix = (*self.ancestor_prefix, str(self.parent_local))
        if len(toks) < len(required_prefix):
            return False
        for i, p in enumerate(required_prefix):
            if toks[i] != p:
                return False
        return True


# -----------------------------------------------------------------------------
# Rect のラッパー（補助ユーティリティ）
# -----------------------------------------------------------------------------


@dataclass(frozen=True)
class Rect:
    """Excelシート上の矩形領域（1始まり座標）。

    - top/left/bottom/right はセルの行・列番号（1-based）
    - completeness は罫線の完全度（0.0〜1.0）。不明な場合は None
    """
    top: int
    left: int
    bottom: int
    right: int
    completeness: Optional[float] = None


def rect_from_tuple(t: tuple[int, int, int, int, float] | tuple[int, int, int, int]) -> Rect:
    """タプルから `Rect` を生成する小ヘルパー。

    受け取る形は `(top, left, bottom, right[, completeness])`。
    """
    if len(t) == 4:
        top, left, bottom, right = t
        return Rect(top=top, left=left, bottom=bottom, right=right, completeness=None)
    if len(t) == 5:
        top5, left5, bottom5, right5, comp5 = t
        return Rect(top=top5, left=left5, bottom=bottom5, right=right5, completeness=comp5)
    # フォールバック（防御）: completeness 不明
    top4, left4, bottom4, right4 = cast(tuple[int, int, int, int], t)
    return Rect(top=top4, left=left4, bottom=bottom4, right=right4, completeness=None)


def detect_rectangular_regions_rects(worksheet, cell_names_map=None) -> List[Rect]:
    """`detect_rectangular_regions` のRect返却版。

    元関数はタプルを返すため、`Rect` へ変換して扱いやすくする。
    """
    regions = detect_rectangular_regions(worksheet, cell_names_map)
    return [rect_from_tuple(r) for r in regions]


def find_bordered_region_rect_around_positions(
    worksheet,
    positions: dict[str, tuple[int, int]],
    *,
    row_margin: int = 12,
    col_margin: int = 8,
) -> Rect | None:
    """与えられた複数のセル位置の周囲に存在する罫線矩形を概算で求める。"""
    t = find_bordered_region_around_positions(
        worksheet, positions, row_margin=row_margin, col_margin=col_margin
    )
    return Rect(top=t[0], left=t[1], bottom=t[2], right=t[3]) if t else None


# -----------------------------------------------------------------------------
# コンテナの領域境界推定（モジュールレベル）
# -----------------------------------------------------------------------------


def find_region_bounds_for_positions(
    workbook,
    container_name: str,
    sheet_name: str,
    pos_map: Dict[str, tuple[int, int]],
) -> RectTuple | None:
    """決定論的境界: コンテナ自身の範囲があればそれ、無ければ pos_map の最小外接矩形を
    1回だけ罫線完全度チェックして採用。"""
    if not pos_map:
        return None
    # 1) コンテナキー自身の範囲
    for sn, coord in iter_defined_name_destinations_all(container_name, workbook):
        eff_sn = sn
        if (not eff_sn) and isinstance(coord, str) and "!" in coord:
            eff_sn = coord.split("!", 1)[0]
        if eff_sn == sheet_name and coord and ":" in str(coord):
            (sc, sr), (ec, er) = parse_range(str(coord).replace("$", "").split("!", 1)[-1])
            return (sr, sc, er, ec)
    # 2) 最小外接矩形 + 罫線完全度チェック（1回）
    ws0 = workbook[sheet_name] if sheet_name in getattr(workbook, "sheetnames", []) else workbook.active
    cols = [c for (c, _r) in pos_map.values()]
    rows = [r for (_c, r) in pos_map.values()]
    top, left, bottom, right = min(rows), min(cols), max(rows), max(cols)
    logger.debug(
        "BOUNDS-CHECK sheet=%s bounds top=%s left=%s bottom=%s right=%s",
        sheet_name,
        top,
        left,
        bottom,
        right,
    )
    comp = calculate_border_completeness(ws0, top, left, bottom, right)
    return (top, left, bottom, right) if comp >= 1.0 else None


@dataclass(frozen=True)
class ProcessingConfig:
    """処理設定を管理するデータクラス"""

    input_files: List[Union[str, Path]] = field(default_factory=list)
    prefix: str = "json"
    trim: bool = False
    output_dir: Optional[Path] = None
    output_format: str = "json"
    schema: Optional[Dict[str, Any]] = None
    containers: Dict[str, Any] = field(default_factory=dict)
    transform_rules: List[str] = field(default_factory=list)
    max_elements: Optional[int] = None
    # 出力形状オプション: 指定したルート名を配列のオブジェクト（{groupKey: {...}}）にラップ
    # ログフォーマット（デフォルトはタイムスタンプ付き）。設定/CLIで上書き可能。
    log_format: Optional[str] = None


@dataclass
class Context:
    """実行時コンテキスト。

    - processing_stats: 処理統計（アクセスは原則 `stats()` アクセサ経由）
    - border_cache / anchor_rects_cache: 罫線/アンカー矩形のキャッシュ
    """
    processing_stats: "ProcessingStats"
    border_cache: dict[tuple, bool] = field(default_factory=dict)
    anchor_rects_cache: dict[tuple, list] = field(default_factory=dict)


# 現在のグローバル実行コンテキスト（後方互換のため初期化時に processing_stats を共有）
_CURRENT_CONTEXT: Context | None = None


def get_current_context() -> Context:
    global _CURRENT_CONTEXT, processing_stats
    if _CURRENT_CONTEXT is None:
        # 初期化: 既存の processing_stats と空キャッシュでコンテキストを生成
        _CURRENT_CONTEXT = Context(processing_stats=processing_stats)
    return _CURRENT_CONTEXT


def set_current_context(ctx: Context) -> None:
    """現在の実行コンテキストを設定し、後方互換のため processing_stats を同期する。

    直接 `processing_stats` グローバルを参照するコードは将来的に削除予定のため、
    新規コードは `stats()` を利用してください。
    """
    global _CURRENT_CONTEXT, processing_stats
    _CURRENT_CONTEXT = ctx
    processing_stats = ctx.processing_stats


def border_cache() -> dict:
    return get_current_context().border_cache


def anchor_rects_cache() -> dict:
    return get_current_context().anchor_rects_cache


def stats() -> "ProcessingStats":
    """現在の処理統計を返す（Context 経由）。"""
    return get_current_context().processing_stats


class Xlsx2JsonConverter:
    """Excel から JSON への変換を行うメインクラス"""

    def __init__(self, config: ProcessingConfig):
        self.config = config
        self.processing_stats = ProcessingStats()
        self.validator = None
        if config.schema:
            # date-time / time などの format 検証を有効化
            self.validator = Draft7Validator(config.schema, format_checker=FormatChecker())

    def process_files(self, input_files: List[Union[str, Path]]) -> int:
        """ファイルリストを処理する"""
        # グローバル統計とインスタンス統計を同一インスタンスに統一
        # これにより、内部関数群が参照する stats()（実体は processing_stats）増分が
        # そのままコンバータのサマリに反映される
        self.processing_stats.start_processing()
        # Context に集約（後方互換のため processing_stats も同期）
        set_current_context(Context(processing_stats=self.processing_stats))

        try:
            xlsx_files = self._collect_xlsx_files(input_files)
            for xlsx_file in xlsx_files:
                try:
                    self._process_single_file(xlsx_file)  # 各ファイルを処理
                except Exception as e:
                    # 個別ファイルのエラーはログに記録するが処理は継続
                    self.processing_stats.add_error(
                        f"ファイル処理エラー {xlsx_file}: {e}"
                    )
                    # 上位ループでの検出・記録（スタックトレース付き）
                    logger.exception(
                        f"ファイル処理中に例外。処理を継続します: {xlsx_file}"
                    )
        except Exception as e:
            self.processing_stats.add_error(f"処理中にエラーが発生: {e}")
            # ここは最上位ハンドラとしてスタックトレースを残す
            logger.exception("処理全体で未処理例外が発生しました")
            return 1
        finally:
            self.processing_stats.end_processing()
            self.processing_stats.log_summary()

        # エラーがあっても処理完了の場合は0を返す（従来の動作を維持）
        return 0

    def _collect_xlsx_files(self, inputs: List[Union[str, Path]]) -> List[Path]:
        """入力からXLSXファイルを収集"""
        files = []
        for input_item in inputs:
            path = Path(input_item)
            if path.is_file() and path.suffix.lower() == ".xlsx":
                files.append(path)
            elif path.is_dir():
                files.extend(path.glob("*.xlsx"))
        return files

    def _process_single_file(self, xlsx_file: Path) -> None:
        """単一ファイルの処理"""
        logger.debug(f"Processing: {xlsx_file}")
        # ワークブック毎にキャッシュをクリア（Context 経由）
        border_cache().clear()
        anchor_rects_cache().clear()
        try:
            # 変換ルールの処理
            array_transform_rules = None
            if self.config.transform_rules:
                array_transform_rules = parse_array_transform_rules(
                    self.config.transform_rules,
                    self.config.prefix,
                    self.config.schema,
                    self.config.trim,
                )

            # 解析を実行（global_max_elements は None の場合は渡さない）
            _extra: Dict[str, Any] = {}
            if self.config.max_elements is not None:
                _extra["global_max_elements"] = self.config.max_elements
            data = parse_named_ranges_with_prefix(
                xlsx_file,
                self.config.prefix,
                array_split_rules=None,
                array_transform_rules=array_transform_rules,
                containers=self.config.containers,
                schema=self.config.schema,
                **_extra,
            )

            # 出力ディレクトリの決定（未指定なら <xlsx_dir>/output）
            out_dir = (
                Path(self.config.output_dir)
                if self.config.output_dir
                else (xlsx_file.parent / "output")
            )
            base_name = xlsx_file.stem
            self._write_output(data, out_dir, base_name)
        except Exception as e:
            # ここはファイル単位の最上位ハンドラ。例外を記録して継続可能。
            logger.exception("単一ファイルの処理中に例外が発生しました")
            self.processing_stats.add_error(f"ファイル処理エラー {xlsx_file}: {e}")

    def _write_output(self, data: dict, output_dir: Path, base_name: str) -> None:
        """データ出力を書き込み（JSON/YAML対応）"""
        # 出力直前にプレフィックス配下をルートへ統合し、プレフィックス配下の重複を除去
        # 期待動作: ルート直下と prefix 配下に同一項目がある場合、prefix 配下は出力しない
        if is_json_dict(data):
            pref_key = self.config.prefix
            if pref_key in data and is_json_dict(data[pref_key]):
                # コンテナから group→root マップを構築（例: lv1 -> ツリー1）
                group_to_root: dict[str, str] = {}
                try:
                    if self.config.containers:
                        for cont_key in self.config.containers.keys():
                            parts = [p for p in cont_key.split(".") if p]
                            if (
                                len(parts) >= 4
                                and parts[0] == pref_key
                                and parts[2]
                                and parts[3].isdigit()
                            ):
                                root_name = parts[1]
                                group_label = parts[2]
                                group_to_root[group_label] = root_name
                except Exception:
                    # ここは「最終出力前のラベルマッピング最適化」のみを行う非本質ロジック。
                    # コンテナ仕様が異常でも致命ではないため、安全側で握りつぶして
                    # マッピング無し（素通し）にフォールバックする。
                    group_to_root = {}
                # ルートへ統合（既存のルート値を優先: 既にルートにあれば上書きしない）
                merged = {k: v for k, v in data.items() if k != pref_key}
                pref_val_any = data[pref_key]
                # 予期しない型の場合は統合スキップ（安全側）
                pref_obj: Dict[str, JSONValue] = (
                    pref_val_any if is_json_dict(pref_val_any) else {}
                )
                for k, v in pref_obj.items():
                    # グループラベルキー（lv1等）は対応するルートキーへ吸収してマージ
                    if k in group_to_root:
                        merged[group_to_root[k]] = v
                        continue
                    merged.setdefault(k, v)
                data = merged

        # 出力フォーマットに応じて拡張子を決定
        extension = ".yaml" if self.config.output_format == "yaml" else ".json"
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"{base_name}{extension}"

        # 空値は常に抑制して出力
        write_data(
            data,
            output_path,
            self.config.output_format,
            self.config.schema,
            self.validator,
            True,
        )


# グローバル統計インスタンス
processing_stats = ProcessingStats()


# =============================================================================
def _is_2d_list(val: Any) -> bool:
    return is_json_list(val) and bool(val) and isinstance(val[0], (list, tuple))


def _coerce_to_1d(val: Any) -> Any:
    if _is_2d_list(val):
        # 2D は 1D へは潰さない（情報落ち防止）。呼び出し側で 2D 優先にするため、通常ここに来ない想定。
        return val
    if not is_json_list(val):
        return [val]
    return val


def _coerce_to_2d(val: Any) -> Any:
    if _is_2d_list(val):
        return val
    if not is_json_list(val):
        return [[val]]
    # 1D → 2D
    if not val or not isinstance(val[0], (list, tuple)):
        return [val]
    return val


def _normalize_field_shapes_in_list_of_dicts(lst: List[Dict[str, Any]]) -> None:
    # フィールドごとの望ましい形状を決定（2D > 1D > scalar）
    field_names: set[str] = set()
    for d in lst:
        if is_json_dict(d):
            field_names.update(d.keys())
    desired: Dict[str, str] = {}
    for f in field_names:
        has_2d = False
        has_1d = False
        for d in lst:
            if not is_json_dict(d) or f not in d:
                continue
            v = d.get(f)
            if _is_2d_list(v):
                has_2d = True
                break
            if is_json_list(v):
                has_1d = True
        if has_2d:
            desired[f] = "2D"
        elif has_1d:
            desired[f] = "1D"
    # 望ましい形状に強制
    for f, shape in desired.items():
        for d in lst:
            if not is_json_dict(d) or f not in d:
                continue
            v = d.get(f)
            if shape == "2D":
                d[f] = _coerce_to_2d(v)
            elif shape == "1D":
                if isinstance(v, list) and bool(v) and isinstance(v[0], (list, tuple)):
                    d[f] = v
                elif not isinstance(v, list):
                    d[f] = [v]
                else:
                    d[f] = v


def _normalize_field_shapes_across_lists_of_dicts(
    nested: List[List[Dict[str, Any]]],
) -> None:
    # 全サブリストを横断してフィールドの望ましい形状を決定（2D > 1D > scalar）
    desired: Dict[str, str] = {}
    # 一旦、観測された最大形状を記録
    for sub in nested:
        if not is_json_list(sub) or not all(
            is_json_dict(x) for x in sub if x is not None
        ):
            continue
        # サブリスト単位の標準化を先に行う
        _normalize_field_shapes_in_list_of_dicts(cast(List[Dict[str, Any]], sub))  # 局所整合
        for d in sub:
            if not is_json_dict(d):
                continue
            for f, v in d.items():
                if _is_2d_list(v):
                    desired[f] = "2D"
                elif is_json_list(v) and desired.get(f) != "2D":
                    desired.setdefault(f, "1D")
    if not desired:
        return
    # 決まった desired に合わせて全体を再強制
    for sub in nested:
        if not is_json_list(sub) or not all(
            is_json_dict(x) for x in sub if x is not None
        ):
            continue
        for d in sub:
            if not is_json_dict(d):
                continue
            for f, shape in desired.items():
                if f not in d:
                    continue
                v = d.get(f)
                if shape == "2D":
                    d[f] = _coerce_to_2d(v)
                elif shape == "1D":
                    if isinstance(v, list) and bool(v) and isinstance(v[0], (list, tuple)):
                        d[f] = v
                    elif not isinstance(v, list):
                        d[f] = [v]
                    else:
                        d[f] = v


def normalize_array_field_shapes(obj: Any) -> JSONValue:
    """配列内の同名フィールドの形状（スカラ/1D/2D）を横並びで統一する（再帰）。
    - list-of-dicts ではフィールドごとに 2D > 1D > scalar の優先で整合
    - list-of-list-of-dicts では全サブリストを横断して整合
    - dict/list は再帰的に処理
    """
    if is_json_list(obj):
        # list-of-list-of-dicts を先に検出
        if all((not it) or is_json_list(it) for it in obj):
            # サブリストの中身が dict であるものがあるか？
            if any(
                is_json_list(it) and all(is_json_dict(x) for x in it if x is not None)
                for it in obj
            ):
                # 形状整合（横断）
                _normalize_field_shapes_across_lists_of_dicts(cast(List[List[Dict[str, Any]]], obj))  # in-place
                # 再帰
                return [normalize_array_field_shapes(it) for it in obj]
        # list-of-dicts の場合
        if all(is_json_dict(it) for it in obj if it is not None):
            _normalize_field_shapes_in_list_of_dicts(cast(List[Dict[str, Any]], obj))  # in-place
            return [normalize_array_field_shapes(it) for it in obj]
        # その他の list は要素を再帰処理
        return [normalize_array_field_shapes(it) for it in obj]
    if is_json_dict(obj):
        return {k: normalize_array_field_shapes(v) for k, v in obj.items()}
    return obj


# =============================================================================
# Validation-time ISO conversion (shared abstraction)
# =============================================================================

def to_iso_for_validation(obj: Any) -> JSONValue:
    """jsonschema 検証前に datetime/date/time を ISO 文字列へ正規化する共通関数。

    - list/dict は再帰処理
    - それ以外はそのまま返す
    """
    if is_json_list(obj):
        return [to_iso_for_validation(x) for x in obj]
    if is_json_dict(obj):
        return {k: to_iso_for_validation(v) for k, v in obj.items()}
    if isinstance(obj, datetime.datetime):
        return obj.isoformat()
    if isinstance(obj, datetime.date):
        return obj.isoformat()
    if isinstance(obj, datetime.time):
        return obj.isoformat()
    return obj


# =============================================================================
def get_generated_names_map(wb) -> Optional[Dict[str, Any]]:
    """ワークブックの生成名オーバーライドマップが存在すれば返し、なければ None を返す。"""
    try:
        return GeneratedNames.for_workbook(wb).as_dict()
    except Exception:
        return None


def set_generated_name(wb, name: str, value: Any) -> None:
    """ワークブックレベルのオーバーライドマップに生成名を登録または上書きします。"""
    if not name:
        return
    try:
        GeneratedNames.for_workbook(wb).set(name, value)
    except Exception:
        logger.debug("failed to set generated name %s", name, exc_info=True)


class GeneratedNames:
    """Thin wrapper around workbook-scoped generated-names map.

    Stores the map on Workbook as attribute `_generated_names` (backwards compatible)
    and provides get/set/iter helpers for consistent access.
    """

    def __init__(self, wb) -> None:
        self._wb = wb

    @classmethod
    def for_workbook(cls, wb) -> "GeneratedNames":
        return cls(wb)

    def _ensure_map(self) -> dict:
        gm = getattr(self._wb, "_generated_names", None)
        if gm is None:
            gm = {}
            try:
                setattr(self._wb, "_generated_names", gm)
            except Exception:
                # fallback: try to attach to properties if available
                props = getattr(self._wb, "properties", None)
                if props is None:
                    class _P:  # pragma: no cover - defensive
                        pass

                    props = _P()
                    try:
                        self._wb.properties = props
                    except Exception:
                        pass
                props.__dict__["_generated_names"] = gm
        return gm

    def as_dict(self) -> Dict[str, Any]:
        gm = getattr(self._wb, "_generated_names", None)
        if not is_json_dict(gm):
            return {}
        return dict(gm)

    def set(self, name: str, value: Any) -> None:
        gm = self._ensure_map()
        gm[name] = value

    def get(self, name: str, default: Any = None) -> Any:
        gm = getattr(self._wb, "_generated_names", None)
        if not is_json_dict(gm):
            return default
        return gm.get(name, default)

    def iter_keys(self):
        gm = getattr(self._wb, "_generated_names", None)
        if not is_json_dict(gm):
            return iter(())
        return iter(gm.keys())


def prepare_containers_and_generated_names(
    wb,
    *,
    prefix: str,
    containers: Optional[Dict[str, Any]],
    global_max_elements: Optional[int],
    extraction_policy: ExtractionPolicy,
) -> tuple[Optional[Dict[str, Any]], bool, Dict[str, Any]]:
    """コンテナ設定のマージ/推論と、生成セル名の登録をまとめて行う。

    返り値: (containers, user_provided_containers, generated_names)
    - containers: 手動/自動推論を反映した最終コンテナ
    - user_provided_containers: 呼び出し元が手動で指定したかどうか（挙動の分岐に使用）
    - generated_names: 生成されたセル名マップ（副作用として wb の _generated_names に登録済み）
    """
    user_provided = containers is not None
    inferred = infer_containers_from_named_ranges(wb, prefix)
    # マージ方針: 手動優先
    if containers and inferred:
        logger.debug(f"コンテナを自動推論（マージ）: {inferred}")
        _merged = inferred.copy()
        _merged.update(containers)
        containers = _merged
    elif not containers and inferred:
        logger.debug(f"コンテナを自動推論: {inferred}")
        containers = inferred

    generated: Dict[str, Any] = {}
    if containers:
        logger.debug(f"コンテナ処理開始: {len(containers)}個のコンテナ")
        generated = generate_cell_names_from_containers(
            containers, wb, global_max_elements, prefix=prefix, extraction_policy=extraction_policy
        )
        # 生成されたセル名を _generated_names に登録（既存定義名があってもオーバーライド可能）
        for name, range_ref in generated.items():
            prefixed_name = name if name.startswith(f"{prefix}.") else f"{prefix}.{name}"
            set_generated_name(wb, prefixed_name, range_ref)
            if prefixed_name in wb.defined_names:
                logger.debug(
                    "生成名を既存定義名に対するオーバーライドとして登録: %s", prefixed_name
                )
            else:
                logger.debug("生成名を登録: %s -> %r", prefixed_name, range_ref)
        logger.debug(f"コンテナ処理完了: {len(generated)}個のセル名を生成")
    return containers, user_provided, generated


def build_all_names_with_generated(wb) -> tuple[Dict[str, Any], set[str]]:
    """定義名辞書に生成名を統合して返す。

    - 返り値: (all_names, defined_only_name_keys)
    - 生成名は DefinedName 互換の簡易オブジェクトを作って destinations を模擬
    """
    all_names: Dict[str, Any] = dict(wb.defined_names.items())
    defined_only_name_keys: set[str] = set(all_names.keys())

    gm = get_generated_names_map(wb)
    if gm:
        logger.debug(f"生成されたセル名を処理対象に追加: {len(gm)}個")
        for gen_name, gen_range in gm.items():
            if gen_name in all_names:
                continue
            # 簡易的なDefinedName互換を提供
            class GeneratedDefinedName:
                def __init__(self, attr_text):
                    self.attr_text = attr_text
                    if "!" in attr_text:
                        sheet_part, range_part = attr_text.split("!")
                        self.destinations = [(sheet_part, range_part)]
                    else:
                        self.destinations = [("Sheet1", attr_text)]

            all_names[gen_name] = GeneratedDefinedName(gen_range)
            logger.debug(f"生成セル名追加: {gen_name} -> {gen_range}")
    return all_names, defined_only_name_keys


# Core Utilities
# =============================================================================


class SchemaLoader:
    """
    JSONスキーマの読み込みと管理を行うクラス。

    公開APIの基準は本クラスです。テスト互換のため、同等機能のトップレベル関数
    load_schema / validate_and_log も提供していますが、将来的には
    SchemaLoader.* への一本化を推奨します（関数は互換用エイリアス扱い）。
    """

    @staticmethod
    def load_schema(schema_path: Optional[Path]) -> Optional[Dict[str, Any]]:
        """指定されたパスからJSONスキーマを読み込む"""
        if not schema_path:
            return None

        if not schema_path.exists():
            raise FileNotFoundError(f"スキーマファイルが見つかりません: {schema_path}")

        if not schema_path.is_file():
            raise ValueError(f"指定されたパスはファイルではありません: {schema_path}")

        try:
            with schema_path.open("r", encoding="utf-8") as f:
                return json.load(f)
        except json.JSONDecodeError as e:
            logger.error(f"無効なJSONフォーマットです: {schema_path} - {e}")
            raise  # 元のJSONDecodeErrorを再発生
        except Exception as e:
            raise IOError(
                f"スキーマファイルの読み込みに失敗: {schema_path} - {e}"
            ) from e

    @staticmethod
    def validate_and_log(
        data: Dict[str, Any], validator: Draft7Validator, log_dir: Path, base_name: str
    ) -> None:
        """JSONデータをバリデートし、エラーがあればファイルに出力"""
        # 可能なら datetime/date/time を ISO 文字列化してからチェック
        data2 = cast(Dict[str, Any], to_iso_for_validation(data))
        errors = sorted(validator.iter_errors(data2), key=lambda e: e.path)
        if not errors:
            return

        log_dir.mkdir(parents=True, exist_ok=True)
        error_log = log_dir / f"{base_name}.error.log"
        with error_log.open("w", encoding="utf-8") as f:
            for err in errors:
                path = ".".join(str(p) for p in err.path) if err.path else "<root>"
                f.write(f"[{path}]: {err.message}\n")


def reorder_json(
    obj: Union[JSONDict, List[Any], Any], schema: Dict[str, Any]
) -> Union[JSONDict, List[Any], Any]:
    """
    スキーマの properties 順に dict のキーを再帰的に並べ替える。
    list の場合は項目ごとに再帰処理。
    その他はそのまま返す。
    """
    if is_json_dict(obj) and is_json_dict(schema):
        ordered: Dict[str, Any] = {}
        props_any = schema.get("properties", {})
        props: Dict[str, Any] = props_any if isinstance(props_any, dict) else {}
        # スキーマ順に追加
        for key, subschema in props.items():
            if key in obj:
                ordered[key] = reorder_json(obj[key], subschema)
        # 追加キー（スキーマ未定義）: 挿入順を維持（特定名称の優先は行わない）
        for key in list(obj.keys()):
            if key in props:
                continue
            ordered[key] = obj[key]
        return ordered

    if is_json_list(obj) and is_json_dict(schema) and "items" in schema:
        items_any = schema.get("items")
        items_schema: Dict[str, Any] = items_any if isinstance(items_any, dict) else {}
        return [reorder_json(item, items_schema) for item in obj]

    return obj
def apply_post_parse_pipeline(
    *,
    result: Dict[str, Any],
    root_first_pos: Dict[str, tuple[int, int, int]],
    prefix: str,
    user_provided_containers: bool,
    containers: Optional[Dict[str, Any]],
    array_transform_rules: Optional[Dict[str, List[ArrayTransformRule]]],
    normalized_prefix: str,
    group_labels: set[str],
    group_to_root: Dict[str, str],
    gen_map: Optional[Dict[str, Any]],
) -> Dict[str, Any]:
    """パース後の出力整形パイプラインを一括適用。

    ステップ:
    1) ルートキー順の安定化（Excel読取順）
    2) ワイルドカード変換の適用
    3) コンテナ出力のリシェイプ（dict of lists → list of dicts）
    4) 明示コンテナ時の prefix 配下の子をトップレベルへ複製（互換）
    5) コンテナ無し時のグループ吸収のフォールバック正規化
    """
    # 1) ルート順安定化
    result2, root_result = reorder_roots_by_sheet_order(
        result, root_first_pos, prefix, user_provided_containers
    )
    # 2) 変換（ワイルドカード／非ワイルドカード含む）
    if array_transform_rules:
        transformed = apply_pattern_transforms(
            root_result, array_transform_rules, normalized_prefix
        )
        if user_provided_containers:
            result2[prefix] = transformed
        else:
            result2 = transformed
    # 3) コンテナ reshape
    result2 = reshape_containers_in_result(
        result2, containers, prefix, user_provided_containers
    )
    # 4) prefix の子をトップに複製（互換）
    if user_provided_containers and is_json_dict(result2) and prefix in result2:
        result2 = replicate_prefix_children_to_top_level(
            result2, prefix, group_labels, root_first_pos
        )
    # 5) コンテナ無しフォールバック
    if not user_provided_containers and is_json_dict(result2):
        result2 = fallback_normalize_root_groups_without_containers(
            result2, group_to_root, gen_map, normalized_prefix
        )
    return result2


def get_named_range_values(wb, defined_name) -> Any:
    """
    Excel の NamedRange からセル値を抽出し、単一セルは値、範囲は行優先のフラットな一次元リストで返す。

    従来互換のデフォルト挙動：
      - 1セル → スカラ
      - 1xN/Nx1 → 長さNの一次元配列
      - MxN     → 長さM*Nの一次元配列（行優先）

    形状保持が必要な特殊ケースは別ヘルパー（get_named_range_values_preserve_shape）で対応する。
    """
    flat_values: List[Any] = []
    single_cell_only = True
    for sheet_name, coord in getattr(defined_name, "destinations", []) or []:
        # #REF! や空座標はスキップ
        try:
            if not coord or (isinstance(coord, str) and "REF" in str(coord).upper()):
                logger.debug("Skipping invalid named range destination: sheet=%r coord=%r", sheet_name, coord)
                continue
            eff_sheet = sheet_name
            eff_coord = coord
            # coord にシート名が含まれる形式に対応
            if isinstance(eff_coord, str) and "!" in eff_coord:
                try:
                    sheet_part, cell_part = eff_coord.split("!", 1)
                    eff_coord = cell_part
                    if not eff_sheet:
                        eff_sheet = sheet_part
                except Exception:
                    pass
            # Worksheet オブジェクトをタイトルへ
            if eff_sheet is not None and not isinstance(eff_sheet, str):
                try:
                    eff_sheet = getattr(eff_sheet, "title", str(eff_sheet))
                except Exception:
                    eff_sheet = str(eff_sheet)
            ws = wb[eff_sheet] if eff_sheet else wb.active
            cell_or_range = ws[eff_coord]
        except Exception as e:
            logger.debug("Skipping destination due to error: sheet=%r coord=%r err=%s", sheet_name, coord, e)
            continue
        if isinstance(cell_or_range, tuple):
            single_cell_only = False
            for row in cell_or_range:
                if isinstance(row, tuple):
                    for cell in row:
                        flat_values.append(getattr(cell, "value", cell))
                else:
                    flat_values.append(getattr(row, "value", row))
        else:
            flat_values.append(getattr(cell_or_range, "value", cell_or_range))

    if not flat_values:
        # すべてのdestinationが無効/スキップされた場合は例外でフォールバックへ
        raise ValueError("No valid destinations for defined name")
    # 単一セルのみの名前付き範囲はスカラを返す
    if single_cell_only and len(flat_values) == 1:
        return flat_values[0]
    return flat_values


def get_named_range_values_preserve_shape(wb, defined_name) -> Any:
    """
    Excel の NamedRange からセル値を抽出し、可能なら形状（1D/2D）を保持して返すヘルパー。
    - 単一セル → スカラ
    - 1xN または Nx1 → 1次元配列
    - MxN (M,N>1) → 2次元配列
    """
    rows_all_sheets: List[List[Any]] = []
    for sheet_name, coord in getattr(defined_name, "destinations", []) or []:
        # #REF! や空座標はスキップ
        try:
            if not coord or (isinstance(coord, str) and "REF" in str(coord).upper()):
                logger.debug("Skipping invalid named range destination (preserve_shape): sheet=%r coord=%r", sheet_name, coord)
                continue
            eff_sheet = sheet_name
            eff_coord = coord
            if isinstance(eff_coord, str) and "!" in eff_coord:
                try:
                    sheet_part, cell_part = eff_coord.split("!", 1)
                    eff_coord = cell_part
                    if not eff_sheet:
                        eff_sheet = sheet_part
                except Exception:
                    pass
            if eff_sheet is not None and not isinstance(eff_sheet, str):
                try:
                    eff_sheet = getattr(eff_sheet, "title", str(eff_sheet))
                except Exception:
                    eff_sheet = str(eff_sheet)
            ws = wb[eff_sheet] if eff_sheet else wb.active
            cell_or_range = ws[eff_coord]
        except Exception as e:
            logger.debug("Skipping destination (preserve_shape) due to error: sheet=%r coord=%r err=%s", sheet_name, coord, e)
            continue
        if isinstance(cell_or_range, tuple):
            local_rows: List[List[Any]] = []
            for row in cell_or_range:
                if isinstance(row, tuple):
                    local_rows.append([getattr(cell, "value", cell) for cell in row])
                else:
                    local_rows.append([getattr(row, "value", row)])
            rows_all_sheets.extend(local_rows)
        else:
            rows_all_sheets.append([getattr(cell_or_range, "value", cell_or_range)])

    if not rows_all_sheets:
        # すべてのdestinationが無効/スキップされた場合は例外でフォールバックへ
        raise ValueError("No valid destinations for defined name (preserve_shape)")
    total_rows = len(rows_all_sheets)
    max_cols = max((len(r) for r in rows_all_sheets), default=0)
    if total_rows == 1 and max_cols == 1:
        return rows_all_sheets[0][0]
    if total_rows == 1:
        return rows_all_sheets[0]
    if max_cols == 1:
        return [r[0] if r else None for r in rows_all_sheets]
    return rows_all_sheets


# =============================================================================
# Container Support Functions
# =============================================================================


def compute_top_left_pos(
    defined_name_obj, sheet_order: Dict[str, int]
) -> tuple[int, int, int]:
    """DefinedName から読み取り順キー (sheet_idx, row, col) を返す共通関数。

    - 並び順は Excel のシート順 → 行 → 列
    - 範囲の場合は左上座標、単一セルの場合はその座標
    - 異常時は非常に大きい値のタプルを返し、末尾に回す
    """
    best: tuple[int, int, int] | None = None
    try:
        for sheet_name, coord in getattr(defined_name_obj, "destinations", []) or []:
            c = coord.replace("$", "") if isinstance(coord, str) else str(coord)
            r: int
            col: int
            if ":" in c:
                (sc, sr), (_ec, _er) = parse_range(c)
                r, col = sr, sc
            else:
                m = re.match(r"^([A-Z]+)(\d+)$", c)
                if not m:
                    continue
                col = column_index_from_string(m.group(1))
                r = int(m.group(2))
            si = sheet_order.get(sheet_name, 10**9)
            cand = (si, r, col)
            if best is None or cand < best:
                best = cand
    except Exception:
        best = None
    return best if best is not None else (10**9, 10**9, 10**9)


def is_nonempty_array_or_dict(x: Any) -> bool:
    """リスト/辞書が実質的に非空かを判定するヘルパー。その他型は False。

    is_completely_empty を用いて空構造（空配列/空オブジェクト/空値のみから成る入れ子）を検出。
    """
    if isinstance(x, list) or isinstance(x, dict):
        return not is_completely_empty(x)
    return False


def collect_root_first_positions(
    normalized_prefix: str,
    defined_only_name_keys: set[str],
    all_names: Dict[str, Any],
    sheet_order: Dict[str, int],
) -> Dict[str, tuple[int, int, int]]:
    """ルートキーごとの初出位置 (sheet,row,col) を収集する共通ヘルパー。

    - 生成名は順序に影響させないため、対象は定義名のみ
    - 返り値は root_key -> 最小位置
    """
    root_first_pos: Dict[str, tuple[int, int, int]] = {}
    for name in defined_only_name_keys:
        if not name.startswith(normalized_prefix):
            continue
        defined_name = all_names.get(name)
        if defined_name is None:
            continue
        keys_probe = [k for k in name.removeprefix(normalized_prefix).split(".") if k]
        if not keys_probe:
            continue
        root_key = keys_probe[0]
        pos_probe = compute_top_left_pos(defined_name, sheet_order)
        prev = root_first_pos.get(root_key)
        if prev is None or pos_probe < prev:
            root_first_pos[root_key] = pos_probe
    return root_first_pos
def _upgrade_shape(cur: Optional[str], new_shape: str) -> str:
    """形状を昇格マージ（'2D' が一度でもあれば '2D'）。"""
    return "2D" if (cur == "2D" or new_shape == "2D") else "1D"


def _infer_shape_from_defined_name_destinations(defined_name) -> Optional[str]:
    """DefinedName風オブジェクトの destinations から 1D/2D を推定。"""
    for _sn, coord in getattr(defined_name, "destinations", []) or []:
        coord_clean = coord.replace("$", "") if isinstance(coord, str) else str(coord)
        if ":" not in coord_clean:
            break
        try:
            (sc, sr), (ec, er) = parse_range(coord_clean)
        except Exception:
            break
        rows = abs(er - sr) + 1
        cols = abs(ec - sc) + 1
        if rows > 1 and cols > 1:
            return "2D"
        if (rows == 1 and cols > 1) or (cols == 1 and rows > 1):
            return "1D"
        break
    return None


def iter_pattern1_field_anchors(
    normalized_prefix: str, all_name_keys: List[str], all_names: Dict[str, Any]
) -> List[Tuple[str, str, Any]]:
    """*.field.1（アンカー指定）の候補を列挙し (array, field, dn) を返す。"""
    results: List[Tuple[str, str, Any]] = []
    for nm in all_name_keys:
        if not nm.startswith(normalized_prefix):
            continue
        parts = [p for p in nm.removeprefix(normalized_prefix).split(".") if p]
        if (
            len(parts) >= 4
            and parts[1].isdigit()
            and parts[-1] == "1"
            and not parts[-2].isdigit()
        ):
            array_name = parts[0]
            field_name = parts[-2]
            dn = all_names.get(nm)
            if dn is None:
                continue
            results.append((array_name, field_name, dn))
    return results


def iter_pattern2_field_ranges(
    normalized_prefix: str, all_name_keys: List[str], all_names: Dict[str, Any]
) -> List[Tuple[str, str, Any]]:
    """*.field（末尾 .1 なし）の候補を列挙し (array, field, dn) を返す。"""
    results: List[Tuple[str, str, Any]] = []
    for nm in all_name_keys:
        if not nm.startswith(normalized_prefix):
            continue
        parts = [p for p in nm.removeprefix(normalized_prefix).split(".") if p]
        if len(parts) == 3 and parts[1].isdigit() and not parts[2].isdigit():
            array_name = parts[0]
            field_name = parts[2]
            dn = all_names.get(nm)
            if dn is None:
                continue
            results.append((array_name, field_name, dn))
    return results


def learn_expected_field_shapes(
    normalized_prefix: str,
    all_name_keys: List[str],
    all_names: Dict[str, Any],
) -> Dict[Tuple[str, str], str]:
    """フィールド形状（1D/2D）を学習し、(array_name, field_name)->shape を返す。

    - パターン1: <array>.<idx>.<field>.1 が範囲なら 1D/2D を記録
    - パターン2: <array>.<idx>.<field>（末尾 .1 なし）が範囲なら 1D/2D を記録
      複数回検出された場合は 2D を優先
    """
    expected_field_shape: Dict[Tuple[str, str], str] = {}

    # パターン1: *.field.1
    for array_name, field_name, dn in iter_pattern1_field_anchors(
        normalized_prefix, all_name_keys, all_names
    ):
        shape = _infer_shape_from_defined_name_destinations(dn)
        if shape:
            key = (array_name, field_name)
            prev = expected_field_shape.get(key)
            expected_field_shape[key] = _upgrade_shape(prev, shape) if prev else shape

    # パターン2: *.field（末尾 .1 なし）
    for array_name, field_name, dn in iter_pattern2_field_ranges(
        normalized_prefix, all_name_keys, all_names
    ):
        shape = _infer_shape_from_defined_name_destinations(dn)
        if shape:
            key = (array_name, field_name)
            prev = expected_field_shape.get(key)
            expected_field_shape[key] = _upgrade_shape(prev, shape) if prev else shape

    return expected_field_shape


def find_arrays_with_double_index(
    normalized_prefix: str,
    all_name_keys: List[str],
    gen_map: Optional[Dict[str, Any]],
) -> set[str]:
    """二重数値インデックス（<array>.<i>.<j>.*）が存在する配列名を収集。

    - 定義名と生成名の両方を対象
    """
    arrays: set[str] = set()
    try:
        for nm in all_name_keys:
            if not nm.startswith(normalized_prefix):
                continue
            parts = [p for p in nm.removeprefix(normalized_prefix).split(".") if p]
            if (
                len(parts) >= 4
                and re.fullmatch(r"\d+", parts[1])
                and any(re.fullmatch(r"\d+", x) for x in parts[2:])
            ):
                arrays.add(parts[0])
        if gen_map is not None:
            for nm in gen_map.keys():
                if not nm.startswith(normalized_prefix):
                    continue
                parts = [p for p in nm.removeprefix(normalized_prefix).split(".") if p]
                if (
                    len(parts) >= 4
                    and re.fullmatch(r"\d+", parts[1])
                    and any(re.fullmatch(r"\d+", x) for x in parts[2:])
                ):
                    arrays.add(parts[0])
    except Exception as e:
        logger.debug("locate_found_local_anchor probe cols build failed: %s", e)
    return arrays
def compute_numeric_root_keys(normalized_prefix: str, all_names: Dict[str, Any]) -> set[str]:
    """先頭トークンの直後が数値インデックスとなるルートキー集合を返す。

    例: json.orders.1.date -> {"orders"}
    生成名も all_names に既に統合済みである前提。
    """
    keys: set[str] = set()
    for nm in all_names.keys():
        if not nm.startswith(normalized_prefix):
            continue
        parts = [p for p in nm.removeprefix(normalized_prefix).split(".") if p]
        if len(parts) >= 2 and re.fullmatch(r"\d+", parts[1]):
            keys.add(parts[0])
    return keys


def compute_container_parents_with_children(
    container_parent_names: set[str],
    all_name_keys: List[str],
    gen_map: Optional[Dict[str, Any]],
) -> set[str]:
    """親コンテナ名のうち、子要素（定義名 or 生成名）が1つ以上存在する親を返す。"""
    result: set[str] = set()
    if not container_parent_names:
        return result
    for cpn in container_parent_names:
        prefix_c = cpn + "."
        has_defined_child = any(
            k.startswith(prefix_c) for k in all_name_keys if k != cpn
        )
        has_generated_child = False
        try:
            if gen_map is not None:
                has_generated_child = any(k.startswith(prefix_c) for k in gen_map.keys())
        except Exception:
            has_generated_child = False
        if has_defined_child or has_generated_child:
            result.add(cpn)
    return result


def should_skip_parent_distribution_for_index(
    *,
    array_name: str,
    array_index: int,
    normalized_prefix: str,
    gen_map: Optional[Dict[str, Any]],
) -> bool:
    """親レベル（array.i.*）の値を子へ分配すべきかの抑止判定。should_skip_array_anchor_insertion に委譲。"""
    try:
        return should_skip_array_anchor_insertion(
            array_name, array_index, normalized_prefix, gen_map
        )
    except Exception:
        return False


def compute_group_to_root_map(
    containers: Optional[Dict[str, Any]],
    prefix: str,
    normalized_prefix: str,
    all_name_keys: List[str],
) -> dict[str, str]:
    """コンテナ/定義名から groupLabel -> root の一意マッピングを構築する。

    - 1) コンテナ定義から（例: json.<root>.<group>.1.<child>）
    - 2) 定義名から（例: json.<root>.<group>.1[.<child>...]）
    複数 root で同じ group が現れた場合は曖昧として除外。
    """
    def _extract_group_root_from_container_key(cont_key: str) -> tuple[str, str] | None:
        parts = [p for p in cont_key.split(".") if p]
        if not (
            len(parts) >= 4 and parts[0] == prefix and parts[2] and parts[3].isdigit()
        ):
            return None
        return parts[2], parts[1]  # (group_label, root_name)

    def _extract_group_root_from_name_key(nm: str) -> tuple[str, str] | None:
        if not nm.startswith(normalized_prefix):
            return None
        parts = [p for p in nm.removeprefix(normalized_prefix).split(".") if p]
        if not (len(parts) >= 3 and parts[1] and parts[2].isdigit()):
            return None
        root_name = parts[0]
        group_label = parts[1]
        if re.fullmatch(r"\d+", group_label):
            return None
        return group_label, root_name

    def _update_group_to_root(
        mapping: dict[str, str], ambiguous: set[str], group_label: str, root_name: str
    ) -> None:
        if group_label in ambiguous:
            return
        if group_label in mapping and mapping[group_label] != root_name:
            ambiguous.add(group_label)
            mapping.pop(group_label, None)
            return
        mapping[group_label] = root_name

    group_to_root: dict[str, str] = {}
    ambiguous_groups: set[str] = set()

    if containers:
        for cont_key in containers.keys():
            extracted = _extract_group_root_from_container_key(cont_key)
            if extracted is None:
                continue
            group_label, root_name = extracted
            _update_group_to_root(group_to_root, ambiguous_groups, group_label, root_name)

    for nm in all_name_keys:
        extracted = _extract_group_root_from_name_key(nm)
        if extracted is None:
            continue
        group_label, root_name = extracted
        _update_group_to_root(group_to_root, ambiguous_groups, group_label, root_name)

    return group_to_root


def compute_anchor_names(normalized_prefix: str, all_name_keys: List[str]) -> set[str]:
    """*.1 で終わるアンカー名集合を返す（normalized_prefix を含むフル名）。"""
    anchors: set[str] = set()
    for nm in all_name_keys:
        if not nm.startswith(normalized_prefix):
            continue
        parts = [p for p in nm.removeprefix(normalized_prefix).split(".") if p]
        if parts and parts[-1] == "1":
            anchors.add(nm)
    return anchors


def compute_group_labels_from_anchors(
    anchor_names: set[str],
    containers: Optional[Dict[str, Any]],
    *,
    prefix: str,
    normalized_prefix: str,
) -> set[str]:
    """アンカー名/コンテナからグループラベル候補（lv1 など）を抽出。"""
    labels: set[str] = set()
    # アンカーから抽出
    for nm in anchor_names:
        parts = [p for p in nm.removeprefix(normalized_prefix).split(".") if p]
        # 期待形: <root>.<group>.1[....]
        if len(parts) >= 3 and parts[1] and parts[2] == "1" and not parts[1].isdigit():
            labels.add(parts[1])
    # コンテナからも補助的に抽出
    if containers:
        for cont_key in containers.keys():
            parts = [p for p in cont_key.split(".") if p]
            if (
                len(parts) >= 4
                and parts[0] == prefix
                and parts[2]
                and parts[3].isdigit()
                and not parts[2].isdigit()
            ):
                labels.add(parts[2])
    return labels


def precompute_generated_indices_for_array(
    gen_map: Optional[Dict[str, Any]],
    normalized_prefix: str,
    array_name: str,
) -> set[int]:
    """配列 `array_name` の生成名に含まれるインデックス集合を抽出して返す。"""
    gen_indices: set[int] = set()
    if gen_map is None:
        return gen_indices
    gen_pref_any = f"{normalized_prefix}{array_name}."
    for _gk in gen_map.keys():
        if not _gk.startswith(gen_pref_any):
            continue
        _tail = [p for p in _gk.removeprefix(gen_pref_any).split(".") if p]
        if _tail and _tail[0].isdigit():
            gen_indices.add(int(_tail[0]))
    return gen_indices


def generate_subarray_names_for_field_anchors(wb, normalized_prefix: str) -> None:
    """フィールド直下の .1 が1D範囲を指す場合、2..N の補助生成名をワークブックに登録。

    - 横1xN/縦Nx1 のみ対象（2Dは生成しない）
    - 既存の定義名がある場合は生成をスキップ
    - 生成名はワークブックの _generated_names に登録（既存実装に準拠）
    """
    try:
        for nm, dn in list(getattr(wb, "defined_names", {} ).items()):
            if not nm or not isinstance(nm, str):
                continue
            if not nm.startswith(normalized_prefix):
                continue
            parts = [p for p in nm.removeprefix(normalized_prefix).split(".") if p]
            # 末尾が '1' かつ その前が非数値 → フィールド直下のインデックス（例: json.X.1.A.1）
            if len(parts) < 2 or parts[-1] != "1" or parts[-2].isdigit():
                continue
            # destinations の先頭のみ使用
            sheet_name = None
            coord = None
            try:
                for sn, c in getattr(dn, "destinations", []) or []:
                    sheet_name, coord = sn, c
                    break
            except Exception:
                sheet_name, coord = None, None
            if not sheet_name or not coord:
                continue
            coord_clean = str(coord).replace("$", "")
            if ":" not in coord_clean:
                # 単一セルは対象外
                continue
            try:
                (sc, sr), (ec, er) = parse_range(coord_clean)
            except Exception:
                continue
            rows = abs(er - sr) + 1
            cols = abs(ec - sc) + 1
            # 1次元のみ（横一列 or 縦一列）
            if rows == 1 and cols > 1:
                total = cols
                orient = "row"
            elif cols == 1 and rows > 1:
                total = rows
                orient = "col"
            else:
                # 2D は生成しない
                continue
            base = normalized_prefix + ".".join(parts[:-1])
            for i in range(2, total + 1):
                if orient == "row":
                    ci = sc + (i - 1)
                    ri = sr
                else:
                    ci = sc
                    ri = sr + (i - 1)
                addr = f"{sheet_name}!${get_column_letter(ci)}${ri}"
                gen_name = f"{base}.{i}"
                # 既存定義名があれば生成しない（尊重）
                if gen_name in getattr(wb, "defined_names", {}):
                    continue
                set_generated_name(wb, gen_name, addr)
    except Exception as _e:
        # 生成補助は必須ではないため失敗しても続行
        logger.debug("generate_subarray_names_for_field_anchors skipped due to error: %s", _e)


def compute_excluded_indexed_field_names(
    normalized_prefix: str,
    all_name_keys: List[str],
    all_names: Dict[str, Any],
) -> set[str]:
    """*.field と *.field.1 が競合する際に、抑制すべきインデックス付きフィールド名を算出。

    規則:
    - 末尾が数値で直前が非数値（*.field.1）の名前について、同一ベース（*.field）が存在する場合は抑制対象。
    - ただし、その *.field.1 が複数セルの範囲（1D/2D）を指す場合は抑制しない。
    - 解析不能時は安全側として抑制に倒す（元コード準拠）。
    """
    excluded: set[str] = set()
    base_name_set = set(all_name_keys)
    for nm in all_name_keys:
        try:
            if not nm.startswith(normalized_prefix):
                continue
            parts = [p for p in nm.removeprefix(normalized_prefix).split(".") if p]
            if len(parts) < 2 or not parts[-1].isdigit() or parts[-2].isdigit():
                continue
            base = normalized_prefix + ".".join(parts[:-1])
            if base not in base_name_set:
                continue
            # ただし *.field.1 が複数セル範囲なら抑制しない
            is_multi = False
            try:
                dn = all_names.get(nm)
                if dn is not None:
                    for _sn, coord in getattr(dn, "destinations", []) or []:
                        coord_clean = str(coord).replace("$", "")
                        if ":" in coord_clean:
                            try:
                                (_sc, _sr), (_ec, _er) = parse_range(coord_clean)
                                if _sc != _ec or _sr != _er:
                                    is_multi = True
                            except Exception:
                                logger.debug("failed to parse range for name=%s coord=%s", nm, coord_clean, exc_info=True)
                        break
            except Exception:
                # 解析失敗時は is_multi=False のまま継続
                logger.debug("failed while inspecting defined name destinations: %s", nm, exc_info=True)
            if not is_multi:
                excluded.add(nm)
        except Exception:
            logger.debug("unexpected error while excluding singleton name=%s", nm, exc_info=True)
            excluded.add(nm)
    return excluded

def should_skip_array_anchor_insertion(
    array_name: str,
    array_index0: int,
    normalized_prefix: str,
    gen_map: Optional[Dict[str, Any]],
) -> bool:
    """親レベルの配列アンカー（json.<array>.<index>）の挿入を抑止すべきか判定。

    規則:
    - 同じ配列・同じ index(1-based) の配下にいずれかの生成名（json.<array>.<index>.<something>）が存在する場合、
      親レベルのアンカー挿入は抑止する（生成名側に委譲）。
    - それ以外は抑止しない。
    """
    if not gen_map:
        return False
    idx1 = array_index0 + 1  # 0-based -> 1-based
    gen_pref_any = f"{normalized_prefix}{array_name}."
    for k2 in gen_map.keys():
        if not isinstance(k2, str) or not k2.startswith(gen_pref_any):
            continue
        tail = [p for p in k2.removeprefix(gen_pref_any).split(".") if p]
        # 期待形: <index>.<something> ...（最低2要素）
        if len(tail) >= 2 and tail[0].isdigit() and int(tail[0]) == idx1 and tail[1]:
            return True
    return False


def ensure_array_and_element(
    root_result: JSONDict, array_name: str, array_index0: int
) -> JSONValue:
    """ルート辞書に配列 `array_name` を用意し、`array_index0` のスロットを確保して返す。

    契約:
    - 入力: ルート`root_result`、配列名`array_name`、0始まりインデックス`array_index0`
    - 出力: 現時点の要素（`None`/`dict`/`list`）
    - 例外: なし（必要に応じて拡張する）
    - 副作用: `root_result[array_name]` の確保と長さ拡張
    """
    if array_name not in root_result or not is_json_list(root_result.get(array_name)):
        root_result[array_name] = []
    array_ref_any = root_result[array_name]
    array_ref = cast(List[Any], array_ref_any)
    _ensure_list_index_capacity(array_ref, array_index0)
    return array_ref[array_index0]


def handle_double_numeric_index(
    wb,
    defined_name,
    value,
    array_ref: List[Any],
    array_name: str,
    array_index: int,
    j_index: int,
    rem_keys: List[str],
    name: str,
    original_path_keys: List[str],
    path_keys: List[str],
    user_provided_containers: bool,
    expected_field_shape: Dict[tuple, str],
    safe_insert,
) -> bool:
    """二重数値インデックス（parent.i.j...）の処理を担う。

    契約:
    - 入力: 親配列参照`array_ref`とi/jインデックス、残りキー`rem_keys`などの文脈
    - 出力: 当該エントリを処理したらTrue（呼び出し側でcontinue）、未処理ならFalse
    - 例外: 呼び出し元で捕捉する前提の通常例外（値挿入はsafe_insert経由）
    """
    if _is_case_index_index_field(rem_keys):
        return _handle_case_index_index_field(
            wb, defined_name, value, array_ref, array_name, array_index, j_index, rem_keys, name,
            original_path_keys, path_keys, user_provided_containers, expected_field_shape, safe_insert
        )
    if _is_case_index_index_field_index(rem_keys):
        return _handle_case_index_index_field_index(
            wb, defined_name, value, array_ref, array_name, array_index, j_index, rem_keys, name,
            original_path_keys, path_keys, user_provided_containers, expected_field_shape, safe_insert
        )
    if _is_case_no_rem_keys(rem_keys):
        return _handle_case_no_rem_keys(
            value, array_ref, array_index, j_index
        )
    return _handle_case_fallback(
        FallbackParams(
            value=value,
            array_ref=array_ref,
            array_index=array_index,
            rem_keys=rem_keys,
            name=name,
            original_path_keys=original_path_keys,
            path_keys=path_keys,
            safe_insert=safe_insert,
        )
    )


def _is_case_index_index_field(rem_keys: List[str]) -> bool:
    """ケースA: 残りキーが1つで非数値（parent.i.j.<field>）。"""
    return len(rem_keys) == 1 and (not rem_keys[0].isdigit())


def _is_case_index_index_field_index(rem_keys: List[str]) -> bool:
    """ケースB: 残りキーが2つで field.index（parent.i.j.<field>.<k>）。"""
    return len(rem_keys) == 2 and (not rem_keys[0].isdigit()) and rem_keys[1].isdigit()


def _is_case_no_rem_keys(rem_keys: List[str]) -> bool:
    """ケースC: 残りキーが0（parent.i.j）。"""
    return len(rem_keys) == 0


def _handle_case_index_index_field(
    wb, defined_name, value, array_ref, array_name, array_index, j_index, rem_keys, name,
    original_path_keys, path_keys, user_provided_containers, expected_field_shape, safe_insert
) -> bool:
    field_token2 = rem_keys[0]
    if user_provided_containers:
        if not is_json_list(array_ref[array_index]):
            prev = array_ref[array_index]
            promoted = _promote_element_to_list_if_appropriate(prev)
            if is_json_list(promoted):
                array_ref[array_index] = promoted
            else:
                if not is_json_dict(array_ref[array_index]):
                    array_ref[array_index] = {}
                target_fallback = cast(Dict[str, Any], array_ref[array_index])
                target_fallback[field_token2] = value
                return True
        inner_list = cast(List[Any], array_ref[array_index])
        target_j = _ensure_nested_dict_at(inner_list, j_index)
        try:
            raw = get_named_range_values_preserve_shape(wb, defined_name)
        except Exception:
            raw = value
        try:
            if isinstance(raw, list):
                if raw and isinstance(raw[0], (list, tuple)):
                    expected_field_shape[(array_name, field_token2)] = "2D"
                else:
                    expected_field_shape[(array_name, field_token2)] = "1D"
        except Exception:
            logger.debug("failed to infer expected_field_shape for %s.%s", array_name, field_token2, exc_info=True)
        coerced = apply_expected_shape_to_value(
            raw,
            field_name=field_token2,
            expected_field_shape=expected_field_shape,
            array_name=array_name,
        )
        _set_or_merge_list_field(target_j, field_token2, coerced)
        return True
    else:
        if not is_json_dict(array_ref[array_index]):
            array_ref[array_index] = {}
        target_obj = cast(Dict[str, Any], array_ref[array_index])
        try:
            raw = get_named_range_values_preserve_shape(wb, defined_name)
        except Exception:
            raw = value
        if isinstance(raw, list) and bool(raw) and isinstance(raw[0], (list, tuple)):
            vals = raw
        elif not isinstance(raw, list):
            vals = [raw]
        else:
            vals = raw
        _set_or_merge_list_field(target_obj, field_token2, vals)
        return True

def _handle_case_index_index_field_index(
    wb, defined_name, value, array_ref, array_name, array_index, j_index, rem_keys, name,
    original_path_keys, path_keys, user_provided_containers, expected_field_shape, safe_insert
) -> bool:
    if not user_provided_containers:
        return False
    field_token2 = rem_keys[0]
    if not is_json_list(array_ref[array_index]):
        array_ref[array_index] = cast(List[Any], _promote_element_to_list_if_appropriate(array_ref[array_index]))
    inner_list = cast(List[Any], array_ref[array_index])
    target_j = _ensure_nested_dict_at(inner_list, j_index)
    try:
        subval = get_named_range_values_preserve_shape(wb, defined_name)
    except Exception:
        subval = value
    subval_list = subval if isinstance(subval, list) else [subval]
    _set_or_merge_list_field(target_j, field_token2, subval_list)
    return True

def _handle_case_no_rem_keys(value, array_ref, array_index, j_index) -> bool:
    cur_elem = array_ref[array_index]
    if not is_json_list(cur_elem):
        promoted = _promote_element_to_list_if_appropriate(cur_elem)
        if is_json_list(promoted):
            array_ref[array_index] = cast(List[Any], promoted)
    if is_json_list(array_ref[array_index]):
        inner_list = cast(List[Any], array_ref[array_index])
        _ensure_list_index_capacity(inner_list, j_index)
        inner_list[j_index] = value
    return True

@dataclass(frozen=True)
class FallbackParams:
    """Parameters for case fallback operations.

    Groups the long argument list of _handle_case_fallback to reduce
    call depth and improve maintainability.
    """

    value: Any
    array_ref: List[Any]
    array_index: int
    rem_keys: List[str]
    name: str
    original_path_keys: List[str]
    path_keys: List[str]
    safe_insert: Callable[[Union[Dict[str, Any], List[Any]], List[str], Any, str, str, List[str], List[str]], None]


def _handle_case_fallback(params: FallbackParams) -> bool:
    """Handle case fallback using FallbackParams dataclass."""
    current_element = params.array_ref[params.array_index]
    if not isinstance(current_element, dict):
        params.array_ref[params.array_index] = {}
        current_element = params.array_ref[params.array_index]
    params.safe_insert(
        current_element,
        params.rem_keys,
        params.value,
        ".".join(params.rem_keys),
        params.name,
        params.original_path_keys,
        params.path_keys,
    )
    return True


def should_skip_distribution_index(
    tgt_idx_int: int,
    array_name: str,
    field_token: Optional[str],
    normalized_prefix: str,
    defined_only_name_keys: set[str],
    gen_map: Optional[Dict[str, Any]],
    gen_indices: set[int],
) -> bool:
    """配列分配時に対象インデックスをスキップすべきか判定する。"""
    # 生成名が存在するインデックスは生成名に委譲
    if tgt_idx_int in gen_indices:
        return True
    if field_token:
        tgt_idx_token = str(tgt_idx_int)
        cand1 = f"{normalized_prefix}{array_name}.{tgt_idx_token}.{field_token}"
        cand2 = f"{normalized_prefix}{array_name}.{field_token}.{tgt_idx_token}"
        if cand1 in defined_only_name_keys or cand2 in defined_only_name_keys:
            return True
        if (gen_map is not None) and (cand1 in gen_map or cand2 in gen_map):
            return True
    return False


def distribute_internal_slice(
    array_ref: List[Any],
    array_name: str,
    array_index: int,
    values: List[Any],
    field_token: Optional[str],
    remaining_keys: List[str],
    remaining_path: str,
    name: str,
    original_path_keys: List[str],
    normalized_prefix: str,
    defined_only_name_keys: set[str],
    gen_map: Optional[Dict[str, Any]],
    safe_insert: Callable[
        [
            Union[Dict[str, Any], List[Any]],
            List[str],
            Any,
            str,
            str,
            List[str],
            List[str],
        ],
        None,
    ],
) -> None:
    """内部スライス（list値）を配列要素へ分配する責務を担う。

    ポリシー:
    - 空値はスキップ
    - 既存の非空フィールドは上書きしない
    - 生成名や定義名による保護を尊重し、該当インデックスはスキップ
    """

    # 配列容量と要素の辞書化を保証
    _ensure_array_slots_as_dicts(array_ref, array_index, len(values))

    # 生成名インデックス集合を事前計算
    gen_indices: set[int] = _precompute_distribution_gen_indices(
        gen_map=gen_map, normalized_prefix=normalized_prefix, array_name=array_name
    )

    # 各値を適切なスロットへ挿入
    for j, vj in enumerate(values):
        _apply_distribution_to_index(
            array_ref=array_ref,
            base_index=array_index,
            j=j,
            value=vj,
            field_token=field_token,
            remaining_keys=remaining_keys,
            remaining_path=remaining_path,
            name=name,
            original_path_keys=original_path_keys,
            array_name=array_name,
            normalized_prefix=normalized_prefix,
            defined_only_name_keys=defined_only_name_keys,
            gen_map=gen_map,
            gen_indices=gen_indices,
            safe_insert=safe_insert,
        )


def _ensure_array_slots_as_dicts(array_ref: List[Any], start_index: int, count: int) -> None:
    """`array_ref[start_index:start_index+count]` を辞書スロットとして確保する。"""
    needed = start_index + count
    while len(array_ref) < needed:
        array_ref.append({})
    for _i in range(start_index, start_index + count):
        if not isinstance(array_ref[_i], dict):
            array_ref[_i] = {}


def _precompute_distribution_gen_indices(
    *, gen_map: Optional[Dict[str, Any]], normalized_prefix: str, array_name: str
) -> set[int]:
    """分配スキップ判定用に、生成名が存在する配列インデックスを抽出する。"""
    return precompute_generated_indices_for_array(
        gen_map=gen_map, normalized_prefix=normalized_prefix, array_name=array_name
    )


def _apply_distribution_to_index(
    *,
    array_ref: List[Any],
    base_index: int,
    j: int,
    value: Any,
    field_token: Optional[str],
    remaining_keys: List[str],
    remaining_path: str,
    name: str,
    original_path_keys: List[str],
    array_name: str,
    normalized_prefix: str,
    defined_only_name_keys: set[str],
    gen_map: Optional[Dict[str, Any]],
    gen_indices: set[int],
    safe_insert: Callable[[Union[Dict[str, Any], List[Any]], List[str], Any, str, str, List[str], List[str]], None],
) -> None:
    """1件の内部スライス分配操作（values[j]を適切なスロットへ挿入）を実施。"""
    vj = value
    if vj in (None, ""):
        return
    tgt_idx_int = base_index + 1 + j
    tgt_idx_token = str(tgt_idx_int)
    if should_skip_distribution_index(
        tgt_idx_int=tgt_idx_int,
        array_name=array_name,
        field_token=field_token,
        normalized_prefix=normalized_prefix,
        defined_only_name_keys=defined_only_name_keys,
        gen_map=gen_map,
        gen_indices=gen_indices,
    ):
        logger.debug("DIST-SKIP idx=%s name=%s", tgt_idx_token, name)
        return
    if field_token and len(remaining_keys) == 1:
        cur = array_ref[base_index + j].get(field_token)
        if not is_completely_empty(cur):
            return
    safe_insert(
        array_ref[base_index + j],
        remaining_keys,
        vj,
        remaining_path,
        name,
        original_path_keys,
        original_path_keys,
    )



def handle_internal_slice_distribution_entry(
    *,
    array_ref: List[Any],
    array_name: str,
    array_index: int,
    value: Any,
    remaining_keys: List[str],
    remaining_path: str,
    name: str,
    original_path_keys: List[str],
    normalized_prefix: str,
    arrays_with_double_index: set[str],
    defined_only_name_keys: set[str],
    gen_map: Optional[Dict[str, Any]],
    safe_insert: Callable[[Union[Dict[str, Any], List[Any]], List[str], Any, str, str, List[str], List[str]], None],
) -> bool:
    """
    内部スライス分配の入口。必要に応じて分配し、処理した場合は True を返す。
    Args:
        array_ref: 配列参照
        array_name: 配列名
        array_index: 開始インデックス
        value: 分配する値
        ...existing code...
    Returns:
        bool: 分配した場合 True
    """
    if not isinstance(value, list):
        return False

    # 親レベルの分配抑止（[i][j] の i に生成名があれば親では分配しない）
    if array_name in arrays_with_double_index and should_skip_parent_distribution_for_index(
        array_name=array_name,
        array_index=array_index,
        normalized_prefix=normalized_prefix,
        gen_map=gen_map,
    ):
        logger.debug(
            "PARENT-SKIP distribute %s[%s] due to generated nested children: %s",
            array_name,
            array_index,
            name,
        )
        return True  # 親では何もしないが処理済みとして扱う

    field_token = remaining_keys[0] if len(remaining_keys) == 1 else None
    distribute_internal_slice(
        array_ref=array_ref,
        array_name=array_name,
        array_index=array_index,
        values=value,
        field_token=field_token,
        remaining_keys=remaining_keys,
        remaining_path=remaining_path,
        name=name,
        original_path_keys=original_path_keys,
        normalized_prefix=normalized_prefix,
        defined_only_name_keys=defined_only_name_keys,
        gen_map=gen_map,
        safe_insert=safe_insert,
    )
    return True


def apply_expected_shape_to_value(
    value: Any,
    field_name: Optional[str],
    expected_field_shape: Dict[tuple, str],
    array_name: str,
) -> Any:
    """
    期待形状に基づいて値の形状を正規化する。
    Args:
        value: 入力値
        field_name: フィールド名
        expected_field_shape: 期待形状マップ
        array_name: 配列名
    Returns:
        Any: 正規化後の値
    """
    if field_name is None:
        return value
    try:
        shape = expected_field_shape.get((array_name, field_name))
        if shape == "1D":
            return _coerce_to_1d(value)
        if shape == "2D":
            return _coerce_to_2d(value)
        return value
    except Exception:
        return value


def should_skip_deep_nested_defined_name(name: str, gen_map: Optional[Dict[str, Any]]) -> bool:
    """深いネスト（数値トークンが2つ以上）かつ生成名配下にある定義名はスキップすべきか。

    - name: 例 'json.ツリー1.lv1.1.lv2.1.field' のようなフル名
    - gen_map が None の場合はスキップ判定せず False
    - 同一サブツリー配下に生成名が存在する場合 True を返す
    """
    try:
        if gen_map is None:
            return False
        parts_nm = [p for p in name.split(".") if p]
        if len(parts_nm) < 4:
            return False
        # 数値トークンが2つ以上なら深いネストとみなす
        digit_count = sum(1 for p in parts_nm if p.isdigit())
        if digit_count < 2:
            return False
        prefix_here = name + "."
        # 自身 or 配下に生成名が存在するか
        for k in gen_map.keys():
            if k == name or k.startswith(prefix_here):
                return True
        return False
    except Exception:
        # 判定失敗時は安全側（スキップしない）
        return False


def match_schema_key(key: str, schema_props: dict) -> str:
    """スキーマのプロパティ名に対し、`_` を `.` として扱うワイルドカード照合でキーを正規化。

    - `schema_props` が空なら `key` をそのまま返す
    - ちょうど1件だけ一致した場合に限り置換
    - 複数マッチ時は警告して置き換えない
    """
    if not schema_props:
        return key
    key = key.strip()
    try:
        pattern = "^" + re.escape(key).replace("_", ".") + "$"
        matches = [
            prop for prop in schema_props if re.fullmatch(pattern, prop, flags=re.UNICODE)
        ]
        logger.debug(f"key={key}, pattern={pattern}, matches={matches}")
        if len(matches) == 1:
            return matches[0]
        elif len(matches) > 1:
            logger.warning(
                f"ワイルドカード照合で複数マッチ: '{key}' → {matches}。ユニークでないため置換しません。"
            )
        return key
    except Exception:
        return key


def resolve_path_keys_with_schema(
    *, path_keys: List[str], schema: Optional[Dict[str, Any]]
) -> Tuple[List[str], bool]:
    """スキーマを用いて `path_keys` を解決し、(解決結果, schema_broken) を返す。

    - 数値キーは items に降りる
    - 非数値キーは properties を使って `match_schema_key` で正規化
    - スキーマが途中で追随できなくなった場合は schema_broken=True を返し、呼び出し側で original を使う
    - 例外時は (path_keys, True) を返して安全側に倒す
    """
    if schema is None:
        return (path_keys, True)
    try:
        props = schema.get("properties", {})
        items = schema.get("items", {})
        current_schema = schema
        resolved: List[str] = []
        schema_broken = False
        for k in path_keys:
            if re.fullmatch(r"\d+", k):
                resolved.append(k)
                if isinstance(current_schema, dict) and "items" in current_schema:
                    current_schema = current_schema["items"]
                    props = (
                        current_schema.get("properties", {})
                        if isinstance(current_schema, dict)
                        else {}
                    )
                    items = (
                        current_schema.get("items", {})
                        if isinstance(current_schema, dict)
                        else {}
                    )
                else:
                    props = {}
                    items = {}
            else:
                if not props or not isinstance(props, dict):
                    schema_broken = True
                    break
                new_k = match_schema_key(k, props)
                resolved.append(new_k)
                next_schema = props.get(new_k, {}) if isinstance(props, dict) else {}
                if isinstance(next_schema, dict) and "properties" in next_schema:
                    current_schema = next_schema
                    props = next_schema["properties"]
                    items = next_schema.get("items", {})
                elif isinstance(next_schema, dict) and "items" in next_schema:
                    current_schema = next_schema
                    props = next_schema.get("properties", {})
                    items = next_schema["items"]
                else:
                    props = {}
                    items = {}
        return (resolved, schema_broken)
    except Exception:
        return (path_keys, True)


def finalize_insertion_for_parent_array_element(
    *,
    current_element: Dict[str, Any],
    array_ref: List[Any],
    array_name: str,
    array_index: int,
    value: Any,
    remaining_keys: List[str],
    name: str,
    original_path_keys: List[str],
    normalized_prefix: str,
    arrays_with_double_index: set[str],
    defined_only_name_keys: set[str],
    gen_map: Optional[Dict[str, Any]],
    safe_insert: Callable[[Union[Dict[str, Any], List[Any]], List[str], Any, str, str, List[str], List[str]], None],
    expected_field_shape: Dict[tuple, str],
) -> None:
    """親配列要素に対する残余キーの処理を一括実行する。

    手順:
    - 単一フィールドの場合、期待形状に基づき値を 1D/2D に正規化
    - 値が list の場合、内部スライス分配（必要に応じて抑止）
    - 分配されなかった場合のみ、上書き抑止を考慮しつつ安全挿入
    """
    remaining_path = ".".join(remaining_keys)
    # 期待形状に基づく値の昇格（スカラ → 配列）
    eff_value = value
    if len(remaining_keys) == 1:
        eff_value = apply_expected_shape_to_value(
            eff_value,
            field_name=remaining_keys[0],
            expected_field_shape=expected_field_shape,
            array_name=array_name,
        )

    handled = handle_internal_slice_distribution_entry(
        array_ref=array_ref,
        array_name=array_name,
        array_index=array_index,
        value=eff_value,
        remaining_keys=remaining_keys,
        remaining_path=remaining_path,
        name=name,
        original_path_keys=original_path_keys,
        normalized_prefix=normalized_prefix,
        arrays_with_double_index=arrays_with_double_index,
        defined_only_name_keys=defined_only_name_keys,
        gen_map=gen_map,
        safe_insert=safe_insert,
    )
    if handled:
        return

    # 既存が配列/辞書ならスカラ上書きを避ける
    if should_skip_scalar_overwrite(current_element, remaining_keys):
        return

    safe_insert(
        current_element,
        remaining_keys,
        eff_value,
        remaining_path,
        name,
        original_path_keys,
        original_path_keys,
    )


def should_suppress_value_insertion(
    *,
    name: str,
    keys: List[str],
    normalized_prefix: str,
    all_name_keys: List[str],
    container_parent_names: set[str],
    container_parents_with_children: set[str],
    group_labels: set[str],
) -> bool:
    """エントリ収集段階で値の挿入を抑止すべきかを判定する。

    最小互換仕様:
    - 親コンテナ名で、かつ子要素（定義名 or 生成名）が存在するものは抑止
    - アンカー（末尾が '1'）で、その直下に非数値の子（フィールド）がある場合は抑止
      例: json.A.1.name があるとき json.A.1 自体の値は挿入しない
    - 配列要素直下のラベル終端（....<idx>.<groupLabel>）で、対応アンカー（....<idx>.1）が存在する場合は抑止
    """
    try:
        # 1) 親コンテナ（子がある）
        if name in container_parent_names and name in container_parents_with_children:
            return True

        # 2) アンカー末端 (.1) で非数値の子を持つ場合
        if keys and keys[-1] == "1":
            parent_prefix = name + "."
            for child in all_name_keys:
                if not child.startswith(parent_prefix) or len(child) <= len(parent_prefix):
                    continue
                tail_first = child[len(parent_prefix):].split(".")[0]
                if tail_first and not tail_first.isdigit():
                    return True

        # 3) ラベル終端（....<idx>.<groupLabel>）で対応アンカーがある場合
        if len(keys) >= 3 and keys[-1] in group_labels and keys[-2].isdigit():
            # 例: json.<root>.<idx>.<label> に対して json.<root>.<idx>.1 またはその子があれば抑止
            base = f"{normalized_prefix}{keys[0]}.{keys[1]}.1"
            for k in all_name_keys:
                if k == base or k.startswith(base + "."):
                    return True

        return False
    except Exception:
        return False


def collect_entries_in_sheet_order(
    *,
    all_names: Dict[str, Any],
    normalized_prefix: str,
    excluded_indexed_field_names: set[str],
    sheet_order: Dict[str, int],
    suppress_ctx: Dict[str, Any],
) -> List[Tuple[tuple[int, int, int], str, Any, List[str]]]:
    """定義名を走査し、抑制ロジックを適用して Excel 読取順のエントリ配列を返す。

    suppress_ctx には抑制判定に必要な文脈を渡す:
    {
      'all_name_keys', 'container_parent_names', 'container_parents_with_children', 'group_labels'
    }
    """
    entries: List[Tuple[tuple[int, int, int], str, Any, List[str]]] = []
    all_name_keys = suppress_ctx.get("all_name_keys", [])
    container_parent_names = suppress_ctx.get("container_parent_names", set())
    container_parents_with_children = suppress_ctx.get(
        "container_parents_with_children", set()
    )
    group_labels = suppress_ctx.get("group_labels", set())

    # ルート最初出現位置（同一ルート内の安定化に使用）
    root_first_pos: Dict[str, tuple[int, int, int]] = suppress_ctx.get(
        "root_first_pos", {}
    )

    for name, defined_name in all_names.items():
        if not name.startswith(normalized_prefix):
            continue
        if name in excluded_indexed_field_names:
            continue
        keys = [k for k in name.removeprefix(normalized_prefix).split(".") if k]
        if should_suppress_value_insertion(
            name=name,
            keys=keys,
            normalized_prefix=normalized_prefix,
            all_name_keys=all_name_keys,
            container_parent_names=container_parent_names,
            container_parents_with_children=container_parents_with_children,
            group_labels=group_labels,
        ):
            continue
        pos_key = compute_top_left_pos(defined_name, sheet_order)
        entries.append((pos_key, name, defined_name, keys))

    def _entry_sort_key(x: Tuple[tuple[int, int, int], str, Any, List[str]]):
        pos, nm, _dn, ks = x
        root = ks[0] if ks else ""
        root_pos = root_first_pos.get(root, (10**9, 10**9, 10**9))
        return (root_pos[0], root_pos[1], root_pos[2], pos[0], pos[1], pos[2], nm)

    entries.sort(key=_entry_sort_key)
    return entries


def reorder_roots_by_sheet_order(
    result: Dict[str, Any],
    root_first_pos: Dict[str, tuple[int, int, int]] | None,
    prefix: str,
    user_provided_containers: bool,
) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    """Excel上の出現順に基づいてルートキー順を安定化する。

    返値は (result, root_result)。例外時は無変更で返す。
    """
    try:
        if is_json_dict(result):
            target_dict: Optional[Dict[str, Any]] = None
            if user_provided_containers and prefix in result and is_json_dict(result[prefix]):
                target_dict = cast(Dict[str, Any], result[prefix])
            else:
                target_dict = result
            if target_dict is not None and isinstance(root_first_pos, dict) and root_first_pos:
                desired_roots = [k for k, _ in sorted(root_first_pos.items(), key=lambda kv: kv[1])]
                existing_keys = list(target_dict.keys())
                new_order: Dict[str, Any] = {}
                for rk in desired_roots:
                    if rk in target_dict:
                        new_order[rk] = target_dict[rk]
                for k in existing_keys:
                    if k not in new_order:
                        new_order[k] = target_dict[k]
                if user_provided_containers and prefix in result and is_json_dict(result[prefix]):
                    result[prefix] = new_order
                else:
                    result = new_order
                # 外側（トップレベル）もルート順に並び替え
                if is_json_dict(result):
                    outer_existing = list(result.keys())
                    outer_new: Dict[str, Any] = {}
                    if prefix in result and prefix not in outer_new:
                        outer_new[prefix] = result[prefix]
                    for rk in desired_roots:
                        if rk in result and rk not in outer_new:
                            outer_new[rk] = result[rk]
                    for k in outer_existing:
                        if k not in outer_new:
                            outer_new[k] = result[k]
                    result = outer_new
        root_result = result if not user_provided_containers else result.get(prefix, {})
        return result, cast(Dict[str, Any], root_result)
    except Exception:
        root_result = result if not user_provided_containers else result.get(prefix, {})
        return result, cast(Dict[str, Any], root_result)


def reshape_containers_in_result(
    result: Dict[str, Any],
    containers: Optional[Dict[str, Any]],
    prefix: str,
    user_provided_containers: bool,
) -> Dict[str, Any]:
    """コンテナ出力（dict of lists）を list of dicts へリシェイプする。"""

    def _reshape_container(value: Any) -> Any:
        if is_json_dict(value) and value:
            list_keys = [k for k, v in value.items() if is_json_list(v)]
            if list_keys:
                max_len = 0
                for lk in list_keys:
                    arrv = value.get(lk)
                    if is_json_list(arrv):
                        max_len = max(max_len, len(cast(List[Any], arrv)))
                rows: List[Dict[str, Any]] = []
                for i in range(max_len):
                    row: Dict[str, Any] = {}
                    for col, arr in value.items():
                        if is_json_list(arr) and i < len(arr):
                            row[col] = arr[i]
                    if row:
                        rows.append(row)
                return rows if rows else value
        return value

    if user_provided_containers and prefix in result and is_json_dict(result[prefix]):
        for cont_key in (containers or {}).keys():
            if cont_key.startswith(prefix + ".") and cont_key.count(".") == 1:
                base_name = cont_key.split(".", 1)[1]
                if base_name in result[prefix]:
                    result[prefix][base_name] = _reshape_container(result[prefix][base_name])
    return result


def replicate_prefix_children_to_top_level(
    result: Dict[str, Any],
    prefix: str,
    group_labels: set[str],
    root_first_pos: Dict[str, tuple[int, int, int]] | None,
) -> Dict[str, Any]:
    """互換性維持のため、prefix 配下のキーをトップレベルへ複製し、ルート順で整列する。"""
    if not (is_json_dict(result) and prefix in result):
        return result
    rep_group_labels: set[str] = {g for g in group_labels if re.fullmatch(r"lv\d+", g)}
    pref_val2 = result[prefix]
    if not is_json_dict(pref_val2):
        return result
    for k, v in cast(Dict[str, JSONValue], pref_val2).items():
        if k in rep_group_labels:
            continue
        if k not in result:
            result[k] = v

    if isinstance(root_first_pos, dict) and root_first_pos:
        desired_roots = [k for k, _ in sorted(root_first_pos.items(), key=lambda kv: kv[1])]
        outer_existing = list(result.keys())
        outer_new: Dict[str, Any] = {}
        if prefix in result and prefix not in outer_new:
            outer_new[prefix] = result[prefix]
        for rk in desired_roots:
            if rk in result and rk not in outer_new:
                outer_new[rk] = result[rk]
        for k in outer_existing:
            if k not in outer_new:
                outer_new[k] = result[k]
        result = outer_new
    return result


def fallback_normalize_root_groups_without_containers(
    result: Dict[str, Any],
    group_to_root: Dict[str, str],
    gen_map: Optional[Dict[str, Any]],
    normalized_prefix: str,
) -> Dict[str, Any]:
    """コンテナ未指定時のフォールバック正規化を適用する。"""
    if not is_json_dict(result):
        return result
    _absorb_root_groups(result, group_to_root)
    _remove_leading_empty_elements(result)
    if gen_map is not None:
        _reconstruct_arrays_from_generated_names(result, gen_map, normalized_prefix)
    return result

def _absorb_root_groups(result: Dict[str, Any], group_to_root: Dict[str, str]) -> None:
    for grp, root_name in list(group_to_root.items()):
        if root_name in result and is_json_dict(result[root_name]):
            continue
        if root_name not in result and grp in result and is_json_list(result[grp]):
            result[root_name] = {grp: result[grp]}
            result.pop(grp, None)

def _remove_leading_empty_elements(result: Dict[str, Any]) -> None:
    def _is_effectively_empty_dict(d: Any) -> bool:
        if not is_json_dict(d):
            return False
        if not d:
            return True
        try:
            return all(DataCleaner.is_empty_value(x) for x in d.values())
        except Exception:
            return False
    for k, v in list(result.items()):
        if is_json_list(v) and v:
            while v and _is_effectively_empty_dict(v[0]):
                v.pop(0)

def _reconstruct_arrays_from_generated_names(result: Dict[str, Any], gen_map: Dict[str, Any], normalized_prefix: str) -> None:
    gkeys: dict = gen_map
    for arr_name, arr_val in list(result.items()):
        if not is_json_list(arr_val):
            continue
        gen_pref = f"{normalized_prefix}{arr_name}."
        has_any = any(k.startswith(gen_pref) for k in gkeys.keys())
        if not has_any:
            continue
        try:
            if any((it is not None) and (not is_json_dict(it)) for it in arr_val):
                continue
        except Exception:
            continue
        idx_map: Dict[int, Dict[str, Any]] = {}
        for gk, gv in gkeys.items():
            if not gk.startswith(gen_pref):
                continue
            tail = [p for p in gk[len(gen_pref) :].split(".") if p]
            if len(tail) < 2:
                continue
            if not tail[0].isdigit():
                continue
            idx = int(tail[0])
            field = tail[1]
            if not field or field.isdigit():
                continue
            idx_map.setdefault(idx, {})[field] = gv
        if not idx_map:
            continue
        max_idx = max(idx_map.keys())
        base_list: List[Dict[str, Any]] = []
        for elem in arr_val:
            base_list.append(dict(elem) if is_json_dict(elem) else {})
        while len(base_list) < max_idx:
            base_list.append({})
        for ix, fields in idx_map.items():
            dst = base_list[ix - 1]
            for fk, fv in fields.items():
                if fv in (None, ""):
                    continue
                if (
                    fk in dst
                    and isinstance(dst.get(fk), (list, dict))
                    and not is_completely_empty(dst.get(fk))
                ):
                    continue
                dst[fk] = fv
        while base_list and all(v in (None, "") for v in base_list[0].values()):
            base_list.pop(0)
        result[arr_name] = base_list


def get_value_for_defined_or_generated_name(
    *, wb: Any, name: str, defined_name: Any, gen_map: Optional[Dict[str, Any]]
) -> Tuple[bool, Any]:
    """生成名を優先して値を取得し、深いネストの定義名は必要に応じてスキップする。

    戻り値: (should_skip, value)
    - should_skip=True の場合は呼び出し側で continue する
    """
    if gen_map is not None and name in gen_map:
        logger.debug("コンテナ生成セル名の値を直接取得: %s -> %r", name, gen_map[name])
        return (False, gen_map[name])
    # 深いネストかつ生成名配下はスキップ
    if should_skip_deep_nested_defined_name(name, gen_map):
        logger.debug("生成名配下の深いネスト定義名をスキップ: %s", name)
        return (True, None)
    try:
        return (False, get_named_range_values(wb, defined_name))
    except Exception as e:
        # #REF! 等で有効な destination が無い場合は、値は取得できないが
        # 形状復元ロジックで null/[]/{} を出力できるよう、空プレースホルダを返す
        logger.debug("Use empty placeholder for invalid destinations: %s (err=%s)", name, e)
        return (False, "")


def process_array_path_entry(
    *,
    wb: Any,
    defined_name: Any,
    value: Any,
    path_keys: List[str],
    name: str,
    original_path_keys: List[str],
    normalized_prefix: str,
    root_result: Dict[str, Any],
    arrays_with_double_index: set[str],
    expected_field_shape: Dict[Tuple[str, str], str],
    gen_map: Optional[Dict[str, Any]],
    group_labels: set[str],
    all_name_keys: List[str],
    container_parent_names: set[str],
    defined_only_name_keys: set[str],
    safe_insert: Callable[[Union[Dict[str, Any], List[Any]], List[str], Any, str, str, List[str], List[str]], None],
    user_provided_containers: bool = False,
) -> bool:
    """配列パス（array.i.*）の処理を行い、処理済みなら True を返す。"""
    if len(path_keys) < 2 or not re.fullmatch(r"\d+", path_keys[1]):
        return False

    array_name = path_keys[0]
    array_index = int(path_keys[1]) - 1

    # アンカー自体の挿入抑止（同 index に生成名がある場合）
    anchor_skip = False
    if len(path_keys) == 2:
        anchor_skip = should_skip_array_anchor_insertion(
            array_name, array_index, normalized_prefix, gen_map
        )

    if anchor_skip:
        return True

    # 配列準備＋要素参照
    current_element = ensure_array_and_element(root_result, array_name, array_index)
    array_ref = root_result[array_name]

    # 二重数値インデックス: array.i.j...
    if len(path_keys) >= 3 and re.fullmatch(r"\d+", path_keys[2]):
        j_index = int(path_keys[2]) - 1
        rem_keys = path_keys[3:] if len(path_keys) > 3 else []
        handled = handle_double_numeric_index(
            wb,
            defined_name,
            value,
            array_ref,
            array_name,
            array_index,
            j_index,
            rem_keys,
            name,
            original_path_keys,
            path_keys,
            user_provided_containers,
            expected_field_shape,
            safe_insert,
        )
        return True if handled else False

    # 1次元配列: array.i.field...
    if (not isinstance(current_element, dict)) and (array_name not in arrays_with_double_index):
        array_ref[array_index] = {}
        current_element = array_ref[array_index]

    if array_name in arrays_with_double_index:
        handled_parent = handle_parent_level_for_double_index_array(
            wb=wb,
            defined_name=defined_name,
            value=value,
            array_ref=array_ref,
            array_name=array_name,
            array_index=array_index,
            path_keys=path_keys,
            name=name,
            normalized_prefix=normalized_prefix,
            gen_map=gen_map,
            expected_field_shape=expected_field_shape,
        )
        if handled_parent:
            return True

    if len(path_keys) > 2:
        # 親レベル処理で未処理の場合、以降は current_element にフィールドを入れるため dict を保証する
        if not isinstance(current_element, dict):
            array_ref[array_index] = {}
            current_element = array_ref[array_index]
        remaining_keys = path_keys[2:]
        # 末端が field.index の場合: 値を形状保持で取得し、1Dで追記
        if (
            len(remaining_keys) == 2
            and not remaining_keys[0].isdigit()
            and remaining_keys[1].isdigit()
        ):
            field_token = remaining_keys[0]
            try:
                subval = get_named_range_values_preserve_shape(wb, defined_name)
            except Exception:
                subval = value
            subval_list = _coerce_to_1d(subval)
            _set_or_merge_list_field(cast(Dict[str, Any], current_element), field_token, subval_list)
            return True
        # ラベル終端の抑制
        if suppress_label_terminal_if_applicable(
            remaining_keys=remaining_keys,
            original_path_keys=original_path_keys,
            group_labels=group_labels,
            normalized_prefix=normalized_prefix,
            all_name_keys=all_name_keys,
            container_parent_names=container_parent_names,
        ):
            return True

        finalize_insertion_for_parent_array_element(
            current_element=cast(Dict[str, Any], current_element),
            array_ref=array_ref,
            array_name=array_name,
            array_index=array_index,
            value=value,
            remaining_keys=remaining_keys,
            name=name,
            original_path_keys=original_path_keys,
            normalized_prefix=normalized_prefix,
            arrays_with_double_index=arrays_with_double_index,
            defined_only_name_keys=defined_only_name_keys,
            gen_map=gen_map,
            safe_insert=safe_insert,
            expected_field_shape=expected_field_shape,
        )
        return True
    else:
        array_ref[array_index] = value
        return True


def should_skip_scalar_overwrite(current_element: Any, remaining_keys: List[str]) -> bool:
    """既に配列/辞書が入っているフィールドにスカラを上書きしない判定。"""
    return (
        len(remaining_keys) == 1
        and isinstance(current_element, dict)
        and remaining_keys[0] in current_element
        and isinstance(current_element[remaining_keys[0]], (list, dict))
        and not is_completely_empty(current_element[remaining_keys[0]])
    )


def _normalize_root_group_for_parse(
    keys: List[str],
    *,
    group_to_root: Dict[str, str],
    numeric_root_keys: set[str],
    root_result: Dict[str, Any],
) -> List[str]:
    """parse_named_ranges_with_prefix 内でのルート配下キー正規化を関数化。

    - group_to_root: グループ名→ルート名のマップ
    - numeric_root_keys: 先頭が配列ルートである既知キー集合
    - root_result: 既存のルート辞書（配列存在チェックに使用）
    """
    k = list(keys)
    if len(k) >= 2 and (k[0] in group_to_root) and re.fullmatch(r"\d+", k[1]):
        k = [group_to_root[k[0]]] + k[1:]
    while len(k) >= 2:
        first, second = k[0], k[1]
        first_is_array_root = (first in numeric_root_keys) or (
            first in root_result and isinstance(root_result.get(first), list)
        )
        if first_is_array_root and not re.fullmatch(r"\d+", second):
            k = k[1:]
            continue
        break
    return k


def merge_into_list_unique(existing: Any, values: Any) -> List[Any]:
    """
    リストへのユニーク追記ヘルパー。
    Args:
        existing: 既存値
        values: 追加値
    Returns:
        List[Any]: 1Dリスト（ユニーク追記済み）
    """
    # 既存をベースのリストに正規化
    base: List[Any]
    if isinstance(existing, list):
        base = list(existing)
    elif existing in (None, ""):
        base = []
    else:
        base = [existing]

    is_list_of_list = bool(base) and isinstance(base[0], (list, tuple))
    if is_list_of_list:
        return base

    cand_iter = values if isinstance(values, list) else [values]
    for v in cand_iter:
        if v in (None, ""):
            continue
        if v not in base:
            base.append(v)
    return base


def _ensure_nested_dict_at(lst: List[Any], idx0: int) -> Dict[str, Any]:
    """リスト `lst[idx0]` に辞書を用意して返す（None/非辞書は辞書に置換）。

    - `lst` の長さが不足していれば `None` 埋めで拡張
    - 既存が `None` もしくは非 dict の場合は `{}` に置換
    - 最終的に `dict` を返す
    """
    _ensure_list_index_capacity(lst, idx0)
    if lst[idx0] is None or not is_json_dict(lst[idx0]):
        lst[idx0] = {}
    return cast(Dict[str, Any], lst[idx0])


def _set_or_merge_list_field(
    target: Dict[str, Any], field: str, new_values: Any
) -> None:
    """target[field] に対して 1D ユニーク追記（2Dはそのまま）で設定するユーティリティ。

    - 既存が None/空相当ならそのまま設定
    - 既存があれば `merge_into_list_unique` で 1D ユニークマージ（例外時は new_values をそのまま代入）
    - `new_values` は呼び出し側で 1D/2D へ事前正規化しておくこと（本関数は形状変換をしない）
    """
    existing = target.get(field)
    if existing is None or is_completely_empty(existing):
        target[field] = new_values
    else:
        try:
            target[field] = merge_into_list_unique(existing, new_values)
        except Exception:
            target[field] = new_values
    # New function to promote elements to list if appropriate
def _promote_element_to_list_if_appropriate(elem: Any) -> Union[List[Any], Any]:
    """空辞書/None 相当の要素は配列に昇格し、それ以外は据え置く。

    - `{}` または `None`/空辞書相当なら `[]` を返す
    - 既に list ならそのまま返す
    - 非空の辞書などはそのまま返す（後方互換のため昇格しない）
    """
    if is_json_list(elem):
        return elem
    if is_json_dict(elem) and len(elem) == 0:
        return []
    if elem in (None, {}):
        return []
    return elem

def parse_range(range_str: str) -> tuple:
    """
    Excel範囲文字列を解析して開始座標と終了座標を返す
    例: "B2:D4" -> ((2, 2), (4, 4))
    例: "$B$2:$D$4" -> ((2, 2), (4, 4))
    """
    # $記号を削除してから解析
    cleaned_range = range_str.replace("$", "")

    match = re.match(r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$", cleaned_range)
    if not match:
        raise ValueError(f"無効な範囲形式: {range_str}")

    start_col, start_row, end_col, end_row = match.groups()
    start_coord = (column_index_from_string(start_col), int(start_row))
    end_coord = (column_index_from_string(end_col), int(end_row))

    return start_coord, end_coord


def detect_instance_count(start_coord: tuple, end_coord: tuple, direction: str) -> int:
    """
    範囲とdirectionから、インスタンス数を検出
    """
    start_col, start_row = start_coord
    end_col, end_row = end_coord

    dir_lc = (direction or "").lower()
    if dir_lc == "row":
        return end_row - start_row + 1
    elif dir_lc == "column":
        return end_col - start_col + 1
    else:
        raise ValueError(f"無効なdirection: {direction}")


def generate_cell_names(
    container_name: str,
    start_coord: tuple,
    end_coord: tuple,
    direction: str,
    items: list,
) -> list:
    """
    コンテナ用のセル名を生成（一般ユーティリティ）
    1-base indexingを使用
    """
    instance_count = detect_instance_count(start_coord, end_coord, direction)
    cell_names = []
    for i in range(1, instance_count + 1):  # 1-base indexing
        for item in items:
            cell_name = f"{container_name}_{i}_{item}"
            cell_names.append(cell_name)
    return cell_names


def load_container_config(config_path: Path) -> Dict[str, Any]:
    """
    config.yaml/config.jsonからコンテナ設定を読み込む
    YAMLフォーマットを優先し、JSONもサポートする（JSONはYAMLのサブセット）
    注意：JSON Schemaファイルは従来通りJSON形式のみサポート
    """
    if not config_path.exists():
        return {}

    try:
        with config_path.open("r", encoding="utf-8") as f:
            # YAMLとして読み込み（JSONはYAMLのサブセットなので自動対応）
            config = yaml.safe_load(f)
            if config is None:
                return {}
            return config.get("containers", {})
    except yaml.YAMLError as e:
        logger.warning(
            f"設定ファイルの読み込みに失敗（YAML解析エラー）: {config_path} - {e}"
        )
        return {}
    except FileNotFoundError:
        logger.warning(f"設定ファイルが見つかりません: {config_path}")
        return {}
    except Exception as e:
        logger.warning(f"設定ファイルの読み込みに失敗: {config_path} - {e}")
        return {}


def infer_containers_from_named_ranges(
    workbook, prefix: str
) -> Dict[str, Dict[str, int | str]]:
    """
    コンテナ未指定時の自動推論:
    - 親セル名（json.X または json.X.Y...）がセル『範囲』を指し、かつその下位階層に他のセル名が存在する場合、
      その親セル名をコンテナ定義として解釈する。
    - 直下の子トークンが数値なら繰り返し（increment=範囲高さ, direction=row）、非数値なら非繰り返し（increment=0, direction=row）。

    注意:
    - 複数シート/複数宛先の範囲がある場合は最初の宛先から高さを決定（仕様不明点。必要なら拡張可能）。
    - 親セル名自体が単一セルの場合は対象外。
    """
    try:
        if not prefix:
            return {}
        prefix_dot = prefix + "."
        # すべての名前を収集
        all_names = [name for name in workbook.defined_names.keys() if name]
        # 範囲高さの収集
        range_heights = _collect_range_heights_for_prefix(workbook, prefix_dot, all_names)
        if not range_heights:
            return {}
        # 繰返し性の仮決定
        preliminary = _determine_preliminary_container_behaviors(range_heights, all_names)
        if not preliminary:
            return {}
        # 抑制ルールの適用
        suppressed = _suppress_redundant_candidates(preliminary, all_names)
        return suppressed
    except Exception as e:
        logger.debug(f"自動コンテナ推論中にエラー: {e}")
        return {}


def _collect_range_heights_for_prefix(workbook, prefix_dot: str, all_names: list[str]) -> Dict[str, int]:
    range_heights: Dict[str, int] = {}
    for name in all_names:
        if not name.startswith(prefix_dot):
            continue
        dn = workbook.defined_names[name]
        height_detected = None
        for _sn, coord in dn.destinations:
            coord_clean = coord.replace("$", "")
            if ":" not in coord_clean:
                continue
            (_sc, _sr), (_ec, _er) = parse_range(coord_clean)
            height_detected = abs(_er - _sr) + 1
            break
        if height_detected is not None:
            range_heights[name] = height_detected
    return range_heights


def _determine_preliminary_container_behaviors(
    range_heights: Dict[str, int], all_names: list[str]
) -> Dict[str, Dict[str, int | str]]:
    preliminary: Dict[str, Dict[str, int | str]] = {}
    for nm, height in range_heights.items():
        if nm.endswith(".1"):
            preliminary[nm] = {"direction": "row", "increment": int(height)}
            continue
        parent_prefix = nm + "."
        child_tokens: set[str] = set()
        for child in all_names:
            if child.startswith(parent_prefix) and len(child) > len(parent_prefix):
                tail = child[len(parent_prefix) :]
                first = tail.split(".")[0]
                if first:
                    child_tokens.add(first)
        if not child_tokens:
            continue
        is_repeating = any(tok.isdigit() for tok in child_tokens)
        preliminary[nm] = {"direction": "row", "increment": 1 if is_repeating else 0}
    return preliminary


def _suppress_redundant_candidates(
    preliminary: Dict[str, Dict[str, int | str]], all_names: list[str]
) -> Dict[str, Dict[str, int | str]]:
    repeating_parents = [p for p, cfg in preliminary.items() if int(cfg.get("increment", 0) or 0) > 0]
    suppressed: Dict[str, Dict[str, int | str]] = {}
    for nm, cfg in preliminary.items():
        nm_parts = [t for t in nm.split(".") if t]
        suppressed_here = False
        for rp in repeating_parents:
            if nm == rp or not nm.startswith(rp + "."):
                continue
            rp_parts = [t for t in rp.split(".") if t]
            if len(nm_parts) == len(rp_parts) + 1 and nm_parts[-1] == "1":
                if not rp.endswith(".1"):
                    suppressed_here = True
                    break
                has_non_numeric_child = False
                nm_prefix = nm + "."
                for child in all_names:
                    if child.startswith(nm_prefix) and len(child) > len(nm_prefix):
                        tail = child[len(nm_prefix) :]
                        first = tail.split(".")[0]
                        if first and (not first.isdigit()):
                            has_non_numeric_child = True
                            break
                if not has_non_numeric_child:
                    suppressed_here = True
                    break
            if len(nm_parts) > len(rp_parts) and not nm_parts[-1].isdigit():
                suppressed_here = True
                break
        if suppressed_here:
            continue
        suppressed[nm] = cfg
    return suppressed


def resolve_container_range(wb, range_spec: str) -> tuple:
    """
    範囲指定を解決して座標を返す
    range_specは名前付き範囲名または範囲文字列（"A1:C10"）
    """
    # 名前付き範囲として試行
    if range_spec in wb.defined_names:
        defined_name = wb.defined_names[range_spec]
        for sheet_name, coord in defined_name.destinations:
            return parse_range(coord)

    # 範囲文字列として解析
    try:
        return parse_range(range_spec)
    except ValueError:
        # 互換: エラーメッセージを統一
        raise ValueError("無効な範囲指定")


class DataCleaner:
    """
    データのクリーニングと検証を行うユーティリティ。

    注意: 実装はモジュールレベル関数に集約されています（重複排除のため）。
    既存コードの互換性維持のため、このクラスの各メソッドは対応する
    モジュール関数に委譲します。
    """

    @staticmethod
    def is_empty_value(value: Any) -> bool:
        """値が空かどうかを判定する（委譲）。"""
        return is_empty_value(value)


# =============================================================================
# JSON Path Operations
# =============================================================================


def parse_json_path(path: str) -> List[str]:
    """
    JSON パス文字列をキーのリストに解析する

    例:
        "data.items[0].value" -> ["data", "items", "0", "value"]
        "users[1].profile.name" -> ["users", "1", "profile", "name"]
    """
    if not path:
        return []

    # 配列アクセス記法 [n] を .n に変換
    path = re.sub(r"\[(\d+)\]", r".\1", path)

    # ドットで分割してキーのリストを作成
    keys = [key for key in path.split(".") if key]

    return keys


@dataclass(frozen=True)
class OutputOrderingPolicy:
    """出力順制御のポリシー（内部拡張ポイント）。

    方針:
    - 通常のユースケースは固定仕様でカバーし、CLI からの細粒度な切替は想定しません。
    - スキーマによる properties 順適用は最終出力直前のみ（読取順の決定性を維持）。
    - スキーマ外プロパティは挿入順を保持し、名称に依存する特例は導入しません。

    パラメータ:
    - schema_first: スキーマの properties 順を優先（True のとき）。
    - align_sibling_list_of_dicts: 同一構造の兄弟 list-of-dicts でヘッドキー順を共有する。
    - keep_extras_in_insertion_order: スキーマ外プロパティは挿入順を保持。
    """

    schema_first: bool = False
    align_sibling_list_of_dicts: bool = True
    keep_extras_in_insertion_order: bool = True


def _align_key_order_by_head(obj: Any) -> Any:
    """list-of-dicts に対してヘッドのキー順を兄弟間で共有する整形。

    - 同一構造パス（数値トークン除外）で最初に観測したキー順を共有して適用。
    - list-of-list-of-dicts も内側に対して同様の整列を行う。
    - 非破壊の新オブジェクトを返す。
    """
    shared_head_keys_by_path: dict[tuple[str, ...], list[str]] = {}

    def _apply_keys_order(d: Dict[str, Any], keys: list[str], path: list[str]) -> Dict[str, Any]:
        rest_keys = [k for k in d.keys() if k not in keys]
        new_d: Dict[str, Any] = {}
        for k in keys:
            if k in d:
                new_d[k] = _recur(d[k], path + [k])
        for k in rest_keys:
            new_d[k] = _recur(d[k], path + [k])
        return new_d

    def _handle_lod_list(x: list[Any], path: list[str]) -> list[Any] | None:
        """list-of-dicts の場合に共有ヘッドキー順を適用。該当しない場合は None。"""
        if not x or not all(is_json_dict(it) for it in x if it is not None):
            return None
        sanitized_path = [p for p in path if not (isinstance(p, str) and p.isdigit())]
        path_key = tuple(sanitized_path)
        if path_key in shared_head_keys_by_path:
            head_keys = shared_head_keys_by_path[path_key]
        else:
            head = next((it for it in x if is_json_dict(it)), None)
            head_keys = list(head.keys()) if head is not None else []
            shared_head_keys_by_path[path_key] = head_keys
        aligned_list: list[Any] = []
        for it in x:
            if not is_json_dict(it):
                aligned_list.append(_recur(it, path))
            else:
                aligned_list.append(_apply_keys_order(it, head_keys, path))
        return aligned_list

    def _handle_lolod_list(x: list[Any], path: list[str]) -> list[Any] | None:
        """list-of-list-of-dicts の場合に内側へヘッドキー順を共有適用。該当しない場合は None。"""
        if not x or not all((not it) or is_json_list(it) for it in x):
            return None
        common_head_keys: list[str] | None = None
        for it in x:
            if not (is_json_list(it) and it):
                continue
            if all(is_json_dict(xx) for xx in it if xx is not None):
                head_candidate = next((xx for xx in it if is_json_dict(xx)), None)
                if head_candidate is not None:
                    common_head_keys = list(head_candidate.keys())
                    break
        if common_head_keys is None:
            return None
        aligned: list[Any] = []
        for it in x:
            if is_json_list(it) and all(is_json_dict(xx) for xx in it if xx is not None):
                sub_aligned: list[Any] = []
                for d in it:
                    if not is_json_dict(d):
                        sub_aligned.append(_recur(d, path))
                    else:
                        sub_aligned.append(_apply_keys_order(d, common_head_keys, path))
                aligned.append(sub_aligned)
            else:
                aligned.append(_recur(it, path))
        return aligned

    def _recur(x: Any, path: list[str]) -> Any:
        if is_json_list(x):
            return (
                _handle_lolod_list(x, path)
                or _handle_lod_list(x, path)
                or [_recur(it, path) for it in x]
            )
        if is_json_dict(x):
            return {k: _recur(v, path + [k]) for k, v in x.items()}
        return x

    return _recur(obj, [])


def order_for_output(data: Dict[str, Any], *, policy: OutputOrderingPolicy, schema: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    """最終出力用の順序整形を適用して返す。

    優先順位:
    1) schema_first=True かつ schema がある → スキーマ順（余剰は挿入順のまま）
    2) align_sibling_list_of_dicts=True → 兄弟 list-of-dicts のキー順共有
    3) keep_extras_in_insertion_order は reorder_json 実装が担保
    """
    out = data
    # まず sibling alignment（表示一貫性）
    if policy.align_sibling_list_of_dicts:
        out = cast(Dict[str, Any], _align_key_order_by_head(out))
    # スキーマ順（必要時）
    if policy.schema_first and schema:
        out = cast(Dict[str, Any], reorder_json(out, schema))
    return out


# =============================================================================
# Cross-cutting policies (common concerns abstraction)
# =============================================================================


@dataclass(frozen=True)
class DataCleaningPolicy:
    """データクリーニングの方針（内部拡張ポイント）。

    注意:
    - 既定値はテストで固定されており、CLI での切替は前提としません。
    - 実運用でのカスタムは、まず --transform かスキーマ定義での整形を推奨します。

    - normalize_array_field_shapes: 配列内の同名フィールドの形状を横並びで統一
    - prune_empty_elements: 完全に空の要素を除去
    - clean_empty_values: 空値を抑制して出力から除去
    """

    normalize_array_field_shapes: bool = True
    prune_empty_elements: bool = True
    clean_empty_values: bool = True


@dataclass(frozen=True)
class JsonPathInsertionPolicy:
    """JSONパス挿入時の方針（内部拡張ポイント）。

    - 通常は固定仕様を使用し、特殊な共存ルールはテストで固定した既定に従います。
    - 既定の value_key は "__value__"。外部からの変更は推奨しません。

    - promote_scalar_to_container_with_value_key: スカラーの上に子を挿入する際、{"__value__": scalar} に昇格
    - replace_empty_container_on_terminal: 末端にスカラを挿入する際、既存が空dictなら置換、非空dictなら __value__ に格納
    - value_key: スカラ保全に用いるキー名（互換のため "__value__" 既定）
    - one_based_index: 配列のインデックス基準（1基準が既定）
    """

    promote_scalar_to_container_with_value_key: bool = True
    replace_empty_container_on_terminal: bool = True
    value_key: str = "__value__"
    one_based_index: bool = True


@dataclass(frozen=True)
class SerializationPolicy:
    """シリアライズ（出力）時の方針（内部拡張ポイント）。

    - CLI の --output-format は引き続き有効です。その他の詳細は既定値で十分なはずです。
    - format: "json" | "yaml"
    - indent: インデント幅（JSON）
    - ensure_ascii: JSONのASCIIエスケープ有無
    - datetime_to_iso: datetime/date/time を ISO 文字列へシリアライズ
    """

    format: str = "json"
    indent: int = 2
    ensure_ascii: bool = False
    datetime_to_iso: bool = True


@dataclass(frozen=True)
class ValidationPolicy:
    """バリデーション時の方針（内部拡張ポイント）。

    - enabled: スキーマ検証を実施するか
    - to_iso_for_validation: 検証前に datetime/date/time を ISO 文字列へ正規化
    """

    enabled: bool = True
    to_iso_for_validation: bool = True


@dataclass(frozen=True)
class LoggingPolicy:
    """ログ出力の方針（内部拡張ポイント）。

    - debug_before_after_prune: prune 前後のサンプルキーをデバッグ出力
    - validation_error_to_file: 検証エラーをファイルへ出力
    """

    debug_before_after_prune: bool = True
    validation_error_to_file: bool = True


# 既定の共通ポリシー（現行の挙動を再現）
DEFAULT_DATA_CLEANING_POLICY = DataCleaningPolicy()
DEFAULT_JSON_PATH_INSERTION_POLICY = JsonPathInsertionPolicy()
DEFAULT_SERIALIZATION_POLICY = SerializationPolicy()
DEFAULT_VALIDATION_POLICY = ValidationPolicy()
DEFAULT_LOGGING_POLICY = LoggingPolicy()


def _is_numeric_key(key: str) -> bool:
    return bool(re.fullmatch(r"\d+", key))


# =============================================================================
# Extraction policies (missing types restored for compatibility)
# =============================================================================

@dataclass(frozen=True)
class NumericTokenPolicy:
    strict_spec_match: bool = True


@dataclass(frozen=True)
class NestedScanPolicy:
    ancestors_first_bounds: bool = True
    row_ownership_without_tokens: bool = True


@dataclass(frozen=True)
class ExtractionPolicy:
    nested_scan: NestedScanPolicy = NestedScanPolicy()
    numeric_tokens: NumericTokenPolicy = NumericTokenPolicy()


# 既定の抽出ポリシー（テストで参照される名前）
_DEFAULT_EXTRACTION_POLICY = ExtractionPolicy()


# =============================================================================
# Exceptions (compatibility)
# =============================================================================

class ConfigurationError(Exception):
    pass


class FileProcessingError(Exception):
    pass


# =============================================================================
# Helpers required by tests restored
# =============================================================================

def suppress_label_terminal_if_applicable(
    *,
    remaining_keys: List[str],
    original_path_keys: List[str],
    group_labels: set[str] | None,
    normalized_prefix: str,
    all_name_keys: List[str],
    container_parent_names: set[str],
) -> bool:
    """配列要素直下のラベル終端（例: json.A.1.lv1）を抑制するか。

    互換仕様:
    - remaining_keys が 1 要素かつ group_labels に含まれる場合
    - 対応する子アンカー（json.A.1.1 系）が all_name_keys に存在する場合 True
    - それ以外は False
    """
    try:
        if not remaining_keys or len(remaining_keys) != 1:
            return False
        label = remaining_keys[0]
        if not group_labels or label not in group_labels:
            return False
        # original_path_keys から親キーのベース（json.A.1）を組み立て、子アンカーの有無を確認
        # normalized_prefix は末尾に '.' を含む前提
        if not normalized_prefix.endswith("."):
            normalized_prefix = normalized_prefix + "."
        base = normalized_prefix + ".".join(original_path_keys[:2])  # json.A.1
        anchor_prefix = base + ".1"  # json.A.1.1
        for k in all_name_keys:
            if k.startswith(anchor_prefix):
                return True
        return False
    except Exception:
        return False


def _ensure_curr_list_or_raise(
    curr: Union[JSONDict, List[Any]],
    parent: Optional[Union[JSONDict, List[Any]]],
    parent_is_list: bool,
    parent_key: Optional[Union[str, int]],
    current_path: str,
) -> List[Any]:
    """カレントをlistとして確保。空dictなら親経由でlistに昇格、不可なら互換エラー。"""
    if is_json_list(curr):
        return cast(List[Any], curr)

    if is_json_dict(curr) and len(cast(Dict[str, Any], curr)) == 0 and parent is not None:
        new_list: List[Any] = []
        if parent_is_list and isinstance(parent_key, int):
            cast(List[Any], parent)[parent_key] = new_list
        elif (not parent_is_list) and isinstance(parent_key, str):
            cast(Dict[str, Any], parent)[parent_key] = new_list
        return new_list

    raise TypeError(f"Expected list at {current_path}, got {type(curr)}")


def _ensure_list_index_capacity(lcurr: List[Any], idx0: int) -> None:
    """`lcurr[idx0]`へ安全に代入できるよう長さを伸長する。"""
    while len(lcurr) <= idx0:
        lcurr.append(None)


def _assign_terminal_in_list(
    lcurr: List[Any], idx0: int, value: JSONValue, _ipol: JsonPathInsertionPolicy
) -> None:
    """終端でlist[idx0]に値を設定。空dictはvalue_key昇格ポリシーに従い代入。"""
    existing_last = lcurr[idx0] if idx0 < len(lcurr) else None
    if is_json_dict(existing_last):
        edict = cast(Dict[str, Any], existing_last)
        if len(edict) == 0:
            if _ipol.replace_empty_container_on_terminal:
                lcurr[idx0] = value
            else:
                edict[_ipol.value_key] = value
        else:
            edict[_ipol.value_key] = value
    else:
        lcurr[idx0] = value


def _prepare_next_container_for_list(
    lcurr: List[Any], idx0: int, next_is_numeric: bool, _ipol: JsonPathInsertionPolicy
) -> Union[JSONDict, List[Any]]:
    """次キーの型（数値/文字列）に応じて配下コンテナ(list/dict)を用意して返す。"""
    nxt_any = lcurr[idx0]
    if nxt_any is None:
        nxt_container: Union[JSONDict, List[Any]] = [] if next_is_numeric else {}
        lcurr[idx0] = nxt_container
        return nxt_container

    if not (is_json_dict(nxt_any) or is_json_list(nxt_any)):
        prev = nxt_any
        nxt_container2: Union[JSONDict, List[Any]]
        if next_is_numeric:
            nxt_container2 = []
        else:
            nxt_container2 = (
                {(_ipol.value_key): prev}
                if _ipol.promote_scalar_to_container_with_value_key
                else {}
            )
        lcurr[idx0] = nxt_container2
        return nxt_container2

    return cast(Union[JSONDict, List[Any]], nxt_any)


def _ensure_curr_dict_or_raise(
    curr: Union[JSONDict, List[Any]],
    parent: Optional[Union[JSONDict, List[Any]]],
    parent_is_list: bool,
    parent_key: Optional[Union[str, int]],
) -> JSONDict:
    """カレントをdictとして確保。空listなら親経由でdictに昇格、不可なら互換エラー。"""
    if is_json_dict(curr):
            return curr

    if is_json_list(curr) and len(cast(List[Any], curr)) == 0 and parent is not None:
        new_dict: Dict[str, Any] = {}
        if parent_is_list and isinstance(parent_key, int):
            cast(List[Any], parent)[parent_key] = new_dict
        elif (not parent_is_list) and isinstance(parent_key, str):
            cast(Dict[str, Any], parent)[parent_key] = new_dict
        return cast(JSONDict, new_dict)

    # 互換のエラーメッセージ
    raise TypeError("insert_json_path: root must be dict")


def _assign_terminal_in_dict(
    dcurr: JSONDict, key: str, value: JSONValue, _ipol: JsonPathInsertionPolicy
) -> None:
    """終端でdict[key]に値を設定。空dictはvalue_key昇格ポリシーに従い代入。"""
    if key in dcurr and is_json_dict(dcurr[key]):
        edict = cast(Dict[str, Any], dcurr[key])
        if len(edict) == 0:
            if _ipol.replace_empty_container_on_terminal:
                dcurr[key] = value
            else:
                edict[_ipol.value_key] = value
        else:
            edict[_ipol.value_key] = value
    else:
        dcurr[key] = value


def _ensure_next_child_for_dict(
    dcurr: JSONDict, key: str, next_is_numeric: bool, _ipol: JsonPathInsertionPolicy
) -> Union[JSONDict, List[Any]]:
    """次キーの型に応じて子(list/dict)を作成/昇格し、空相互変換も許容して返す。"""
    if key not in dcurr:
        dcurr[key] = [] if next_is_numeric else {}
        return cast(Union[JSONDict, List[Any]], dcurr[key])

    child = dcurr[key]
    # スカラーを辞書に昇格
    if not (is_json_dict(child) or is_json_list(child)):
        prev_value = child
        dcurr[key] = (
            {(_ipol.value_key): prev_value}
            if _ipol.promote_scalar_to_container_with_value_key
            else {}
        )
    else:
        # 空コンテナの相互変換
        if next_is_numeric and is_json_dict(child) and len(cast(Dict[str, Any], child)) == 0:
            dcurr[key] = []
        elif (not next_is_numeric) and is_json_list(child) and len(cast(List[Any], child)) == 0:
            dcurr[key] = {}

    return cast(Union[JSONDict, List[Any]], dcurr[key])


def insert_json_path(
    root: Union[JSONDict, List[Any]],
    keys: Union[Sequence[str], str],
    value: JSONValue,
    full_path: str = "",
    *,
    insertion_policy: JsonPathInsertionPolicy | None = None,
) -> None:
    """
    ドット区切りキーのリストまたは文字列から JSON 構造を構築し、値を挿入する。
    数字キーは list、文字列キーは dict として扱う。
    配列要素の構築時には適切に辞書から配列への変換も行う。
    """
    # パス正規化（空チェック含む）
    keys = _normalize_json_path_keys(keys)

    # ルート型チェック: dict または list のみ許可
    if not (is_json_dict(root) or is_json_list(root)):
        raise TypeError("insert_json_path: root must be dict or list")

    # ポリシー（既定は互換）
    _ipol = insertion_policy or DEFAULT_JSON_PATH_INSERTION_POLICY

    # 逐次処理（親参照を保持して dict<->list の昇格を安全に行う）
    parent: Optional[Union[JSONDict, List[Any]]] = None
    parent_is_list: bool = False
    parent_key: Optional[Union[str, int]] = None
    curr: Union[JSONDict, List[Any]] = root

    for i, key in enumerate(keys):
        is_last = i == len(keys) - 1
        current_path = f"{full_path}.{key}" if full_path else key
        is_num = _is_numeric_key(key)

        if is_num:
            parent, parent_is_list, parent_key, curr, inserted = _insert_path_numeric_step(
                parent=parent,
                parent_is_list=parent_is_list,
                parent_key=parent_key,
                curr=curr,
                key=key,
                is_last=is_last,
                value=value,
                next_key=(None if is_last else keys[i + 1]),
                current_path=current_path,
                _ipol=_ipol,
            )
            if inserted:
                return
        else:
            parent, parent_is_list, parent_key, curr, inserted = _insert_path_string_step(
                parent=parent,
                parent_is_list=parent_is_list,
                parent_key=parent_key,
                curr=curr,
                key=key,
                is_last=is_last,
                value=value,
                next_key=(None if is_last else keys[i + 1]),
                _ipol=_ipol,
            )
            if inserted:
                return


def _normalize_json_path_keys(keys: Union[Sequence[str], str]) -> List[str]:
    """`insert_json_path` 用にキー列を正規化し、空なら例外を出す。"""
    if isinstance(keys, str):
        norm = parse_json_path(keys)
    else:
        norm = list(keys)
    if not norm:
        raise ValueError(
            "JSONパスが空です。値を挿入するには少なくとも1つのキーが必要です。"
        )
    return norm


def _insert_path_numeric_step(
    *,
    parent: Optional[Union[JSONDict, List[Any]]],
    parent_is_list: bool,
    parent_key: Optional[Union[str, int]],
    curr: Union[JSONDict, List[Any]],
    key: str,
    is_last: bool,
    value: JSONValue,
    next_key: Optional[str],
    current_path: str,
    _ipol: JsonPathInsertionPolicy,
) -> tuple[Optional[Union[JSONDict, List[Any]]], bool, Optional[Union[str, int]], Union[JSONDict, List[Any]], bool]:
    """数値キー1ステップ分の処理を行い、親/現在参照を更新。終端なら挿入を完了して True を返す。"""
    idx0 = int(key) - 1
    if idx0 < 0:
        raise ValueError(f"配列インデックスは1以上である必要があります: {key}")

    lcurr = _ensure_curr_list_or_raise(curr, parent, parent_is_list, parent_key, current_path)
    _ensure_list_index_capacity(lcurr, idx0)

    if is_last:
        _assign_terminal_in_list(lcurr, idx0, value, _ipol)
        return parent, parent_is_list, parent_key, curr, True

    next_is_numeric = _is_numeric_key(next_key) if next_key is not None else False
    nxt_container = _prepare_next_container_for_list(lcurr, idx0, next_is_numeric, _ipol)
    return lcurr, True, idx0, nxt_container, False


def _insert_path_string_step(
    *,
    parent: Optional[Union[JSONDict, List[Any]]],
    parent_is_list: bool,
    parent_key: Optional[Union[str, int]],
    curr: Union[JSONDict, List[Any]],
    key: str,
    is_last: bool,
    value: JSONValue,
    next_key: Optional[str],
    _ipol: JsonPathInsertionPolicy,
) -> tuple[Optional[Union[JSONDict, List[Any]]], bool, Optional[Union[str, int]], Union[JSONDict, List[Any]], bool]:
    """文字列キー1ステップ分の処理を行い、親/現在参照を更新。終端なら挿入を完了して True を返す。"""
    dcurr = _ensure_curr_dict_or_raise(curr, parent, parent_is_list, parent_key)
    if is_last:
        _assign_terminal_in_dict(dcurr, key, value, _ipol)
        return parent, parent_is_list, parent_key, curr, True

    next_is_numeric = _is_numeric_key(next_key) if next_key is not None else False
    next_child = _ensure_next_child_for_dict(dcurr, key, next_is_numeric, _ipol)
    return dcurr, False, key, next_child, False


# =============================================================================
# Array Transform Rules
# =============================================================================


class ArrayTransformRule:
    """配列変換ルールを表すクラス"""

    def __init__(
        self,
        path: str,
        transform_type: str,
        transform_spec: str,
        trim_enabled: bool = False,
        *,
        apply_to_list_as_whole: bool = False,
    ):
        # パラメータの基本検証
        if not path:
            raise ValueError("pathは空ではない文字列である必要があります。")
        if not transform_type:
            raise ValueError("transform_typeは空ではない文字列である必要があります。")
        if not transform_spec:
            raise ValueError("transform_specは空ではない文字列である必要があります。")

        self.path = path
        self.transform_type = transform_type  # 'function', 'command', 'split'
        self.transform_spec = transform_spec
        self.trim_enabled = trim_enabled
        # リスト値に対して関数を配列全体へ1回だけ適用するか（既定: 各要素へ個別適用）
        self._transform_func: Optional[Callable] = None
        self._setup_transform()

    def _setup_transform(self):
        """変換関数をセットアップ"""
        if self.transform_type == "function":
            self._setup_python_function()
        elif self.transform_type == "command":
            self._setup_command()
        elif self.transform_type == "split":
            self._setup_split()
        else:
            raise ValueError(f"Unknown transform type: {self.transform_type}")

    def _setup_split(self):
        """split変換のセットアップ"""
        # transform_specは区切り文字（複数の場合は|で区切り）
        delimiter_str = self.transform_spec

        # パイプ文字のエスケープ処理
        if r"\|" in delimiter_str:
            delimiter_str = delimiter_str.replace(r"\|", "PIPE_ESCAPE_TEMP")

        # 区切り文字を分割
        delimiters = delimiter_str.split("|")

        # エスケープを元に戻す
        delimiters = [d.replace("PIPE_ESCAPE_TEMP", "|") for d in delimiters]

        # 特殊文字の置換
        for i, delimiter in enumerate(delimiters):
            delimiter = delimiter.replace("\\n", "\n")
            delimiter = delimiter.replace("\\t", "\t")
            delimiter = delimiter.replace("\\r", "\r")
            delimiters[i] = delimiter

        # 変換関数を設定
        def split_func(value):
            # 1区切りのみの場合は一次元配列として分割
            if len(delimiters) == 1:
                return convert_string_to_array(value, delimiters[0])
            else:
                return convert_string_to_multidimensional_array(value, delimiters)

        self._transform_func = split_func

    def _setup_python_function(self):
        """Python関数のセットアップ"""
        # 形式: module_path:function_name または file_path:function_name
        if ":" not in self.transform_spec:
            raise ValueError(
                f"Python function spec must be 'module:function' or 'file.py:function': {self.transform_spec}"
            )
        module_or_file, func_name = self.transform_spec.rsplit(":", 1)
        try:
            # 外部ファイルまたはモジュールの処理
            if module_or_file.endswith(".py"):
                # ファイルから関数を読み込み
                file_path = Path(module_or_file)
                if not file_path.exists():
                    raise FileNotFoundError(f"Transform file not found: {file_path}")

                spec = importlib.util.spec_from_file_location(
                    "transform_module", file_path
                )
                if spec is None or spec.loader is None:
                    raise ImportError(f"Cannot load module from {file_path}")

                module = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(module)
                self._transform_func = getattr(module, func_name)
            else:
                # モジュールから関数を読み込み (フォールバックで .py 探索)
                if module_or_file not in sys.modules and str(Path.cwd()) not in sys.path:
                    sys.path.insert(0, str(Path.cwd()))
                try:
                    module = importlib.import_module(module_or_file)
                    self._transform_func = getattr(module, func_name)
                except ModuleNotFoundError:
                    candidate = Path(module_or_file)
                    if candidate.suffix != ".py":
                        candidate = candidate.with_suffix(".py")
                    if not candidate.exists():
                        raise
                    spec = importlib.util.spec_from_file_location(
                        f"transform_module_{candidate.stem}", candidate
                    )
                    if spec is None or spec.loader is None:
                        raise ImportError(f"Cannot load module from {candidate}")
                    module = importlib.util.module_from_spec(spec)
                    spec.loader.exec_module(module)
                    self._transform_func = getattr(module, func_name)
        except Exception as e:
            raise ValueError(
                f"Failed to load transform function: {self.transform_spec}: {e}"
            ) from e

        logger.debug(f"Loaded transform function: {self.transform_spec}")

    def _setup_command(self):
        """外部コマンドのセットアップ"""
        # コマンドが実行可能かチェック
        try:
            result = subprocess.run(
                self.transform_spec.split()[0],
                capture_output=True,
                text=True,
                input="test",
                timeout=5,
            )
            logger.debug(f"Command available: {self.transform_spec}")
        except (subprocess.TimeoutExpired, FileNotFoundError) as e:
            logger.warning(f"Command check failed for '{self.transform_spec}': {e}")

    def transform(self, value: Any, workbook=None) -> Any:
        """値を変換

        valueがExcelの名前付き範囲の値である場合、そのデータ形式
        （値、1次元配列、2次元配列、さらに高次元配列）に応じて適切に変換関数に渡す

        変換関数が辞書を返した場合、動的セル名構築として処理する
        """
        if self.transform_type == "function":
            result = self._transform_with_function(value)
            # trim指定時は配列要素をstrip()
            if self.trim_enabled and isinstance(result, list):
                return self._apply_trim_recursively(result)
            return result
        elif self.transform_type == "command":
            return self._transform_with_command(value)
        elif self.transform_type == "split":
            return self._apply_split_recursively(value)
        else:
            return value

    def _apply_trim_recursively(self, data: Any) -> Any:
        """多次元配列に対して再帰的にstripを適用"""
        if isinstance(data, list):
            return [self._apply_trim_recursively(item) for item in data]
        elif isinstance(data, str):
            return data.strip()
        else:
            return data

    def _apply_split_recursively(self, value: Any, depth: Optional[int] = None) -> Any:
        """多次元 split 変換。リスト要素も個別に処理する。"""
        if isinstance(value, list):
            # リストの各要素を個別に変換
            result = []
            for item in value:
                processed_item = self._apply_split_recursively(item, depth)
                # 各要素を独立して処理（extendではなくappend）
                result.append(processed_item)
            return result
        elif isinstance(value, str):
            # 文字列を変換関数で処理
            if self._transform_func is None:
                return value
            return self._transform_func(value)
        else:
            return value

    def _transform_with_function(self, value: Any) -> Any:
        """関数による変換（標準出力をキャプチャしつつ、例外は握りつぶさない）"""
        if self._transform_func is None:
            logger.warning(f"Transform function not initialized: {self.transform_spec}")
            return value

        stdout_capture = io.StringIO()
        stderr_capture = io.StringIO()

        with redirect_stdout(stdout_capture), redirect_stderr(stderr_capture):
            result = self._transform_func(value)

        # ログは無加工（トリム無し）で全量出力
        stdout_content = stdout_capture.getvalue()
        stderr_content = stderr_capture.getvalue()
        if stdout_content:
            logger.debug(
                "Transform function stdout: %s",
                stdout_content,
                extra={
                    "transform_spec": self.transform_spec,
                    "transform_type": "function",
                },
            )
        if stderr_content:
            logger.warning(
                "Transform function stderr: %s",
                stderr_content,
                extra={
                    "transform_spec": self.transform_spec,
                    "transform_type": "function",
                },
            )

        return result

    def _transform_with_command(self, value: Any) -> Any:
        """外部コマンドで変換（タイムアウトのみ明示的に捕捉）

        入力正規化仕様（2025-09 改訂）:
        - dict / ネストを含む list/tuple は JSON 文字列 (ensure_ascii=False) として渡す
        - フラット（全要素がスカラー）の list/tuple は従来通り 改行結合
        - それ以外のスカラー値は str() 変換
        この方針で "構造" を壊さずにコマンドへ受け渡し可能にする。
        """

        def _is_scalar(x: Any) -> bool:
            return isinstance(x, (str, int, float, bool)) or x is None

        def _is_flat_scalar_list(v: Any) -> bool:
            return isinstance(v, (list, tuple)) and all(_is_scalar(e) for e in v)

        def _json_default(o: Any):  # JSON化できない set 等を救済
            if isinstance(o, set):
                try:
                    return sorted(list(o))
                except Exception:  # noqa: BLE001
                    return list(o)
            raise TypeError(f"Object of type {type(o)} is not JSON serializable")

        treat_multiline_as_list = False

        if isinstance(value, dict):
            try:
                input_str = json.dumps(value, ensure_ascii=False, default=_json_default)
            except Exception:  # noqa: BLE001
                input_str = str(value)
        elif isinstance(value, (list, tuple)):
            if _is_flat_scalar_list(value):
                # フラット: 改行結合（既存コマンド sort -u 等との親和性を保つ）
                input_str = "\n".join("" if v is None else str(v) for v in value)
                treat_multiline_as_list = True
            else:
                try:
                    input_str = json.dumps(value, ensure_ascii=False, default=_json_default)
                except Exception:  # noqa: BLE001
                    # 失敗したら安全側で repr 文字列
                    input_str = str(value)
        else:
            input_str = str(value) if value is not None else ""
        try:
            result = subprocess.run(
                shlex.split(self.transform_spec),
                input=input_str,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                timeout=30,
            )
        except subprocess.TimeoutExpired:
            logger.error(
                "Command timeout: %s",
                self.transform_spec,
                extra={"transform_spec": self.transform_spec},
            )
            return value
        except Exception as e:
            logger.error(
                "Command execution error: %s: %s",
                self.transform_spec,
                e,
                extra={"transform_spec": self.transform_spec},
            )
            return value

        # 標準出力・標準エラーは無加工で全量ログ（存在時）
        if result.stdout:
            logger.debug(
                "Command stdout: %s",
                result.stdout,
                extra={
                    "transform_spec": self.transform_spec,
                    "returncode": result.returncode,
                },
            )
        if result.stderr:
            logger.warning(
                "Command stderr: %s",
                result.stderr,
                extra={
                    "transform_spec": self.transform_spec,
                    "returncode": result.returncode,
                },
            )

        if result.returncode != 0:
            logger.warning(
                "Command failed: %s (returncode=%s)",
                self.transform_spec,
                result.returncode,
                extra={
                    "transform_spec": self.transform_spec,
                    "returncode": result.returncode,
                },
            )
            return value

        output = result.stdout

        # 1. JSON として解釈できれば優先（後続加工を避ける）
        try:
            return json.loads(output.strip())
        except json.JSONDecodeError:
            pass

        # 2. フラットスカラ配列入力だった場合は複数行を行配列へ
        if treat_multiline_as_list and "\n" in output:
            lines = [ln for ln in output.split("\n") if ln.strip()]
            return lines

        # 3. それ以外は文字列のまま返す
        return output


def parse_array_transform_rules(  # noqa: PLR0915 長さは後続リファクタ対象
    array_transform_rules: Sequence[Union[str, Dict[str, str]]],
    prefix: str,
    schema: Optional[Dict[str, Any]] = None,
    trim_enabled: bool = False,
) -> Dict[str, List[ArrayTransformRule]]:
    """
    配列変換ルールのパース。
    形式: "json.path=function:module:func_name" または "json.path=command:cat"
    ワイルドカード対応: "json.arr.*.name=split:," または "json.arr.*=range:A1:B2:function:builtins:len"
    連続適用対応: 同一セル名に対する複数の--transform指定を順次適用
    """
    if not prefix:
        raise ValueError("prefixは空ではない文字列である必要があります。")

    rules: Dict[str, List[ArrayTransformRule]] = {}
    wildcard_rules: Dict[str, List[ArrayTransformRule]] = {}

    # prefixの末尾にドットがなければ自動で追加
    normalized_prefix = prefix if prefix.endswith(".") else prefix + "."

    def _split_rule(raw: str) -> tuple[str, str] | None:
        if "=" not in raw:
            logger.warning(f"無効な変換設定: {raw}")
            return None
        left, right = raw.split("=", 1)
        return left, right

    def _normalize_path(path: str) -> tuple[str, bool]:
        has_wildcard = "*" in path
        if path.startswith(normalized_prefix):
            path = path[len(normalized_prefix) :]
        elif path.startswith(prefix):
            path = path[len(prefix) :]
            if not path.startswith("."):
                path = "." + path if path else ""
            if path.startswith("."):
                path = path[1:]
        return path, has_wildcard

    def _parse_spec(spec: str) -> tuple[str, str]:
        if spec.startswith("function:"):
            return "function", spec[len("function:") :]
        if spec.startswith("command:"):
            return "command", spec[len("command:") :]
        if spec.startswith("split:"):
            return "split", spec[len("split:") :]
        return "function", spec

    def _insert_rule(dst: Dict[str, List[ArrayTransformRule]], key: str, rule_obj: ArrayTransformRule) -> None:
        if key not in dst:
            dst[key] = []
        dst[key].append(rule_obj)

    normalized_inputs: List[tuple[str, str]] = []
    for raw in array_transform_rules:
        if isinstance(raw, dict):
            key = raw.get("key")
            func = None
            if "function" in raw:
                func = f"function:{raw['function']}" if not str(raw['function']).startswith(("function:", "command:", "split:")) else raw['function']
            elif "command" in raw:
                func = f"command:{raw['command']}"
            elif "split" in raw:
                func = f"split:{raw['split']}"
            if not key or not func:
                logger.warning(f"無効な変換設定: {raw}")
                continue
            normalized_inputs.append((key, func))
            continue
        splitted = _split_rule(raw)
        if not splitted:
            continue
        path, transform_spec = splitted
        normalized_inputs.append((path, transform_spec))

    for path, transform_spec in normalized_inputs:
        path, has_wildcard = _normalize_path(path)

        try:
            transform_type, actual_spec = _parse_spec(transform_spec)
            rule_obj = ArrayTransformRule(path, transform_type, actual_spec, trim_enabled)
            if has_wildcard:
                _insert_rule(wildcard_rules, path, rule_obj)
            else:
                _insert_rule(rules, path, rule_obj)
        except Exception as e:
            logger.error(f"変換ルール作成エラー: {raw}, エラー: {e}")
            continue

    if wildcard_rules:
        # 既存仕様: 記載順（後勝ち）を保持するためそのまま更新
        rules.update(wildcard_rules)

    return rules


def get_applicable_transform_rules(
    transform_rules: TransformRulesMap,
    normalized_path_keys: List[str],
    original_path_keys: List[str],
) -> Optional[List[ArrayTransformRule]]:
    """normalized/original の両方で、完全一致→親キー→ワイルドカードの順に適用可能な変換ルールを返す。

    優先順位:
    1) 完全一致（normalized → original）
    2) 親キー一致（normalized → original）
    3) ワイルドカード（*）パターンマッチ（normalized → original）
    """
    if not transform_rules:
        return None

    key_path = ".".join(normalized_path_keys)
    orig_key_path = ".".join(original_path_keys)

    def _find_exact_match() -> Optional[List[ArrayTransformRule]]:
        if key_path in transform_rules:
            return transform_rules[key_path]
        if orig_key_path in transform_rules:
            return transform_rules[orig_key_path]
        return None

    def _find_parent_match() -> Optional[List[ArrayTransformRule]]:
        if len(normalized_path_keys) > 1:
            for i in range(len(normalized_path_keys) - 1, 0, -1):
                parent_norm = ".".join(normalized_path_keys[:i])
                if parent_norm in transform_rules:
                    return transform_rules[parent_norm]
        if len(original_path_keys) > 1:
            for i in range(len(original_path_keys) - 1, 0, -1):
                parent_orig = ".".join(original_path_keys[:i])
                if parent_orig in transform_rules:
                    return transform_rules[parent_orig]
        return None

    def _find_wildcard_match() -> Optional[List[ArrayTransformRule]]:
        for rule_key, rule_list in transform_rules.items():
            if "*" in rule_key and (
                wildcard_match_path(rule_key, key_path)
                or wildcard_match_path(rule_key, orig_key_path)
            ):
                return rule_list
        return None

    return _find_exact_match() or _find_parent_match() or _find_wildcard_match()


def apply_transform_rules_for_path(
    *,
    rules: List[ArrayTransformRule],
    value: Any,
    workbook,
    insert_keys: List[str],
    root_result: JSONDict,
    transform_rules_map: TransformRulesMap,
    prefix: str,
) -> Any:
    """順次ルール適用と辞書戻り値の動的セル名処理を行い、最終値を返す。

    - rules を順に適用
    - dict 戻り値は絶対/相対キーへ展開（再帰的に適用ルールも探索して追適用）
    - list は各要素に対して関数適用
    - 最終スカラ/配列は呼び出し側で insert する
    """
    current_value = value

    # Build a TransformContext for internal helpers to consume
    tctx = TransformContext(
        workbook=workbook,
        prefix=prefix,
        transform_rules_map=transform_rules_map,
        insert_keys=insert_keys,
        root_result=root_result,
    )

    def _apply_one(rule: ArrayTransformRule, val: Any):
        # Pass workbook via context; transform implementations expect (val, workbook)
        return rule.transform(val, tctx.workbook)

    for i, tr in enumerate(rules):
        log_transform_progress(
            step_index=i + 1, total_steps=len(rules), insert_keys=insert_keys, rule=tr
        )
        current_value = _apply_one(tr, current_value)

    # 新仕様: 辞書戻り値はそのまま値として扱う（動的セル名展開はしない）
    if rules:
        logger.debug(f"変換ルール{len(rules)}後の値: {current_value}")

    return current_value


# =============================================================================
# Wildcard helpers (module-level)
# =============================================================================


def wildcard_match_path(pattern: str, actual_path: str) -> bool:
    """拡張ワイルドカードマッチ。

    仕様:
    - セグメント数は一致すること
    - セグメント全体が '*' の場合: 任意1セグメント
    - セグメントに部分ワイルドカード（例: '*items', 'pre*', 'mid*post'）を含められる
      → '*' は同一セグメント内で 0 文字以上にマッチ
    - 正規表現は使用せずキャッシュ不要な軽量実装
    """
    pattern_parts = [p for p in pattern.split('.') if p]
    actual_parts = [p for p in actual_path.split('.') if p]
    if len(pattern_parts) != len(actual_parts):
        return False

    def seg_match(pp: str, ap: str) -> bool:
        if pp == '*':
            return True
        if '*' not in pp:
            return pp == ap
        # 部分ワイルドカード: 連続 '*' は一つに畳んで処理
        # 例: '*items' -> suffix match, 'pre*' -> prefix match, 'a*b*c' -> subsequence match
        tokens = [t for t in pp.split('*')]
        if len(tokens) == 1:
            # '*' が末尾にあったケースで split 結果が ['xxx', ''] になるのでここは通常到達しない
            return pp == ap
        # 位置合わせ: 最初のトークンは先頭一致、最後のトークンは末尾一致、中間は順序出現
        # 空トークンはスキップ
        cur = 0
        first_token_consumed = False
        for i, tk in enumerate(tokens):
            if tk == '':
                continue
            if i == 0:  # 先頭
                if not ap.startswith(tk):
                    return False
                cur = len(tk)
                first_token_consumed = True
                continue
            # 中間/末尾: 残りを検索
            idx = ap.find(tk, cur)
            if idx < 0:
                return False
            cur = idx + len(tk)
            if i == len(tokens) - 1 and tokens[-1] != '' and not ap.endswith(tk):
                # 末尾トークンは末尾一致
                return False
        # 末尾が '' の場合は何でも可
        # 先頭トークンが '' のとき（pp 始まりが '*'）も上記ロジックで許容
        return True

    return all(seg_match(pp, ap) for pp, ap in zip(pattern_parts, actual_parts))


def handle_parent_level_for_double_index_array(
    *,
    wb: Any,
    defined_name: Any,
    value: Any,
    array_ref: List[Any],
    array_name: str,
    array_index: int,
    path_keys: List[str],
    name: str,
    normalized_prefix: str,
    gen_map: Optional[Dict[str, Any]],
    expected_field_shape: Dict[Tuple[str, str], str],
) -> bool:
    """二重数値インデックス配列に対する親レベル（parent.i.*）の処理を行う。

    契約:
    - 入力: [i] までの準備は済んでおり、`array_ref[array_index]` が None/{} or list で存在
    - path_keys は少なくとも [array, i, ...] の形をとる
    - 当該 i に既に生成名が存在する場合、親レベルからの挿入をスキップする
    - 末端が field.index の場合は形状保持で値を再取得し、1D追加マージ（2Dはそのまま）
    - 末端が通常の field の場合は expected_field_shape に従い 1D/2D へ昇格

    戻り値:
    - True: 親レベルで処理を完了（呼び出し元は continue）
    - False: 親レベルでは未処理（呼び出し元で続行）
    """
    try:
        # 小さなヘルパーで分岐を整理（挙動は完全維持）
        def _has_generated_for_index() -> bool:
            return should_skip_array_anchor_insertion(
                array_name, array_index, normalized_prefix, gen_map
            )

        def _ensure_list_at_index() -> Optional[List[Any]]:
            """array_ref[array_index] をリストへ昇格できたら返す。昇格不可（非空dict等）なら None。"""
            if is_json_list(array_ref[array_index]):
                return cast(List[Any], array_ref[array_index])
            promoted0 = _promote_element_to_list_if_appropriate(array_ref[array_index])
            if is_json_list(promoted0):
                array_ref[array_index] = promoted0
                return cast(List[Any], promoted0)
            return None

        def _ensure_first_dict(inner: List[Any]) -> Dict[str, Any]:
            if not inner:
                inner.append(None)
            return _ensure_nested_dict_at(inner, 0)

        def _is_field_index(rem: List[str]) -> bool:
            return len(rem) == 2 and (not rem[0].isdigit()) and rem[1].isdigit()

        def _is_field_only(rem: List[str]) -> bool:
            return len(rem) == 1 and (not rem[0].isdigit())

        def _apply_field_index_case(target: Dict[str, Any], field_token: str) -> bool:
            try:
                subval = get_named_range_values_preserve_shape(wb, defined_name)
            except Exception:
                subval = value
            subval_list = _coerce_to_1d(subval)
            _set_or_merge_list_field(target, field_token, subval_list)
            logger.debug(
                "NEST-MERGE [i][0] %s[%s][0].%s += %r (from field.index)",
                array_name,
                array_index,
                field_token,
                subval_list,
            )
            return True

        def _apply_field_only_case(target: Dict[str, Any], fld: str) -> bool:
            coerced = apply_expected_shape_to_value(
                value,
                field_name=fld,
                expected_field_shape=expected_field_shape,
                array_name=array_name,
            )
            if not (fld in target and is_nonempty_array_or_dict(target.get(fld))):
                target[fld] = coerced
                logger.debug(
                    "NEST-SET [i][0] %s[%s][0].%s=%r (coerced)",
                    array_name,
                    array_index,
                    fld,
                    coerced,
                )
            return True

        # 生成名がある場合は親レベルをスキップ
        if _has_generated_for_index():
            logger.debug(
                "PARENT-SKIP %s[%s] due to generated nested under this index: %s",
                array_name,
                array_index,
                name,
            )
            return True

        # 昇格に失敗（非空dict等）した場合は親レベル未処理
        inner0 = _ensure_list_at_index()
        if inner0 is None:
            return False

        target0 = _ensure_first_dict(inner0)
        remaining_keys = path_keys[2:]

        if _is_field_index(remaining_keys):
            return _apply_field_index_case(target0, remaining_keys[0])
        if _is_field_only(remaining_keys):
            return _apply_field_only_case(target0, remaining_keys[0])
    except Exception:
        # 例外時は親レベルでは未処理として呼び出し側へ返す（安全側）
        return False
    return False


def get_nested_value(obj: JSONDict, path: str) -> Any:
    """dict のドット区切りパスから値を取得（存在しなければ None）。"""
    parts = [p for p in path.split(".") if p]
    cur: JSONValue = obj
    for part in parts:
        if is_json_dict(cur) and part in cur:
            cur = cast(JSONDict, cur)[part]
        else:
            return None
    return cur


def set_nested_value(obj: JSONDict, path: str, value: JSONValue) -> None:
    """dict のドット区切りパスに値を設定（中間 dict は自動生成）。"""
    parts = [p for p in path.split(".") if p]
    cur: JSONDict = obj
    for part in parts[:-1]:
        if part not in cur or not is_json_dict(cur.get(part)):
            cur[part] = {}
        cur = cast(JSONDict, cur[part])
    cur[parts[-1]] = value


def find_matching_paths(
    obj: JSONValue, pattern: str, current_path: str = ""
) -> List[str]:
    """
    dict 構造から、パターンにマッチするドット区切りパスを列挙する。
    リストは配列インデックスをキーとみなさず、子要素を同じ current_path 配下として再帰探索する。
    """
    matches: List[str] = []
    if is_json_dict(obj):
        for key, value in obj.items():
            new_path = f"{current_path}.{key}" if current_path else key
            if wildcard_match_path(pattern, new_path):
                matches.append(new_path)
            if is_json_dict(value) or is_json_list(value):
                matches.extend(find_matching_paths(value, pattern, new_path))
    elif is_json_list(obj):
        # リスト: 各要素に対し同じ current_path を継続しつつ要素(dict) 自体がパターンに合致するケースを許容
        for item in obj:
            # 要素が dict の場合 current_path 自体を対象として再評価する（例えば *.items.* で items 配下の要素をマッチさせる）
            if is_json_dict(item) and current_path and wildcard_match_path(pattern, current_path):
                matches.append(current_path)
            matches.extend(find_matching_paths(item, pattern, current_path))
        # 特別処理: pattern が current_path + '.*' に近く、要素が dict の場合は current_path を候補に含める
        if current_path and any(is_json_dict(it) for it in obj):
            # pattern セグメント数と current_path セグメント数の差が1 以上で最後のパターンセグメントが '*' 含む場合
            p_parts = [p for p in pattern.split('.') if p]
            c_parts = [p for p in current_path.split('.') if p]
            if len(p_parts) >= len(c_parts) and wildcard_match_path(".".join(p_parts[:len(c_parts)]), current_path):
                # 直後のパターンセグメントに '*' を含むならノード自体を追加
                if len(p_parts) > len(c_parts) and '*' in p_parts[len(c_parts)]:
                    if current_path not in matches:
                        matches.append(current_path)
    return matches


def log_transform_progress(
    *,
    step_index: int,
    total_steps: int,
    insert_keys: List[str],
    rule: ArrayTransformRule,
) -> None:
    """変換進捗のデバッグログを統一出力。"""
    logger.debug(
        "変換ルール%s/%sで変換: %s -> rule=%s:%s",
        step_index,
        total_steps,
        insert_keys,
        getattr(rule, "transform_type", None),
        getattr(rule, "transform_spec", None),
    )


def extract_abs_path_from_prefixed_key(key: str, prefix: str) -> str:
    """
    prefix 付きキーから絶対パス部分を抽出する。
    - 'json.foo.bar' のようなキーに対して prefix='json' なら 'foo.bar' を返す
    - 'json' 単体や prefix に一致しない場合は空文字列
    """
    if not key.startswith(prefix):
        return ""
    if key == prefix:
        return ""
    if key.startswith(prefix + "."):
        return key[len(prefix + ".") :]
    # 'jsonX...' のような紛らわしいケースは prefix とはみなさない
    return ""


def _apply_dynamic_rules_if_any(
    *,
    value: Any,
    transform_rules_map: Optional[TransformRulesMap],
    parts: List[str],
    workbook,
    apply_dynamic_rules: bool,
) -> Any:
    if apply_dynamic_rules and transform_rules_map is not None:
        dyn_rules = get_applicable_transform_rules(transform_rules_map or {}, parts, parts)
        if dyn_rules:
            for dr in dyn_rules:
                value = dr.transform(value, workbook)
    return value


def _ensure_dict_path(root: JSONDict, parts: List[str]) -> JSONDict:
    cur = root
    for ap in parts:
        if ap not in cur:
            cur[ap] = {}
        cur = cast(JSONDict, cur[ap])
    return cur


def expand_and_insert_dict(
    *,
    result_dict: JSONDict,
    base_path: str,
    prefix: str,
    root_result: JSONDict,
    transform_rules_map: Optional[TransformRulesMap] = None,
    workbook=None,
    apply_dynamic_rules: bool = False,
) -> None:
    """
    変換の結果が dict の場合に、各キーを絶対/相対指定として評価して root_result に書き込む共通処理。

    - 絶対指定: 'json.foo.bar'（prefix が json の場合）
    - 相対指定: 'baz'（base_path の直下に挿入）
    - transform_rules_map が与えられる場合は挿入前に該当ルールを追適用する
    """
    for key, val in result_dict.items():
        if key.startswith(prefix):
            abs_path = extract_abs_path_from_prefixed_key(key, prefix)
            if not abs_path:
                continue
            abs_parts = [p for p in abs_path.split(".") if p]
            new_val = _apply_dynamic_rules_if_any(
                value=val,
                transform_rules_map=transform_rules_map,
                parts=abs_parts,
                workbook=workbook,
                apply_dynamic_rules=apply_dynamic_rules,
            )
            cur = _ensure_dict_path(root_result, abs_parts[:-1])
            cur[abs_parts[-1]] = new_val
        else:
            rel_path = f"{base_path}.{key}" if base_path else key
            rel_parts = [p for p in rel_path.split(".") if p]
            new_val = _apply_dynamic_rules_if_any(
                value=val,
                transform_rules_map=transform_rules_map,
                parts=rel_parts,
                workbook=workbook,
                apply_dynamic_rules=apply_dynamic_rules,
            )
            set_nested_value(root_result, rel_path, new_val)


def convert_string_to_multidimensional_array(value: Any, delimiters: List[str]) -> Any:
    """
    文字列を指定された区切り文字のリストで多次元配列に変換する。
    最終次元は個別の文字列要素として展開される。

    Args:
        value: 変換対象の値
        delimiters: 区切り文字のリスト（1次元目、2次元目、3次元目...の順）

    Returns:
        多次元配列に変換された値
    """
    if not isinstance(value, str):
        return value

    if not value.strip():
        return []

    if not delimiters:
        return value

    def split_recursively(text: str, delimiter_list: List[str], depth: int = 0) -> Any:
        if not delimiter_list:
            return text.strip()

        current_delimiter = delimiter_list[0]
        remaining_delimiters = delimiter_list[1:]

        # 現在の区切り文字で分割
        parts = text.split(current_delimiter)

        result = []
        for part in parts:
            part = part.strip()
            if part:
                if remaining_delimiters:
                    # まだ区切り文字が残っている場合は再帰処理
                    sub_result = split_recursively(
                        part, remaining_delimiters, depth + 1
                    )
                    result.append(sub_result)
                else:
                    # 最後の区切り文字の場合：個別の文字列として追加
                    result.append(part)

        # 最終次元の場合、配列を個別要素として展開
        if len(delimiter_list) == 1:
            return result  # 最終次元なので配列のまま返す

        return result if result else []

    return split_recursively(value, delimiters)


def convert_string_to_array(value: Any, delimiter: str) -> Any:
    """
    文字列を指定された区切り文字で配列に変換する。
    （後方互換性のため残存）
    """
    if not isinstance(value, str):
        return value

    if not value.strip():
        return []

    # 区切り文字で分割
    parts = value.split(delimiter)
    # 前後の空白を削除
    result = [part.strip() for part in parts if part.strip()]

    return result if result else []


def parse_array_split_rules(
    rules: Optional[List[Optional[str]]], prefix: str
) -> Dict[str, List[str]]:
    r"""配列分割ルールの解析ヘルパー。

    入力例:
        ["json.field1=,", "json.nested.field2=;|\\n", "json.field3=\\t|\\|"]
        ["json.data=split:,", "json.items=split:;|\\n"] も許可

    返却:
        {"field1": [","], "nested.field2": [";", "\n"], "field3": ["\t", "|"]}

    - prefix は "json" または "json." いずれも受け付ける
    - 無効な行は warning を出して無視
    - デリミタ列は '|' で区切る。'\|' はリテラルのパイプを表す
    - エスケープシーケンス \n, \t, \r, \\ を展開
    """
    result: Dict[str, List[str]] = {}
    if rules is None or len(rules) == 0:
        return result
    if not isinstance(prefix, str) or prefix == "":
        raise ValueError("prefixは空ではない文字列である必要があります。")

    norm_pref = prefix if prefix.endswith(".") else prefix + "."

    def _decode_delims(spec: str) -> List[str]:
        tokens: List[str] = []
        buf: List[str] = []
        escape = False
        for ch in spec:
            if escape:
                buf.append({"n": "\n", "t": "\t", "r": "\r"}.get(ch, ch))
                escape = False
                continue
            if ch == "\\":
                escape = True
                continue
            if ch == "|":
                tokens.append("".join(buf))
                buf = []
            else:
                buf.append(ch)
        if escape:
            buf.append("\\")
        tokens.append("".join(buf))
        return [t for t in tokens if t != ""]

    def _normalize_left_path(left: str) -> str:
        path = left.strip()
        if path.startswith(norm_pref):
            path = path[len(norm_pref) :]
        elif path.startswith(prefix):
            path = path[len(prefix) :]
            if path.startswith("."):
                path = path[1:]
        return path.strip().strip(".")

    def _parse_rule_line(raw: str) -> tuple[str, str] | None:
        if not raw or not isinstance(raw, str):
            return None
        if "=" not in raw:
            logger.warning("無効な配列化設定: '='がありません: %r", raw)
            return None
        left, right = raw.split("=", 1)
        path = _normalize_left_path(left)
        if not path:
            logger.warning("無効な配列化設定: 空のパス: %r", raw)
            return None
        spec = right.strip()
        if spec.startswith("split:"):
            spec = spec[len("split:") :]
        return path, spec

    for raw in rules:
        parsed = _parse_rule_line(raw)  # type: ignore[arg-type]
        if not parsed:
            continue
        path, spec = parsed
        delims: List[str]
        if spec == "":
            delims = [","]
        else:
            delims = _decode_delims(spec)
            if not delims:
                logger.warning("無効な配列化設定: デリミタ指定が無効: %r", raw)
                continue
        result[path] = delims

    return result


def try_apply_transform_and_insert(
    *,
    array_transform_rules: Optional[Dict[str, List[ArrayTransformRule]]],
    path_keys: List[str],
    original_path_keys: List[str],
    value: Any,
    workbook,
    root_result: Dict[str, Any],
    prefix: str,
    safe_insert: Callable[[Union[Dict[str, Any], List[Any]], List[str], Any, str, str, List[str], List[str]], None],
) -> bool:
    """パスにマッチする変換ルールを適用し、挿入まで実施する。

    変換ルールが見つからない場合は False を返し、呼び出し元は通常処理を継続する。
    見つかった場合は適用の上で挿入し True を返す。
    """
    transform_rules = get_applicable_transform_rules(
        array_transform_rules or {}, path_keys, original_path_keys
    )
    if transform_rules is None:
        return False
    insert_keys = path_keys
    # 内部使用のための TransformContext を構築
    tctx = TransformContext(
        workbook=workbook,
        prefix=prefix,
        transform_rules_map=array_transform_rules or {},
        insert_keys=insert_keys,
        root_result=root_result,
    )

    current_value = apply_transform_rules_for_path(
        rules=transform_rules,
        value=value,
        workbook=tctx.workbook,
        insert_keys=tctx.insert_keys,
        root_result=tctx.root_result,
        transform_rules_map=tctx.transform_rules_map,
        prefix=tctx.prefix,
    )
    safe_insert(
        root_result,
        insert_keys,
        current_value,
        ".".join(insert_keys),
        ".".join([prefix] + original_path_keys),
        original_path_keys,
        path_keys,
    )
    return True


def preseed_root_keys(
    *, root_result: Dict[str, Any], root_first_pos: Dict[str, tuple[int, int, int]]
) -> None:
    """ルートキーを出現順に事前作成して dict の順序を安定化させる。

    既存値は温存し、存在しないキーのみ空 dict をセットする。
    root_first_pos が空のときは何もしない。
    """
    try:
        if is_json_dict(root_result) and root_first_pos:
            desired_roots = [k for k, _ in sorted(root_first_pos.items(), key=lambda kv: kv[1])]
            for rk in desired_roots:
                root_result.setdefault(rk, {})
    except Exception as e:
        # ここでの安定化は最適化目的のため、失敗しても致命ではない
        logger.debug("preseed_root_keys skipped due to: %s", e)


# =============================================================================
# Named Range Parsing
# =============================================================================


def _prepare_parsing_prelude(
    *,
    wb,
    prefix: str,
    containers: Optional[Dict[str, Dict]],
    global_max_elements: Optional[int],
    extraction_policy: ExtractionPolicy,
) -> Dict[str, Any]:
    """パース前の派生情報をまとめて構築し、状態辞書を返す。"""
    # コンテナ準備
    containers, user_provided_containers, _generated_names = prepare_containers_and_generated_names(
        wb,
        prefix=prefix,
        containers=containers,
        global_max_elements=global_max_elements,
        extraction_policy=extraction_policy,
    )

    # prefix の正規化
    normalized_prefix = prefix if prefix.endswith(".") else prefix + "."

    # 定義名 + 生成名を統合（初期スナップショット）
    all_names, defined_only_name_keys = build_all_names_with_generated(wb)

    # 既存仕様: フィールド直下の *.field.1 用に補助生成名を後から追加
    try:
        generate_subarray_names_for_field_anchors(wb, normalized_prefix)
    except Exception as _e:
        logger.debug("subarray name generation skipped due to error: %s", _e)

    all_name_keys = list(all_names.keys())
    # 生成名マップ（補助生成名含む、最新状態）
    gen_map = get_generated_names_map(wb)

    # *.field と *.field.1 の重複抑止集合
    excluded_indexed_field_names: set[str] = compute_excluded_indexed_field_names(
        normalized_prefix, all_name_keys, all_names
    )

    # 各フィールドの期待形状（1D/2D）
    expected_field_shape: Dict[Tuple[str, str], str] = learn_expected_field_shapes(
        normalized_prefix, all_name_keys, all_names
    )

    # ルート直下の数値キー
    numeric_root_keys = compute_numeric_root_keys(normalized_prefix, all_names)

    # 二重数値インデックス（array.i.j.*）が存在する配列名集合
    arrays_with_double_index: set[str] = find_arrays_with_double_index(
        normalized_prefix=normalized_prefix, all_name_keys=all_name_keys, gen_map=gen_map
    )

    container_parent_names: set[str] = set(containers.keys()) if containers else set()
    container_parents_with_children: set[str] = compute_container_parents_with_children(
        container_parent_names=container_parent_names,
        all_name_keys=all_name_keys,
        gen_map=gen_map,
    )

    # groupLabel -> rootName
    group_to_root: dict[str, str] = compute_group_to_root_map(
        containers=containers,
        prefix=prefix,
        normalized_prefix=normalized_prefix,
        all_name_keys=all_name_keys,
    )

    # アンカー名とグループラベル
    anchor_names: set[str] = compute_anchor_names(normalized_prefix, all_name_keys)
    group_labels: set[str] = compute_group_labels_from_anchors(
        anchor_names, containers, prefix=prefix, normalized_prefix=normalized_prefix
    )

    # シート順序とルート最初の出現位置
    sheet_order: Dict[str, int] = {ws.title: idx for idx, ws in enumerate(wb.worksheets)}
    root_first_pos: Dict[str, tuple[int, int, int]] = collect_root_first_positions(
        normalized_prefix, defined_only_name_keys, all_names, sheet_order
    )

    return {
        "containers": containers,
        "user_provided_containers": user_provided_containers,
        "normalized_prefix": normalized_prefix,
        "all_names": all_names,
        "defined_only_name_keys": defined_only_name_keys,
        "all_name_keys": all_name_keys,
        "gen_map": gen_map,
        "excluded_indexed_field_names": excluded_indexed_field_names,
        "expected_field_shape": expected_field_shape,
        "numeric_root_keys": numeric_root_keys,
        "arrays_with_double_index": arrays_with_double_index,
        "container_parent_names": container_parent_names,
        "container_parents_with_children": container_parents_with_children,
        "group_to_root": group_to_root,
        "anchor_names": anchor_names,
        "group_labels": group_labels,
        "sheet_order": sheet_order,
        "root_first_pos": root_first_pos,
    }


def _iterate_and_fill_entries(
    *,
    wb,
    schema: Optional[Dict[str, Any]],
    array_transform_rules: Optional[Dict[str, List[ArrayTransformRule]]],
    prefix: str,
    root_result: Dict[str, Any],
    state: Dict[str, Any],
    safe_insert,
    user_provided_containers: bool,
) -> None:
    """定義名/生成名を走査し、値の取得→変換→配分→挿入までを一括処理。"""
    # ルートキーの事前挿入で順序を安定化
    preseed_root_keys(root_result=root_result, root_first_pos=state["root_first_pos"])

    entries = collect_entries_in_sheet_order(
        all_names=state["all_names"],
        normalized_prefix=state["normalized_prefix"],
        excluded_indexed_field_names=state["excluded_indexed_field_names"],
        sheet_order=state["sheet_order"],
        suppress_ctx={
            "all_name_keys": state["all_name_keys"],
            "container_parent_names": state["container_parent_names"],
            "container_parents_with_children": state["container_parents_with_children"],
            "group_labels": state["group_labels"],
            "root_first_pos": state["root_first_pos"],
        },
    )

    for _pos, name, defined_name, original_path_keys in entries:
        path_keys = original_path_keys.copy()

        if schema is not None:
            schema_path_keys, schema_broken = resolve_path_keys_with_schema(
                path_keys=path_keys, schema=schema
            )
            if not schema_broken:
                path_keys = schema_path_keys

        _skip, value = get_value_for_defined_or_generated_name(
            wb=wb, name=name, defined_name=defined_name, gen_map=state["gen_map"]
        )
        if _skip:
            continue

        # 末端（スカラー）で None の場合は空プレースホルダに変換して形状復元の対象にする
        # （同階層に有効データがあるときに null として出力されるようにするため）
        if value is None or (not isinstance(value, (list, dict)) and DataCleaner.is_empty_value(value)):
            value = ""

        # 1. 変換ルール
        if try_apply_transform_and_insert(
            array_transform_rules=array_transform_rules,
            path_keys=path_keys,
            original_path_keys=original_path_keys,
            value=value,
            workbook=wb,
            root_result=root_result,
            prefix=prefix,
            safe_insert=safe_insert,
        ):
            continue

        logger.debug(f"配列化後の値: {value}")

        # 2. 配列パス要素の処理（array.i.*）
        if process_array_path_entry(
            wb=wb,
            defined_name=defined_name,
            value=value,
            path_keys=path_keys,
            name=name,
            original_path_keys=original_path_keys,
            normalized_prefix=state["normalized_prefix"],
            root_result=root_result,
            arrays_with_double_index=state["arrays_with_double_index"],
            expected_field_shape=state["expected_field_shape"],
            gen_map=state["gen_map"],
            group_labels=state["group_labels"],
            all_name_keys=state["all_name_keys"],
            container_parent_names=state["container_parent_names"],
            defined_only_name_keys=state["defined_only_name_keys"],
            safe_insert=safe_insert,
            user_provided_containers=user_provided_containers,
        ):
            continue

        # 3. ルート配下の正規化（配列パスで未処理の場合に実行）
        normalized_keys = _normalize_root_group_for_parse(
            path_keys,
            group_to_root=state["group_to_root"],
            numeric_root_keys=state["numeric_root_keys"],
            root_result=root_result,
        )
        safe_insert(
            root_result,
            normalized_keys,
            value,
            ".".join(normalized_keys),
            name,
            original_path_keys,
            normalized_keys,
        )


def _finalize_result(
    *,
    result: Dict[str, Any],
    prefix: str,
    state: Dict[str, Any],
    array_transform_rules: Optional[Dict[str, List[ArrayTransformRule]]],
    user_provided_containers: bool,
    containers: Optional[Dict[str, Dict]],
) -> Dict[str, Any]:
    """ポストパース整形パイプラインを適用し、最終出力を返す。"""
    return apply_post_parse_pipeline(
        result=result,
        root_first_pos=state["root_first_pos"],
        prefix=prefix,
        user_provided_containers=user_provided_containers,
        containers=containers,
        array_transform_rules=array_transform_rules,
        normalized_prefix=state["normalized_prefix"],
        group_labels=state["group_labels"],
        group_to_root=state["group_to_root"],
        gen_map=state["gen_map"],
    )


def parse_named_ranges_with_prefix(
    xlsx_path: Path,
    prefix: str,
    array_split_rules: Optional[Dict[str, List[str]]] = None,
    array_transform_rules: Optional[Dict[str, List[ArrayTransformRule]]] = None,
    containers: Optional[Dict[str, Dict]] = None,
    schema: Optional[Dict[str, Any]] = None,
    global_max_elements: Optional[int] = None,
    extraction_policy: Optional[ExtractionPolicy] = None,
) -> Dict[str, Any]:
    """
    Excel 名前付き範囲(prefix) を解析してネスト dict/list を返す。
    prefixはデフォルトで"json"。
    array_split_rules: 配列化設定の辞書 {path: [delimiter1, delimiter2, ...]}
    array_transform_rules: 配列変換設定の辞書 {path: ArrayTransformRule}
    extraction_policy: 抽出時の共通ポリシー（未指定時は既定の現行仕様を適用）
    """
    # 文字列/PathLike を Path に正規化
    xlsx_path = Path(xlsx_path)
    if not xlsx_path:
        raise ValueError(
            "xlsx_pathは有効なPathオブジェクトまたは文字列パスである必要があります。"
        )

    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excelファイルが見つかりません: {xlsx_path}")

    if not xlsx_path.is_file():
        raise ValueError(f"指定されたパスはファイルではありません: {xlsx_path}")

    if not prefix:
        raise ValueError("prefixは空ではない文字列である必要があります。")

    try:
        wb = load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        raise ValueError(f"Excelファイルの読み込みに失敗しました: {xlsx_path} - {e}")

    # ポリシー決定（未指定なら既定）
    policy = extraction_policy or _DEFAULT_EXTRACTION_POLICY

    # 事前準備を一括計算
    _state = _prepare_parsing_prelude(
        wb=wb,
        prefix=prefix,
        containers=containers,
        global_max_elements=global_max_elements,
        extraction_policy=policy,
    )
    containers = _state["containers"]
    user_provided_containers = _state["user_provided_containers"]

    result: Dict[str, Any] = {}
    # 明示コンテナ指定時のみ prefix 配下に格納（自動推論時は従来通りトップ直下）
    root_result: Dict[str, Any] = (
        result if not user_provided_containers else result.setdefault(prefix, {})
    )

    if array_split_rules is None:
        array_split_rules = {}
    if array_transform_rules is None:
        array_transform_rules = {}

    # デバッグ支援: 挿入時の文脈を付与する安全ラッパー
    def _safe_insert(
        target_root: Union[Dict[str, Any], List[Any]],
        keys: List[str],
        val: Any,
        full_path_hint: str,
        context_name: str,
        original_keys: List[str],
        normalized_keys: List[str],
    ):
        try:
            insert_json_path(target_root, keys, val, full_path_hint)
        except Exception as e:
            raise type(e)(
                f"insert_json_pathで例外発生: name='{context_name}' "
                f"original_keys={original_keys} normalized_keys={normalized_keys} "
                f"target_type={type(target_root).__name__} keys={keys} full_path_hint='{full_path_hint}': {e}"
            ) from e

    # エントリ走査と挿入
    _iterate_and_fill_entries(
        wb=wb,
        schema=schema,
        array_transform_rules=array_transform_rules,
        prefix=prefix,
        root_result=root_result,
        state=_state,
        safe_insert=_safe_insert,
        user_provided_containers=user_provided_containers,
    )

    # パース後の出力整形
    result = _finalize_result(
        result=result,
        prefix=prefix,
        state=_state,
        array_transform_rules=array_transform_rules,
        user_provided_containers=user_provided_containers,
        containers=containers,
    )

    return result


# =============================================================================
# File Operations
# =============================================================================


def _has_non_empty_content(obj: Any) -> bool:
    """オブジェクト全体に空でないコンテンツが含まれているかチェック"""
    if obj is None:
        return False

    if is_json_dict(obj):
        return any(_has_non_empty_content(value) for value in obj.values())

    if is_json_list(obj):
        return any(_has_non_empty_content(item) for item in obj)

    # スカラー値は空でないと判定
    return not DataCleaner.is_empty_value(obj)


def _is_empty_container(obj: Any) -> bool:
    """空のコンテナ（辞書・リスト）かどうかをチェック"""
    return (is_json_list(obj) or is_json_dict(obj)) and len(obj) == 0


def prune_empty_elements(obj: Any, *, _has_sibling_data: Optional[bool] = None, schema: Optional[Dict[str, Any]] = None) -> Any:
    """
    再帰的に dict/list から空要素を除去する

    Args:
        obj: 処理対象のオブジェクト
        _has_sibling_data: 内部用パラメータ（外部使用禁止）

    Returns:
        空要素を除去した結果:
        - dict: 空要素のみの場合は {} または None
        - list: 空要素のみの場合は []
        - その他: そのまま返す
    """
    # 最初の呼び出し時のみ、全体的な非空データの存在をチェック
    if _has_sibling_data is None:
        _has_sibling_data = _has_non_empty_content(obj)

    # 型に応じた処理
    if is_json_dict(obj):
        return _prune_dict(obj, _has_sibling_data, schema=schema)
    elif is_json_list(obj):
        return _prune_list(obj, _has_sibling_data, schema=schema)
    else:
        return obj


def _prune_dict(obj: dict, has_sibling_data: bool, *, schema: Optional[Dict[str, Any]] = None) -> Optional[dict]:
    """辞書の空要素を除去"""
    if not obj:
        return {}

    # 子要素をプルーニング（Walrus演算子とdict内包表記を使用）
    props = schema.get("properties", {}) if isinstance(schema, dict) else {}
    pruned_items: Dict[str, Any] = {}
    for key, value in obj.items():
        sub_schema = props.get(key) if isinstance(props, dict) else None
        pruned_value = prune_empty_elements(value, _has_sibling_data=has_sibling_data, schema=sub_schema)
        if pruned_value is not None:
            pruned_items[key] = pruned_value

    if not pruned_items:
        return None

    # 非空要素の存在をチェック（ジェネレータ式を使用）
    has_non_empty = any(
        not _is_empty_container(value) for value in pruned_items.values()
    )
    # スキーマが array/object を要求しているキーがあり、値が空配列/空オブジェクトでも保持したい場合
    keep_due_to_schema = False
    if isinstance(props, dict):
        for k, v in pruned_items.items():
            t = None
            sub = props.get(k)
            if isinstance(sub, dict):
                t = sub.get("type")
            if t == "array" and is_json_list(v) and len(v) == 0:
                keep_due_to_schema = True
                break
            if t == "object" and is_json_dict(v) and len(v) == 0:
                keep_due_to_schema = True
                break

    # 三項演算子を使用（読みやすさのため分割）
    if has_non_empty or keep_due_to_schema:
        return pruned_items
    else:
        return {} if has_sibling_data else None


def _prune_list(obj: list, has_sibling_data: bool, *, schema: Optional[Dict[str, Any]] = None) -> list:
    """リストの空要素を除去"""
    # 空でない要素のみを残す（Walrus演算子とリスト内包表記を使用）
    item_schema = schema.get("items") if isinstance(schema, dict) else None
    pruned_items = []
    for item in obj:
        pruned_item = prune_empty_elements(item, _has_sibling_data=has_sibling_data, schema=item_schema)
        if pruned_item is not None:
            pruned_items.append(pruned_item)

    if not pruned_items:
        return []

    # 非空要素の存在をチェック（ジェネレータ式を使用）
    has_non_empty = any(not _is_empty_container(item) for item in pruned_items)

    # 三項演算子を使用
    return pruned_items if has_non_empty else []


def _restore_shapes_tree(original_after_prune: Any, data: Any, schema: Optional[Dict[str, Any]]) -> Any:
    """完全空のフィールド形状（[],{} ,None）を兄弟データの有無に応じて復元する。"""
    def _is_completely_empty(v: Any) -> bool:
        return is_completely_empty(v)

    def _empty_shape_for(value: Any) -> Any:
        if isinstance(value, list):
            return []
        if isinstance(value, dict):
            return {}
        return None

    def _restore_shapes(orig: Any, cur: Any, schema_here: Optional[Dict[str, Any]] = None) -> Any:
        if not isinstance(orig, (dict, list)):
            return cur
        if isinstance(orig, dict) and isinstance(cur, dict):
            has_non_empty_sibling = any(not _is_completely_empty(v) for v in orig.values())
            props = None
            if isinstance(schema_here, dict):
                maybe_props = schema_here.get("properties")
                if isinstance(maybe_props, dict):
                    props = maybe_props
            for k, v in orig.items():
                sub_schema = props.get(k) if isinstance(props, dict) else None
                if k not in cur:
                    if _is_completely_empty(v):
                        # 兄弟に非空がある場合は従来通り復元
                        # 兄弟がすべて空でも、スキーマが与えられている/該当プロパティが定義されていれば復元（パターン②）
                        if has_non_empty_sibling or (sub_schema is not None or schema_here is not None):
                            # スキーマから形状推定（array->[], object->{}）、なければ元の型から
                            base: Any
                            if isinstance(sub_schema, dict) and sub_schema.get("type") == "array":
                                base = []
                            elif isinstance(sub_schema, dict) and sub_schema.get("type") == "object":
                                base = {}
                            else:
                                base = _empty_shape_for(v)
                            # 再帰的に内側も復元（スキーマを渡して再帰埋め）
                            cur[k] = _restore_shapes(v, base, sub_schema)
                else:
                    cur[k] = _restore_shapes(v, cur[k], sub_schema)
            return cur
        if isinstance(orig, list) and isinstance(cur, list):
            restored_list = []
            for i, cur_item in enumerate(cur):
                if i < len(orig):
                    item_schema = None
                    if isinstance(schema_here, dict):
                        it = schema_here.get("items")
                        if isinstance(it, dict):
                            item_schema = it
                    restored_list.append(_restore_shapes(orig[i], cur_item, item_schema))
                else:
                    restored_list.append(cur_item)
            cur[:] = restored_list
            return cur
        return cur

    return _restore_shapes(original_after_prune, data, schema)


def _validate_and_log_errors(
    *,
    data: Dict[str, Any],
    schema: Optional[Dict[str, Any]],
    validator: Optional[Draft7Validator],
    validation_policy: ValidationPolicy,
    serialization_policy: SerializationPolicy,
    output_dir: Path,
    base_name: str,
) -> None:
    if not (validator and validation_policy.enabled):
        return
    data_for_validation = cast(Dict[str, Any], to_iso_for_validation(data) if validation_policy.to_iso_for_validation else data)
    errors = list(validator.iter_errors(data_for_validation))
    if not errors:
        return
    log_file = output_dir / f"{base_name}.error.log"
    log_file.parent.mkdir(parents=True, exist_ok=True)
    with open(log_file, "w", encoding="utf-8") as f:
        for error in errors:
            path_str = ".".join(str(p) for p in error.absolute_path)
            msg = f"Validation error at {path_str}: {error.message}\n"
            f.write(msg)
    first_error = errors[0]
    logger.error(f"Validation error: {first_error.message}")


def _dump_to_file(*, data: Dict[str, Any], output_path: Path, output_format: str, sp: SerializationPolicy) -> None:
    def json_default(obj):
        if sp.datetime_to_iso:
            if isinstance(obj, datetime.datetime):
                return obj.isoformat()
            if isinstance(obj, datetime.date):
                return obj.isoformat()
        return str(obj)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    eff_format = (sp.format or output_format).lower()
    if eff_format == "yaml" or output_format == "yaml":
        with output_path.open("w", encoding="utf-8") as f:
            yaml_data = json.loads(json.dumps(data, default=json_default))
            yaml.dump(yaml_data, f, default_flow_style=False, allow_unicode=True, indent=2)
    else:
        with output_path.open("w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=sp.ensure_ascii, indent=sp.indent, default=json_default)


def write_data(
    data: Dict[str, Any],
    output_path: Path,
    output_format: str = "json",
    schema: Optional[Dict[str, Any]] = None,
    validator: Optional[Draft7Validator] = None,
    suppress_empty: bool = True,
    *,
    cleaning_policy: DataCleaningPolicy | None = None,
    ordering_policy: OutputOrderingPolicy | None = None,
    serialization_policy: SerializationPolicy | None = None,
    validation_policy: ValidationPolicy | None = None,
    logging_policy: LoggingPolicy | None = None,
) -> None:
    """
    データをファイルに書き出し（JSON/YAML対応）。
    バリデーションとソートはオプション。
    """
    base_name = output_path.stem
    output_dir = output_path.parent

    # デバッグ: 出力前の一部構造確認（特定キー名に依存しない汎用ログ）
    sample_keys = list(data.keys())[:5] if is_json_dict(data) else []
    logger.debug("BEFORE-PRUNE keys(sample)=%r", sample_keys)

    # ポリシー（デフォルト併用）
    _cp = cleaning_policy or DEFAULT_DATA_CLEANING_POLICY
    _sp = serialization_policy or DEFAULT_SERIALIZATION_POLICY
    _vp = validation_policy or DEFAULT_VALIDATION_POLICY
    _lp = logging_policy or DEFAULT_LOGGING_POLICY

    # 形状正規化
    if _cp.normalize_array_field_shapes:
        data = cast(Dict[str, Any], normalize_array_field_shapes(data))

    # 全フィールドが未設定の要素を除去
    if _cp.prune_empty_elements:
        data = prune_empty_elements(data, schema=schema)

    # 空値の除去（この時点では空キーは落ちる）
    original_after_prune = data  # 形状復元のため保持
    if _cp.clean_empty_values and suppress_empty:
        cleaned_data = clean_empty_values(original_after_prune, suppress_empty, schema=schema)
        if cleaned_data is None:
            data = {}
        else:
            data = cast(Dict[str, Any], cleaned_data)

    if suppress_empty:
        data = cast(Dict[str, Any], _restore_shapes_tree(original_after_prune, data, schema))

    # デバッグ: プルーニング後の構造確認（特定キー名に依存しない汎用ログ）
    sample_keys_after = list(data.keys())[:5] if is_json_dict(data) else []
    logger.debug("AFTER-PRUNE keys(sample)=%r", sample_keys_after)

    # 出力順ポリシーの適用
    if ordering_policy is None:
        ordering_policy = OutputOrderingPolicy(
            schema_first=bool(schema),
            align_sibling_list_of_dicts=True,
            keep_extras_in_insertion_order=True,
        )
    data = order_for_output(data, policy=ordering_policy, schema=schema)

    # バリデーション → エラーログ
    _validate_and_log_errors(
        data=data,
        schema=schema,
        validator=validator,
        validation_policy=_vp,
        serialization_policy=_sp,
        output_dir=output_dir,
        base_name=base_name,
    )

    # スキーマ順は order_for_output で適用済み（必要時）

    # ファイル書き出し
    _dump_to_file(data=data, output_path=output_path, output_format=output_format, sp=_sp)

    logger.debug(f"ファイルの出力に成功しました: {output_path}")


# =============================================================================
# Border Rectangle Detection
# =============================================================================


def has_border(worksheet, row, col, side):
    """
    指定セルの指定方向に罫線があるかチェック。
    自セルの辺に罫線があるか、または隣接セルの対応辺に罫線があれば True。
    """
    # メモ化（ワークブック処理中にクリアされる）
    # DummySheet のように title が無い場合はキャッシュしない（id 再利用などで誤検知を避ける）
    sheet_title = getattr(worksheet, "title", None)
    cache_enabled = sheet_title is not None
    cache_key = (id(worksheet), sheet_title, row, col, side)
    if cache_enabled:
        cached = border_cache().get(cache_key)
        if cached is not None:
            return cached
    # 自セル側
    cell = worksheet.cell(row=row, column=col)
    border = getattr(cell.border, side, None)
    if border is not None and border.style is not None:
        if cache_enabled:
            border_cache()[cache_key] = True
        return True

    # 隣接セル側
    adj_row, adj_col = row, col
    adj_side = side
    if side == "top":
        adj_row = row - 1
        adj_side = "bottom"
    elif side == "bottom":
        adj_row = row + 1
        adj_side = "top"
    elif side == "left":
        adj_col = col - 1
        adj_side = "right"
    elif side == "right":
        adj_col = col + 1
        adj_side = "left"
    if adj_row > 0 and adj_col > 0:
        try:
            # 境界チェックは緩めにし、DummySheet等でも必ずアクセスして評価できるようにする
            acell = worksheet.cell(row=adj_row, column=adj_col)
            ab = getattr(acell.border, adj_side, None)
            if ab is not None and getattr(ab, "style", None) is not None:
                if cache_enabled:
                    border_cache()[cache_key] = True
                return True
        except Exception:
            # ワークシート実装に依存せず安全にフォールバック
            pass
    if cache_enabled:
        border_cache()[cache_key] = False
    return False
def compute_scan_bounds_for_rect_detection(worksheet, cell_names_map=None):
    """四角形検出のスキャン範囲 (min_row, min_col, max_row, max_col) を返す。

    セル名マップがある場合は最小外接矩形に十分なマージンを付けて探索範囲を絞る。
    ない場合はワークシートの実質的な有効範囲（30行×30列を上限）に限定する。
    """
    if cell_names_map:
        named_rows = [row for row, _ in cell_names_map.keys()]
        named_cols = [col for _, col in cell_names_map.keys()]
        if not named_rows or not named_cols:
            # フォールバック: 極小範囲
            min_row = 1
            min_col = 1
            max_row = min(worksheet.max_row or 30, 30)
            max_col = min(worksheet.max_column or 30, 30)
        else:
            ws_max_row = worksheet.max_row or (max(named_rows) + 5)
            ws_max_col = worksheet.max_column or (max(named_cols) + 5)
            # 少し広めのマージン（現行仕様に合わせる）
            ROW_MARGIN = 20
            COL_MARGIN = 10
            min_row = max(1, min(named_rows) - ROW_MARGIN)
            min_col = max(1, min(named_cols) - COL_MARGIN)
            max_row = min(ws_max_row, max(named_rows) + ROW_MARGIN)
            max_col = min(ws_max_col, max(named_cols) + COL_MARGIN)
    else:
        # セル名がない場合は実際のワークシートの有効範囲内に制限
        min_row, min_col = 1, 1
        actual_max_row = worksheet.max_row if worksheet.max_row else 30
        actual_max_col = worksheet.max_column if worksheet.max_column else 30
        max_row, max_col = min(actual_max_row, 30), min(actual_max_col, 30)
    return min_row, min_col, max_row, max_col


def build_area_sorted_size_combinations(max_width: int, max_height: int) -> List[Tuple[int, int, int]]:
    """(area, width, height) の組を面積降順で生成する。

    巨大な領域から探索して最初に完全矩形を見つけたら採用、という既存戦略を維持。
    """
    size_combinations: List[Tuple[int, int, int]] = []
    for width in range(1, max_width + 1):
        for height in range(1, max_height + 1):
            area = width * height
            size_combinations.append((area, width, height))
    size_combinations.sort(reverse=True)
    return size_combinations


def detect_regions_from_anchors(
    worksheet, min_row: int, min_col: int, max_row: int, max_col: int, cell_names_map
) -> List[Tuple[int, int, int, int, float]]:
    """セル名アンカーから右端候補→下端確定で矩形を検出（1アンカー最大1件）。"""
    regions: List[Tuple[int, int, int, int, float]] = []
    anchors = sorted(set(cell_names_map.keys()))  # (row,col)
    for top, left in anchors:
        try:
            _sheet_name = getattr(worksheet, "title", "")
        except Exception:
            _sheet_name = ""
        logger.debug("ANCHOR-RECT-CHECK sheet=%s top=%s left=%s", _sheet_name, top, left)
        # 起点要件: 左上の外枠が存在
        if not (
            has_border(worksheet, top, left, "top")
            and has_border(worksheet, top, left, "left")
        ):
            continue
        r_limit = max_row
        c_limit = max_col
        # 右端の最大候補: top 行で連続する上辺ボーダーがある範囲
        right_max = left
        c = left
        while c <= c_limit and has_border(worksheet, top, c, "top"):
            right_max = c
            c += 1
        # 右端候補を広い方から狭めて最初に閉じる矩形を採用
        for right in range(right_max, left - 1, -1):
            cand = _find_rect_from_anchor(
                worksheet, left, right, top, max_bottom=r_limit
            )
            if not cand:
                continue
            l2, t2, r2, b2 = cand
            comp = calculate_border_completeness(worksheet, t2, l2, b2, r2)
            logger.debug(
                "ANCHOR-RECT-CAND sheet=%s bounds top=%s left=%s bottom=%s right=%s completeness=%.3f",
                _sheet_name,
                t2,
                l2,
                b2,
                r2,
                comp,
            )
            if comp >= 1.0:
                # 領域内に少なくとも1つセル名が含まれること
                cell_names_in_region = get_cell_names_in_region(
                    cell_names_map, t2, l2, b2, r2
                )
                if not cell_names_in_region:
                    continue
                regions.append((t2, l2, b2, r2, comp))
                break  # このアンカーでは最大の1つを採用
    return regions


def detect_regions_bruteforce(
    worksheet, min_row: int, min_col: int, max_row: int, max_col: int, cell_names_map=None
) -> List[Tuple[int, int, int, int, float]]:
    """各セルを左上起点として全探索で矩形を検出。"""
    regions: List[Tuple[int, int, int, int, float]] = []
    for top in range(min_row, max_row + 1):
        for left in range(min_col, max_col + 1):
            if not (
                has_border(worksheet, top, left, "top")
                and has_border(worksheet, top, left, "left")
            ):
                continue
            max_width = min(max_col - left + 1, max_col - min_col + 1)
            max_height = min(max_row - top + 1, max_row - min_row + 1)
            size_combinations = build_area_sorted_size_combinations(
                max_width, max_height
            )
            for _area, width, height in size_combinations:
                right = left + width - 1
                bottom = top + height - 1
                if bottom > max_row or right > max_col:
                    continue
                completeness = calculate_border_completeness(
                    worksheet, top, left, bottom, right
                )
                if completeness >= 1.0:
                    if cell_names_map:
                        cell_names_in_region = get_cell_names_in_region(
                            cell_names_map, top, left, bottom, right
                        )
                        if not cell_names_in_region:
                            continue
                    regions.append((top, left, bottom, right, completeness))
    return regions


def dedup_and_sort_regions(
    regions: List[Tuple[int, int, int, int, float]]
) -> List[Tuple[int, int, int, int, float]]:
    """同一座標を completeness 最大で残し、大きい順→完成度→左上でソート。"""
    uniq: dict[tuple[int, int, int, int], float] = {}
    for t, left, b, r, c in regions:
        key = (t, left, b, r)
        if key not in uniq or c > uniq[key]:
            uniq[key] = c
    regions = [
        (t, left, b, r, uniq[(t, left, b, r)]) for (t, left, b, r) in uniq.keys()
    ]
    regions.sort(
        key=lambda r: (-(r[2] - r[0] + 1) * (r[3] - r[1] + 1), -r[4], r[0], r[1])
    )
    return regions


def filter_overlapping_regions(
    regions: List[Tuple[int, int, int, int, float]]
) -> List[Tuple[int, int, int, int, float]]:
    """包含関係で冗長な小領域を除外し、大きい領域を優先して保持。"""
    filtered: List[Tuple[int, int, int, int, float]] = []
    for region in regions:
        top, left, bottom, right, completeness = region
        region_area = (bottom - top + 1) * (right - left + 1)
        is_redundant = False
        for existing_region in list(filtered):
            ex_top, ex_left, ex_bottom, ex_right, _ex_comp = existing_region
            ex_area = (ex_bottom - ex_top + 1) * (ex_right - ex_left + 1)
            # 既存の大きな領域に完全包含される小領域は除外
            if (
                top >= ex_top
                and left >= ex_left
                and bottom <= ex_bottom
                and right <= ex_right
                and region_area < ex_area
            ):
                is_redundant = True
                break
            # 新領域が既存の小領域を包含するなら既存を削除
            if (
                ex_top >= top
                and ex_left >= left
                and ex_bottom <= bottom
                and ex_right <= right
                and ex_area < region_area
            ):
                filtered.remove(existing_region)
        if not is_redundant:
            filtered.append(region)
    return filtered


def detect_rectangular_regions(worksheet, cell_names_map=None):
    """
    罫線で囲まれた四角形領域を検出
    左上から大きい順にソートして返す
    """
    regions: List[Tuple[int, int, int, int, float]] = []

    # スキャン対象範囲を計算（セル名マップの有無で分岐）
    min_row, min_col, max_row, max_col = compute_scan_bounds_for_rect_detection(
        worksheet, cell_names_map
    )

    lr, ur = (min(min_row, max_row), max(min_row, max_row))
    lc, uc = (min(min_col, max_col), max(min_col, max_col))
    logger.debug(f"四角形検出範囲: 行{lr}-{ur}, 列{lc}-{uc}")

    # まず、セル名マップがある場合は『名前付きセルをアンカーにした高速検出』を優先
    if cell_names_map:
        regions.extend(
            detect_regions_from_anchors(
                worksheet, min_row, min_col, max_row, max_col, cell_names_map
            )
        )

    # 各セルを起点として四角形を検出（大きい領域から小さい領域へ）
    # 各セルを起点として四角形を検出（大きい領域から小さい領域へ）
    regions.extend(
        detect_regions_bruteforce(
            worksheet, min_row, min_col, max_row, max_col, cell_names_map
        )
    )

    # 重複を除去してから、大きい順、完成度順、左上位置順でソート
    regions = dedup_and_sort_regions(regions)

    # 重複する領域を除去（より意味のある大きな領域を優先）
    filtered_regions = filter_overlapping_regions(regions)
    logger.debug(f"検出された四角形領域数: {len(filtered_regions)}")
    return filtered_regions


def _row_has_horizontal_border(ws, row: int, left: int, right: int, side: str) -> bool:
    """指定行の区間で水平ボーダーが連続しているか（キャッシュ活用）。"""
    for c in range(left, right + 1):
        if not has_border(ws, row, c, side):
            return False
    return True


def _col_has_vertical_border(ws, col: int, top: int, bottom: int, side: str) -> bool:
    """指定列の区間で垂直ボーダーが連続しているか（キャッシュ活用）。"""
    for r in range(top, bottom + 1):
        if not has_border(ws, r, col, side):
            return False
    return True


def find_bordered_region_around_positions(
    worksheet,
    positions: dict[str, tuple[int, int]],
    *,
    row_margin: int = 12,
    col_margin: int = 8,
) -> tuple[int, int, int, int] | None:
    """
    基準座標群（フィールド -> (col,row)）の最小外接矩形から外側へ拡張し、
    四辺が罫線で閉じた領域をヒューリスティックに検出する高速版。
    失敗時は None を返す。
    """
    if not positions:
        return None
    cols = [c for (c, _r) in positions.values()]
    rows = [r for (_c, r) in positions.values()]
    if not cols or not rows:
        return None
    ws = worksheet
    max_r = getattr(ws, "max_row", 200) or 200
    max_c = getattr(ws, "max_column", 50) or 50

    left = max(1, min(cols) - col_margin)
    right = min(max_c, max(cols) + col_margin)
    top = max(1, min(rows) - row_margin)
    bottom = min(max_r, max(rows) + row_margin)

    # 上下の水平ボーダーを探す
    steps = 0
    while (
        top > 1
        and not _row_has_horizontal_border(ws, top, left, right, "top")
        and steps < row_margin * 2
    ):
        top -= 1
        steps += 1
    steps = 0
    while (
        bottom < max_r
        and not _row_has_horizontal_border(ws, bottom, left, right, "bottom")
        and steps < row_margin * 2
    ):
        bottom += 1
        steps += 1

    # 左右の垂直ボーダーを探す
    steps = 0
    while (
        left > 1
        and not _col_has_vertical_border(ws, left, top, bottom, "left")
        and steps < col_margin * 2
    ):
        left -= 1
        steps += 1
    steps = 0
    while (
        right < max_c
        and not _col_has_vertical_border(ws, right, top, bottom, "right")
        and steps < col_margin * 2
    ):
        right += 1
        steps += 1

    # 最終確認
    ok = (
        _row_has_horizontal_border(ws, top, left, right, "top")
        and _row_has_horizontal_border(ws, bottom, left, right, "bottom")
        and _col_has_vertical_border(ws, left, top, bottom, "left")
        and _col_has_vertical_border(ws, right, top, bottom, "right")
    )
    return (top, left, bottom, right) if ok else None


def get_cell_names_in_region(cell_names_map, top, left, bottom, right):
    """指定領域内のセル名を取得"""
    cell_names = []
    for row in range(top, bottom + 1):
        for col in range(left, right + 1):
            if (row, col) in cell_names_map:
                cell_names.append(cell_names_map[(row, col)])
    return cell_names


# 罫線矩形スキャン（ネスト処理用の最小ヘルパー群を再導入）
def _find_rect_from_anchor(ws, left, right, top, max_bottom=None):
    """(left, right, top) を与えて、下方向に罫線が閉じる bottom を探す。
    返り値は (left, top, right, bottom)。max_bottom を超える場合は None。
    厳密に左右幅固定・横ズレ許容なし。
    """
    # 上辺の連続性チェック（left..right の全列で top の上辺があること）
    for c in range(left, right + 1):
        if not has_border(ws, top, c, "top"):
            return None

    # 探索上限
    hard_limit = (
        max_bottom if max_bottom is not None else (getattr(ws, "max_row", 200) or 200)
    )

    # 下方向に走査し、左右辺が連続し、かつ下辺が閉じている最初の bottom を採用
    for bottom in range(top, hard_limit + 1):
        # 左右辺の連続性
        ok_vertical = True
        for r in range(top, bottom + 1):
            if not has_border(ws, r, left, "left") or not has_border(
                ws, r, right, "right"
            ):
                ok_vertical = False
                break
        if not ok_vertical:
            continue

        # 候補 bottom の下辺が全列で存在するか
        ok_bottom = True
        for c in range(left, right + 1):
            if not has_border(ws, bottom, c, "bottom"):
                ok_bottom = False
                break
        if ok_bottom:
            return (left, top, right, bottom)

    return None


def _scan_rects_seq(ws, left, right, top, col_tolerance=0, max_bottom=None):
    """(left,right,top) を基準に、同幅で縦に連なる矩形を上から順に検出。
    返り値は [(left, top, right, bottom), ...]
    col_tolerance は列ズレ許容（0 で厳密一致）。
    """
    rects: List[Tuple[int, int, int, int]] = []
    base = _find_rect_from_anchor(ws, left, right, top, max_bottom=max_bottom)
    if not base:
        return rects
    b_left, b_top, b_right, b_bottom = base
    rects.append(base)

    cur_top = b_bottom + 1
    while True:
        if max_bottom is not None and cur_top > max_bottom:
            break
        cand = _find_rect_from_anchor(
            ws, b_left, b_right, cur_top, max_bottom=max_bottom
        )
        if not cand:
            break
        _l, _t, _r, _b = cand
        if col_tolerance == 0:
            if _l != b_left or _r != b_right:
                break
        else:
            if not (
                abs(_l - b_left) <= col_tolerance and abs(_r - b_right) <= col_tolerance
            ):
                break
        rects.append(cand)
        cur_top = _b + 1
    return rects


def _get_anchor_rects_naive(
    workbook, anchor_name: str, target_sheet: str, *, col_tolerance: int = 0
) -> List[Tuple[int, int, int, int]]:
    """
    指定アンカー名に対応する矩形列を、キャッシュを使わずに毎回スキャンして取得する。
    """
    # キャッシュ参照（ワークブック処理中有効）
    key = (id(workbook), target_sheet, anchor_name, col_tolerance)
    if key in anchor_rects_cache():
        return list(anchor_rects_cache()[key])
    pr_left = pr_right = pr_top = pr_bottom = None
    for sn, coord in iter_defined_name_destinations_all(anchor_name or "", workbook):
        if sn != target_sheet:
            continue
        coord_clean = coord.replace("$", "")
        if ":" in coord_clean:
            (sc, sr), (ec, er) = parse_range(coord_clean)
            pr_left, pr_top = sc, sr
            pr_right, pr_bottom = ec, er
        break
    rects: List[Tuple[int, int, int, int]] = []
    if pr_left and pr_right and pr_top:
        ws0 = (
            workbook[target_sheet]
            if target_sheet in getattr(workbook, "sheetnames", [])
            else workbook.active
        )
        rects = _scan_rects_seq(
            ws0, pr_left, pr_right, pr_top, col_tolerance=col_tolerance
        )
    # キャッシュ保存
    anchor_rects_cache()[key] = list(rects)
    return rects


def extract_cell_names_from_workbook(workbook, prefix: str = "json"):
    """ワークブックから名前付き範囲の <prefix>.* セル名を抽出"""
    cell_names_map = {}

    # 名前付き範囲から座標とセル名を取得
    for name, defined_name in workbook.defined_names.items():
        if name.startswith(f"{prefix}."):
            try:
                for sheet_name, coord in defined_name.destinations:
                    # 座標文字列を解析（例: "$S$2" -> (19, 2)）
                    coord_clean = coord.replace("$", "")
                    if ":" in coord_clean:
                        # 範囲の場合は左上のセルを使用
                        coord_clean = coord_clean.split(":")[0]

                    # 列文字を数値に変換
                    col_match = ""
                    row_match = ""
                    for char in coord_clean:
                        if char.isalpha():
                            col_match += char
                        elif char.isdigit():
                            row_match += char

                    if col_match and row_match:
                        col_num = column_index_from_string(col_match)
                        row_num = int(row_match)
                        cell_names_map[(row_num, col_num)] = name

            except Exception as e:
                logger.warning(f"名前付き範囲の解析に失敗: {name} - {e}")

    return cell_names_map


def extract_cell_names_for_sheet(workbook, sheet_name: str, prefix: str = "json"):
    """特定シートの <prefix>.* セル名を (row,col)->name で抽出"""
    cell_names_map = {}
    for name, defined_name in workbook.defined_names.items():
        if not name.startswith(f"{prefix}."):
            continue
        for sn, coord in defined_name.destinations:
            if sn != sheet_name:
                continue
            coord_clean = coord.replace("$", "")
            if ":" in coord_clean:
                coord_clean = coord_clean.split(":")[0]
            col_part = ""
            row_part = ""
            for ch in coord_clean:
                if ch.isalpha():
                    col_part += ch
                elif ch.isdigit():
                    row_part += ch
            if col_part and row_part:
                col_num = column_index_from_string(col_part)
                row_num = int(row_part)
                cell_names_map[(row_num, col_num)] = name
    return cell_names_map


def calculate_border_completeness(worksheet, top, left, bottom, right):
    """四角形の罫線完全度を計算（0.0-1.0）。直接辺の罫線のみを評価する。"""
    # 注意: テスト等で罫線が動的に変更されるケースに対応するため、
    # 完全度の再計算直前にキャッシュをクリアして整合性を担保する。
    # これにより本関数内ではキャッシュの恩恵は薄れるが、他の探索ロジックでは
    # has_border のメモ化が有効に働く（トレードオフ）。
    border_cache().clear()
    try:
        _sheet_name = getattr(worksheet, "title", "")
    except Exception:
        _sheet_name = ""
    logger.debug(
        "BORDER-COMP sheet=%s bounds top=%s left=%s bottom=%s right=%s",
        _sheet_name,
        top,
        left,
        bottom,
        right,
    )
    total_segments = 0
    bordered_segments = 0

    # 上辺をチェック
    for col in range(left, right + 1):
        total_segments += 1
        if has_border(worksheet, top, col, "top"):
            bordered_segments += 1

    # 下辺をチェック
    for col in range(left, right + 1):
        total_segments += 1
        if has_border(worksheet, bottom, col, "bottom"):
            bordered_segments += 1

    # 左辺をチェック
    for row in range(top, bottom + 1):
        total_segments += 1
        if has_border(worksheet, row, left, "left"):
            bordered_segments += 1

    # 右辺をチェック
    for row in range(top, bottom + 1):
        total_segments += 1
        if has_border(worksheet, row, right, "right"):
            bordered_segments += 1

    return bordered_segments / total_segments if total_segments > 0 else 0.0


def extract_field_names_from_pattern(pattern, prefix: str = "json"):
    """パターンからフィールド名を抽出"""
    parts = pattern.split(".")
    field_names = []

    for part in parts:
        if part != prefix and part != "*" and not part.isdigit():
            field_names.append(part)

    return field_names if field_names else ["value"]


def _replace_nth_from_end_numeric(key: str, n_from_end: int, new_index: int) -> str:
    """
    指定された文字列 `key` のドット区切り部分のうち、末尾から数えて `n_from_end` 番目の
    数値部分を `new_index` で置き換えます。
    """
    parts = key.split(".")
    idx_positions = [i for i, p in enumerate(parts) if p.isdigit()]
    if len(idx_positions) < n_from_end or n_from_end <= 0:
        return key
    target_pos = idx_positions[-n_from_end]
    parts[target_pos] = str(new_index)
    return ".".join(parts)


def _set_parent_index_in_key(key: str, parent_index: int) -> str:
    """
    指定されたキー文字列内の親インデックス部分を更新する。
    """
    parts = key.split(".")
    idx_positions = [i for i, p in enumerate(parts) if p.isdigit()]
    if len(idx_positions) < 2:
        return key
    parts[idx_positions[-2]] = str(parent_index)
    return ".".join(parts)


# =============================================================================
# Container Functions
# =============================================================================


def _parse_container_mapping(arg_text: str) -> Dict[str, Any]:
    """--container で渡された1オブジェクトを YAML としてパースする（JSONはYAMLのサブセット）。

    戻り値は dict（マッピング型）のみ許可。その他型なら ValueError。
    例外メッセージは既存テスト互換のため『無効なJSON形式』表現を維持。
    """
    try:
        obj = yaml.safe_load(arg_text)
        if isinstance(obj, dict):
            return obj
        raise ValueError(
            "--container はオブジェクト(JSON/YAMLのマッピング)である必要があります"
        )
    except Exception as e:
        # 互換のため JSON 文言を維持
        raise ValueError(f"無効なJSON形式: {e}")


def parse_container_args(container_args, config_containers=None):
    """CLIのcontainerオプションを解析し、設定ファイルとマージ（JSON/YAML対応）"""
    combined_containers = config_containers.copy() if config_containers else {}

    if not container_args:
        return combined_containers

    for container_arg in container_args:
        container_def = _parse_container_mapping(container_arg)
        combined_containers.update(container_def)
    return combined_containers


def validate_cli_containers(container_args, prefix: str = "json"):
    """CLIの--containerオプション専用検証（YAMLのみ。JSONはYAMLのサブセットとして解釈）"""
    for i, container_arg in enumerate(container_args):
        # 互換: 明らかに JSON 風だがコロンが無い等のケースは、既存テスト期待に合わせて JSON エラーを返す
        if "{" in container_arg and "}" in container_arg and ":" not in container_arg:
            raise ValueError(f"無効なJSON形式: 引数{i+1}")

        try:
            container_def = _parse_container_mapping(container_arg)
        except ValueError as e:
            # メッセージ互換
            raise ValueError(f"無効なJSON形式: 引数{i+1}: {e}")

        for cell_name in container_def.keys():
            if not cell_name.startswith(f"{prefix}."):
                raise ValueError(
                    f"セル名は'{prefix}.'で始まる必要があります: {cell_name}"
                )


def calculate_hierarchy_depth(cell_name, prefix: str = "json"):
    """数値インデックスを除外した階層深度を計算"""
    parts = cell_name.split(".")
    # 空の部分も除外し、数値でない部分のみを階層として扱う
    hierarchy_parts = [part for part in parts if part and not part.isdigit()]
    # 先頭がprefixならそれを除外してカウント
    if hierarchy_parts and hierarchy_parts[0] == prefix:
        depth = len(hierarchy_parts) - 1
    else:
        depth = len(hierarchy_parts)
    logger.debug(f"階層深度計算: {cell_name} -> 部分={hierarchy_parts} -> 深度={depth}")
    return depth


def validate_container_config(containers, prefix: str = "json"):
    """コンテナ設定の妥当性を検証"""
    errors: List[str] = []

    for container_name, container_def in containers.items():
        # セル名の形式チェック
        if not container_name.startswith(f"{prefix}."):
            errors.append(
                f"コンテナ名は'{prefix}.'で始まる必要があります: {container_name}"
            )

        # direction の検証
        if "direction" in container_def:
            direction = container_def["direction"]
            if direction not in ["row", "column"]:
                errors.append(
                    f"コンテナ{container_name}の'direction'は'row'または'column'である必要があります: {direction}"
                )

        # increment の検証
        if "increment" in container_def:
            increment = container_def["increment"]
            if not isinstance(increment, int) or increment < 0:
                errors.append(
                    f"コンテナ{container_name}の'increment'は0以上の整数である必要があります: {increment}"
                )

        # 親要素のincrement推奨値チェック
    if _is_parent_container(container_name, prefix=prefix):
        increment = container_def.get("increment", 0)
        if increment != 0:
            logger.warning(
                f"親要素のincrementは0推奨: {container_name} increment={increment}"
            )

    return errors


def _is_parent_container(container_key, prefix: str = "json"):
    """親要素かどうかを判定"""
    # re imported at top
    return re.match(rf"^{re.escape(prefix)}\.[^.]+$", container_key) is not None


def validate_hierarchy_consistency(containers, prefix: str = "json"):
    """コンテナの階層構造の整合性を検証"""
    errors: List[str] = []

    # 親子関係の検証（新仕様では基本的に自由だが、論理的な整合性をチェック）
    for container_name, container_def in containers.items():
        # 子要素の場合、対応する親要素が推奨される
        if _is_child_container(container_name, prefix=prefix):
            parent_name = _extract_parent_key_from_child(container_name, prefix=prefix)
            if parent_name and parent_name not in containers:
                logger.info(
                    f"子要素 {container_name} に対応する親要素 {parent_name} が未定義です（任意）"
                )

    # 循環参照の検証は新仕様では不要（階層が明確）
    return errors


def _is_child_container(container_key, prefix: str = "json"):
    """子要素かどうかを判定"""
    # re imported at top
    return re.match(rf"^{re.escape(prefix)}\.[^.]+\.\d+", container_key) is not None


def _extract_parent_key_from_child(child_key, prefix: str = "json"):
    """子要素キーから親要素キーを抽出"""
    # "json.orders.1.items.1" → "json.orders"
    parts = child_key.split(".")
    if len(parts) >= 3:
        parent_parts = []
        for part in parts[1:-1]:  # prefix と末尾数値を除外
            if not part.isdigit():
                parent_parts.append(part)

        if parent_parts:
            return f"{prefix}." + ".".join(parent_parts)

    return None


def filter_cells_with_names(
    range_coords: Iterable[Tuple[int, int]],
    cell_names: Mapping[Tuple[int, int], str],
    prefix: str = "json",
) -> Dict[Tuple[int, int], str]:
    """指定された範囲から、接頭辞に一致するセル名のみを抽出して返す。

    契約:
    - 入力: `range_coords` は (col, row) のタプル列。
            `cell_names` は (col, row) -> セル名 のマッピング。
            `prefix` はフィルタ対象の接頭辞（例: "json"）。
    - 出力: 条件に一致した (col, row) -> セル名 の辞書。
    - 例外: なし（不正な型は型チェックで検出される想定）。
    """
    result: Dict[Tuple[int, int], str] = {}
    for cell_coord in range_coords:
        cell_name = cell_names.get(cell_coord)
        if cell_name and cell_name.startswith(f"{prefix}."):
            result[cell_coord] = cell_name
    return result


def generate_cell_names_from_containers(
    containers,
    workbook,
    global_max_elements: Optional[int] = None,
    *,
    prefix: str = "json",
    extraction_policy: Optional[ExtractionPolicy] = None,
):
    """
    コンテナ定義からセル名を自動生成
    """
    generated_names: Dict[str, Any] = {}

    if not containers:
        return generated_names

    # コンテナ設定の妥当性を検証
    config_errors = validate_container_config(containers, prefix=prefix)
    if config_errors:
        for error in config_errors:
            logger.error(f"コンテナ設定エラー: {error}")
        return generated_names

    # 階層構造の整合性を検証
    hierarchy_errors = validate_hierarchy_consistency(containers, prefix=prefix)
    if hierarchy_errors:
        for error in hierarchy_errors:
            logger.error(f"階層構造エラー: {error}")

    # 各コンテナを処理（親→子の階層順に安定化）
    ordered_containers = sort_containers_by_hierarchy(containers, prefix=prefix)
    for container_name, container_def in ordered_containers:
        logger.debug(f"コンテナ処理開始: {container_name}")
        stats().containers_processed += 1

        try:
            process_container(
                container_name,
                container_def,
                workbook,
                generated_names,
                global_max_elements,
                extraction_policy=extraction_policy,
            )
        except Exception as e:
            # コンテナ単位の最上位でエラーを検出・記録（トレース出力）
            logger.exception(f"コンテナ {container_name} の処理中にエラー")

    logger.debug(f"生成されたセル名と値: {generated_names}")
    return generated_names
def enumerate_sheeted_ranges_sorted(workbook) -> list[tuple[Optional[str], Optional[str]]]:
    """ワークブックからシート名の列を取得し、ワークブック順に整列した (sheet, None) の配列を返す。"""
    sheeted_ranges: list[tuple[Optional[str], Optional[str]]] = []
    for title in list(getattr(workbook, "sheetnames", [])):
        try:
            if not title:
                continue
            sheeted_ranges.append((title, None))
        except Exception:
            continue
    order = {name: idx for idx, name in enumerate(getattr(workbook, "sheetnames", []))}
    sheeted_ranges.sort(key=lambda x: order.get(x[0], 10**6))
    return sheeted_ranges


def build_nameful_sheets_and_positions(
    container_name: str,
    workbook,
    sheeted_ranges: list[tuple[Optional[str], Optional[str]]],
) -> tuple[set[str], dict[str, dict[str, tuple[int, int]]]]:
    """直下フィールドの定義名から、シートごとの座標を解決し nameful シート集合と座標マップを返す。"""
    nameful_sheets: set[str] = set()
    per_sheet_positions: dict[str, dict[str, tuple[int, int]]] = {}
    for _sheet, _ in sheeted_ranges:
        sheet_field_names = get_cell_names_in_container_range(
            container_name, workbook, _sheet
        )
        cur_pos: dict[str, tuple[int, int]] = {}
        if sheet_field_names:
            if _sheet is not None:
                nameful_sheets.add(_sheet)
            for _field, _nm in sheet_field_names.items():
                _pos = get_cell_position_from_name(_nm, workbook, _sheet)
                if _pos:
                    cur_pos[_field] = _pos
            if cur_pos and _sheet is not None:
                per_sheet_positions[_sheet] = cur_pos
        logger.debug(
            "NAMEFUL-CHECK %s sheet=%s fields=%s",
            container_name,
            _sheet,
            sorted(list(cur_pos.keys())),
        )
    return nameful_sheets, per_sheet_positions


def select_template_positions(
    nameful_sheets: set[str],
    per_sheet_positions: dict[str, dict[str, tuple[int, int]]],
    sheeted_ranges: list[tuple[Optional[str], Optional[str]]],
) -> tuple[dict[str, tuple[int, int]], Optional[str]]:
    """nameful な最初のシートの座標群をテンプレートとして返す。"""
    template_positions: dict[str, tuple[int, int]] = {}
    template_sheet: Optional[str] = None
    for _sheet, _ in sheeted_ranges:
        if _sheet in nameful_sheets:
            template_positions = dict(per_sheet_positions.get(_sheet, {}))
            if template_positions:
                template_sheet = _sheet
                break
    return template_positions, template_sheet


# ---- Extracted helpers to reduce process_container complexity ----
def gather_current_positions_and_template(
    *,
    container_name: str,
    container_def: dict[str, Any],
    workbook,
    target_sheet: Optional[str],
    nameful_sheets: set[str],
    template_positions: dict[str, tuple[int, int]],
    template_sheet: Optional[str],
    global_field_names: dict[str, str] | None,
) -> tuple[dict[str, tuple[int, int]], bool]:
    """Resolve current positions for a sheet and optionally apply template composition.

    Returns (current_positions, using_template).
    Mirrors original inline logic in process_container.
    """
    using_template = False
    current_positions: dict[str, tuple[int, int]] = {}

    # Resolve positions from defined names on the target sheet
    base_cell_names = global_field_names or {}
    if base_cell_names:
        for field_name, original_cell_name in base_cell_names.items():
            pos = get_cell_position_from_name(original_cell_name, workbook, target_sheet)
            if pos:
                current_positions[field_name] = pos
        # Apply template positions if only one nameful sheet and no range on the container
        if (
            not current_positions
            and template_positions
            and template_sheet is not None
            and len(nameful_sheets) == 1
        ):
            _has_digit_token = any(part.isdigit() for part in container_name.split("."))
            if (
                (container_def.get("increment", 0) or 0) > 0
                and _has_digit_token
                and not container_def.get("range")
            ):
                using_template = True
                current_positions = dict(template_positions)
            else:
                # When no range, skip will be decided by caller; if range exists, we'll try composition below
                pass

    # If still empty, try composing from explicit range (with border completeness check)
    if not current_positions:
        rng = container_def.get("range")
        if rng and template_positions:
            try:
                rng_str = str(rng)
                rng_a1 = rng_str.split("!", 1)[1] if "!" in rng_str else rng_str
                (sc, sr), (ec, er) = parse_range(rng_a1.replace("$", ""))
                # Border completeness on the target sheet
                if target_sheet in getattr(workbook, "sheetnames", []):
                    ws_check = workbook[target_sheet]
                else:
                    ws_check = workbook.active
                comp = calculate_border_completeness(ws_check, sr, sc, er, ec)
                if comp >= 1.0:
                    target_top_left = (sc, sr)
                    (_tsc, _tsr), (_tec, _ter) = parse_range(rng_a1.replace("$", ""))
                    tpl_top_left = (_tsc, _tsr)
                    for f, (c, r) in template_positions.items():
                        dx = c - tpl_top_left[0]
                        dy = r - tpl_top_left[1]
                        current_positions[f] = (target_top_left[0] + dx, target_top_left[1] + dy)
                    using_template = True
            except Exception:
                current_positions = {}

    return current_positions, using_template


def estimate_element_count_and_step(
    *,
    container_name: str,
    container_def: dict[str, Any],
    workbook,
    target_sheet: Optional[str],
    current_positions: dict[str, tuple[int, int]],
    base_cell_names: dict[str, str] | None,
    global_max_elements: Optional[int],
    template_sheet: Optional[str],
    using_template: bool,
    direction: str,
    eff_increment: int,
    anchor_range_span: Optional[int],
    labels: list[str],
) -> tuple[int, Optional[int], Optional[int], Optional[int], int]:
    """Compute element_count and optional step_override for a container on a sheet.

    Returns (element_count, step_override, parent_range_span_for_container, range_count, internal_slice_count).
    """
    # Internal slice detection
    internal_slice_count = detect_internal_slice_count(base_cell_names, workbook)

    # Parent range span
    parent_range_span_for_container = compute_parent_range_span_for_container(
        container_name=container_name,
        workbook=workbook,
        target_sheet=target_sheet,
        direction=direction,
    )

    # Range-based count
    range_count = compute_range_count_from_explicit_range(
        container_def=container_def, direction=direction
    )

    # Label-first count analysis
    element_count = analyze_element_count_without_range(
        pos_map=current_positions,
        eff_increment=eff_increment,
        direction=direction,
        labels=labels,
        workbook=workbook,
        container_name=container_name,
        target_sheet=target_sheet,
        using_template=using_template,
    )

    # Skip template-applied sheets with zero detection unless range deterministically provides count
    if using_template and element_count == 0 and target_sheet != template_sheet:
        if not (isinstance(range_count, int) and range_count > 0):
            return 0, None, parent_range_span_for_container, range_count, internal_slice_count

    # Apply range limits
    if isinstance(range_count, int) and range_count > 0:
        if labels:
            if element_count > 0:
                element_count = min(element_count, range_count)
        else:
            element_count = range_count

    # Clip by parent range span
    if (
        isinstance(parent_range_span_for_container, int)
        and parent_range_span_for_container > 0
        and (element_count or 0) > 0
    ):
        element_count = min(element_count, parent_range_span_for_container)

    # Clip by anchor range span
    if (
        isinstance(anchor_range_span, int)
        and anchor_range_span > 0
        and (element_count or 0) > 0
    ):
        element_count = min(element_count, anchor_range_span)

    step_override: Optional[int] = None

    # Internal slice priority
    if internal_slice_count > 1:
        element_count = internal_slice_count
        step_override = 1
        # Global cap
        if isinstance(global_max_elements, int) and global_max_elements > 0:
            # next_index は呼び出し側で加算するため、ここでは全体上限のみ安全クリップは不可
            # 呼び出し側で追加のクリップを行う設計のため、この段階では不変
            pass
        # Clip by anchor range & parent range again
        if (
            isinstance(anchor_range_span, int)
            and anchor_range_span > 0
            and element_count > 0
        ):
            element_count = min(element_count, anchor_range_span)
        if (
            isinstance(parent_range_span_for_container, int)
            and parent_range_span_for_container > 0
            and element_count > 0
        ):
            element_count = min(element_count, parent_range_span_for_container)
        # Clip by bordered region bounds as safety
        try:
            if target_sheet is not None and current_positions:
                b2 = find_region_bounds_for_positions(
                    workbook, container_name, target_sheet, current_positions
                )
                if b2:
                    top2, left2, bottom2, right2 = b2
                    if direction == "row":
                        anchor_row2 = min((row for (col, row) in current_positions.values()))
                        cap2 = ((bottom2 - anchor_row2) // 1) + 1 if bottom2 >= anchor_row2 else 0
                    else:
                        anchor_col2 = min((col for (col, row) in current_positions.values()))
                        cap2 = ((right2 - anchor_col2) // 1) + 1 if right2 >= anchor_col2 else 0
                    if cap2 > 0:
                        element_count = min(element_count, cap2)
        except Exception:
            logger.debug("bounds clipping failed for container=%s sheet=%s", container_name, target_sheet, exc_info=True)

    element_count, step_override = _finalize_element_count_and_step(
        element_count=element_count,
        step_override=step_override,
        internal_slice_count=internal_slice_count,
        target_sheet=target_sheet,
        direction=direction,
        container_name=container_name,
        workbook=workbook,
        parent_range_span_for_container=parent_range_span_for_container,
        current_positions=current_positions,
        eff_increment=eff_increment,
        global_max_elements=global_max_elements,
        range_count=range_count,
        labels=labels,
        using_template=using_template,
    )

    return element_count, step_override, parent_range_span_for_container, range_count, internal_slice_count


def _finalize_element_count_and_step(
    *,
    element_count: int,
    step_override: Optional[int],
    internal_slice_count: int,
    target_sheet: Optional[str],
    direction: str,
    container_name: str,
    workbook,
    parent_range_span_for_container: Optional[int],
    current_positions: dict[str, tuple[int, int]],
    eff_increment: int,
    global_max_elements: Optional[int],
    range_count: Optional[int],
    labels: list[str],
    using_template: bool,
) -> tuple[int, Optional[int]]:
    """Apply final overrides and clipping to element_count and step_override.

    This helper consolidates the late-stage clips and deterministic overrides.
    Behavior is preserved from the original inline logic.
    """
    # Deterministic rectangle series override when labels and internal slice are not active
    _ec2, _step2 = compute_count_step_from_anchor_rect_series(
        labels=labels,
        internal_slice_count=internal_slice_count,
        target_sheet=target_sheet,
        direction=direction,
        container_name=container_name,
        workbook=workbook,
        parent_range_span_for_container=parent_range_span_for_container,
    )
    if isinstance(_ec2, int) and _ec2 > 0:
        element_count = _ec2
    if isinstance(_step2, int) and _step2 > 0:
        step_override = _step2

    # Clip by parent bottom for single numeric child
    element_count = clip_element_count_by_parent_bottom_for_single_numeric_child(
        container_name=container_name,
        workbook=workbook,
        target_sheet=target_sheet,
        current_positions=current_positions,
        eff_increment=eff_increment,
        step_override=step_override,
        element_count=element_count,
    )

    return element_count, step_override

# ---- Small, behavior-preserving helpers extracted from process_container ----
def compute_effective_increment_and_anchor_span(
    *,
    container_name: str,
    container_def: dict[str, Any],
    workbook,
    target_sheet: Optional[str],
    global_field_names: dict[str, str] | None,
) -> tuple[int, Optional[int]]:
    """Estimate the effective increment (step) and anchor range span for a container.

    Mirrors the original inline logic:
    - If the container itself ends with '.1', use its defined range height as increment.
    - Else, when increment is not specified (<=0) and fields hang under '<parent>.1.<field>',
      derive increment from that child anchor range height.
    Returns (eff_increment, anchor_range_span).
    """
    increment = int(container_def.get("increment", 0) or 0)
    eff_increment = increment
    anchor_range_span: Optional[int] = None

    try:
        # Case 1: container itself is an anchor '*.1'
        if container_name.endswith(".1"):
            for sn, coord in iter_defined_name_destinations_all(container_name, workbook):
                if sn != target_sheet or not coord:
                    continue
                coord_clean = str(coord).replace("$", "")
                if ":" in coord_clean:
                    (_sc, _sr), (_ec, _er) = parse_range(coord_clean)
                    eff_increment = max(1, abs(_er - _sr) + 1)
                    anchor_range_span = eff_increment
                else:
                    eff_increment = 1
                    anchor_range_span = 1
                break
        # Case 2: not specified and fields hang under '<parent>.1.<field>'
        if (eff_increment or 0) <= 0 and global_field_names:
            try:
                sample_field, sample_name = next(iter(global_field_names.items()))
            except StopIteration:
                sample_field, sample_name = None, None
            if sample_field and sample_name:
                parts = sample_name.split(".")
                if len(parts) >= 2 and parts[-1] == sample_field:
                    anchor_guess = ".".join(parts[:-1])
                    if anchor_guess.split(".")[-1].isdigit():
                        for sn, coord in iter_defined_name_destinations_all(anchor_guess, workbook):
                            if sn != target_sheet or not coord:
                                continue
                            cc = str(coord).replace("$", "")
                            if ":" in cc:
                                (_sc, _sr), (_ec, _er) = parse_range(cc)
                                eff_increment = max(1, abs(_er - _sr) + 1)
                                anchor_range_span = eff_increment
                            else:
                                eff_increment = 1
                                anchor_range_span = 1
                            break
    except Exception:
        # Fail-safe: keep computed defaults
        pass

    return int(eff_increment or 0), anchor_range_span


def resolve_labels_in_positions(labels: list[str] | None, pos_map: dict[str, tuple[int, int]] | None) -> list[str]:
    """与えられた `labels` のうち、現在の座標マップ `pos_map` に存在するフィールド名のみを返す。

    安全側で例外は握りつぶし、空リストを返す。
    """
    try:
        if not labels or not pos_map:
            return []
        return [lf for lf in labels if isinstance(lf, str) and lf in pos_map]
    except Exception:
        return []


# ---- Tiny helpers extracted for process_container readability ----
def _normalize_labels(container_def: dict[str, Any]) -> list[str]:
    try:
        return [str(x) for x in (container_def.get("labels", []) or []) if isinstance(x, str)]
    except Exception:
        return []


def _apply_global_max_cap(element_count: int, global_max_elements: Optional[int], next_index: int) -> int:
    try:
        if (
            isinstance(global_max_elements, int)
            and global_max_elements > 0
            and element_count > 0
        ):
            remaining = max(0, global_max_elements - (next_index - 1))
            return max(0, min(element_count, remaining))
    except Exception:
        pass
    return element_count


def _try_nested_scan_and_emit(
    *,
    workbook,
    target_sheet: Optional[str],
    container_name: str,
    direction: str,
    current_positions: dict[str, tuple[int, int]],
    labels: list[str],
    policy: ExtractionPolicy,
    generated_names: dict,
) -> bool:
    try:
        parts_cn = container_name.split(".")
        num_positions = [i for i, t in enumerate(parts_cn) if t.isdigit()]
        if len(num_positions) < 2:
            return False
        anchor_positions = num_positions
        anchor_names_chain = [".".join(parts_cn[: p + 1]) for p in anchor_positions]
        ends_with_numeric = parts_cn[-1].isdigit()
        parent_anchor = (
            anchor_names_chain[-2] if (ends_with_numeric and len(anchor_names_chain) >= 2)
            else anchor_names_chain[-1]
        )
        if target_sheet is None:
            return False
        ws0 = (
            workbook[target_sheet]
            if target_sheet in getattr(workbook, "sheetnames", [])
            else workbook.active
        )
        parent_rects, ancestors_rects_chain = build_parent_and_ancestor_rects_for_nested_scan(
            workbook=workbook,
            target_sheet=target_sheet,
            parent_anchor=parent_anchor,
            direction=direction,
            generated_names=generated_names,
            anchor_names_chain=anchor_names_chain,
            ends_with_numeric=ends_with_numeric,
        )
        params = NestedScanParams(
            workbook=workbook,
            target_sheet=target_sheet,
            ws0=ws0,
            container_name=container_name,
            direction=direction,
            current_positions=current_positions,
            labels=labels,
            policy=policy,
            parent_anchor=parent_anchor,
            parent_rects=parent_rects,
            ancestors_rects_chain=ancestors_rects_chain,
            generated_names=generated_names,
            num_positions=num_positions,
            ends_with_numeric=ends_with_numeric,
        )
        handled = scan_and_emit_nested_children(params=params)
        return bool(handled)
    except Exception as e:
        logger.debug("NESTED-ERR %s: %s", container_name, e, exc_info=True)
        return False


# Thin wrapper data classes and adapters were removed in favor of direct
# calls to the underlying functions to reduce unnecessary call depth and
# surface area. This keeps behavior identical while simplifying the codebase.


def detect_internal_slice_count(
    base_cell_names: dict[str, str] | None, workbook
) -> int:
    """index=1 のフィールドが複数セルの範囲を指す場合の内部スライス数を検出する。

    これらの範囲の長さの最大値を返す。既存の挙動を踏襲する。
    """
    internal_slice_count = 0
    try:
        for field_name, nm in (base_cell_names or {}).items():
            try:
                dn = workbook.defined_names[nm]
            except Exception:
                continue
            vals = get_named_range_values(workbook, dn)
            if isinstance(vals, list) and len(vals) > 1:
                internal_slice_count = max(internal_slice_count, len(vals))
    except Exception:
        internal_slice_count = 0
    return internal_slice_count


def compute_parent_range_span_for_container(
    *, container_name: str, workbook, target_sheet: Optional[str], direction: str
) -> Optional[int]:
    """親コンテナ（末尾が数値でない）が定義範囲を持つ場合、direction に応じて
    その高さ（row）または幅（col）を返す。なければ None を返す。
    """
    parent_range_span_for_container: Optional[int] = None
    try:
        tail_tok = container_name.split(".")[-1]
        if tail_tok and (not tail_tok.isdigit()):
            for sn, coord in iter_defined_name_destinations_all(container_name, workbook):
                if sn != target_sheet or not coord:
                    continue
                cc = str(coord).replace("$", "")
                if ":" in cc:
                    (sc0, sr0), (ec0, er0) = parse_range(cc)
                    if (direction or "row").lower() == "row":
                        parent_range_span_for_container = max(1, abs(er0 - sr0) + 1)
                    else:
                        parent_range_span_for_container = max(1, abs(ec0 - sc0) + 1)
                break
    except Exception:
        parent_range_span_for_container = None
    return parent_range_span_for_container


def compute_range_count_from_explicit_range(
    *, container_def: dict[str, Any], direction: str
) -> Optional[int]:
    """container_def に明示的な 'range' がある場合、direction に基づいて件数を算出する。
    該当しない場合やエラー時は None を返す。
    """
    try:
        rng = container_def.get("range")
        if not rng:
            return None
        rng_str = str(rng)
        rng_a1 = rng_str.split("!", 1)[1] if "!" in rng_str else rng_str
        (sc, sr), (ec, er) = parse_range(rng_a1.replace("$", ""))
        return detect_instance_count((sc, sr), (ec, er), direction)
    except Exception:
        return None


def analyze_element_count_without_range(
    *,
    pos_map: dict[str, tuple[int, int]] | None,
    eff_increment: int,
    direction: str,
    labels: list[str],
    workbook,
    container_name: str,
    target_sheet: Optional[str],
    using_template: bool,
) -> int:
    """
    ラベル優先の件数分析（range が無い場合の分析）。

    件数算出ポリシー:
    - increment(eff_increment) を最優先。
      - labels がある場合: 増分スキャンし、labels のいずれかが空(None/"")になった時点で停止（矩形は参照しない）。
      - labels が無い場合: 矩形上限内で幾何学的に最大件数を算出（アンカー基準）。
    - eff_increment <= 0 は 1 件固定（非繰り返し）。
    - 矩形は labels が無い場合の上限、または安全装置としてのみ採用。
    """
    try:
        if (eff_increment or 0) <= 0:
            return 1
        if not pos_map:
            return 0

        # 可能なら矩形境界を取得（labelsが無い場合の上限として使用）
        if target_sheet is None:
            return 0

        # 直接呼び出しに置換: 内部の薄いラッパー関数は不要なので除去
        bounds = find_region_bounds_for_positions(workbook, container_name, target_sheet, pos_map)
        if bounds:
            top, left, bottom, right = bounds
            # 全基準が矩形内であることを確認（矩形が見つかった場合のみ）
            for col, row in pos_map.values():
                if not (left <= col <= right and top <= row <= bottom):
                    bounds = None
                    break

        # アンカー（row: 最上段, column: 最左）を抽出
        anchor_col: Optional[int] = None
        anchor_row: Optional[int] = None
        for _fname, (col, row) in pos_map.items():
            if anchor_col is None:
                anchor_col, anchor_row = col, row
                continue
            if (direction or "row").lower() == "row":
                if anchor_row is None or row < anchor_row:
                    anchor_col, anchor_row = col, row
            else:
                if anchor_col is None or col < anchor_col:
                    anchor_col, anchor_row = col, row
        if anchor_col is None:
            return 0

        # labels がある場合は増分スキャンを優先（矩形は参照しない）
        resolved_labels = resolve_labels_in_positions(labels, pos_map)
        if resolved_labels:
            try:
                ws0 = (
                    workbook[target_sheet]
                    if target_sheet in getattr(workbook, "sheetnames", [])
                    else workbook.active
                )
                count = 0
                # 安全上限（無限ループ防止）。現実的な大きさに設定。
                SAFE_MAX = 2000
                for idx in range(1, SAFE_MAX + 1):
                    # 全ての label が非空である限り継続
                    all_non_empty = True
                    for lf in resolved_labels:
                        base = pos_map[lf]
                        c, r = calculate_target_position(
                            base, (direction or "row").lower(), idx, eff_increment
                        )
                        v = read_cell_value((c, r), ws0)
                        if v in (None, ""):
                            all_non_empty = False
                            break
                    if not all_non_empty:
                        break
                    count += 1
                return count
            except Exception:
                # 失敗時は矩形上限へフォールバック
                pass

        # labels が無い場合、または labelsスキャン失敗時は矩形上限を採用
        if not bounds:
            # 矩形が無ければスキップ（テンプレート適用時は特に読み取り禁止）
            return 0 if using_template else 1

        dir_norm = (direction or "row").lower()
        if dir_norm == "row":
            if anchor_row is None:
                return 0
            geometric_max = (bottom - anchor_row) // max(1, eff_increment) + 1
        else:
            if anchor_col is None:
                return 0
            geometric_max = (right - anchor_col) // max(1, eff_increment) + 1
        return max(0, geometric_max)
    except Exception:
        return 0


def compute_count_step_from_anchor_rect_series(
    *,
    labels: list[str],
    internal_slice_count: int,
    target_sheet: Optional[str],
    direction: str,
    container_name: str,
    workbook,
    parent_range_span_for_container: Optional[int],
) -> tuple[Optional[int], Optional[int]]:
    """
    ラベル無し・内部スライス無しの通常ケースで、矩形列による決定論的カウント/ステップを適用。

    戻り値: (element_count_override, step_override)。適用不可時は (None, None)。
    例外は内部で握りつぶし、(None, None) を返す。
    """
    try:
        if labels:
            return None, None
        if internal_slice_count > 1:
            return None, None
        if target_sheet is None:
            return None, None

        step_override: Optional[int] = None
        element_count: Optional[int] = None

        tail_tok2 = container_name.split(".")[-1]
        if tail_tok2 and (not tail_tok2.isdigit()):
            # 親（非 .1）自身の範囲が矩形列のアンカー
            rects_for_parent = _get_anchor_rects_naive(
                workbook, container_name, target_sheet, col_tolerance=0
            )
            if rects_for_parent:
                # 矩形の高さ
                _pl, _pt, _pr, _pb = rects_for_parent[0]
                rect_height = max(1, _pb - _pt + 1)
                # 件数は矩形数
                element_count = len(rects_for_parent)
                # ステップは矩形高さ
                step_override = rect_height
                # 親範囲や .1 アンカーによる上限がある場合はクリップ
                if (
                    isinstance(parent_range_span_for_container, int)
                    and parent_range_span_for_container > 0
                    and element_count > 0
                ):
                    element_count = min(element_count, parent_range_span_for_container)
        else:
            # .1 アンカー自身を起点に矩形列をスキャン
            rects_for_child = _get_anchor_rects_naive(
                workbook, container_name, target_sheet, col_tolerance=0
            )
            if rects_for_child:
                _cl, _ct, _cr, _cb = rects_for_child[0]
                rect_h2 = max(1, _cb - _ct + 1)
                element_count = len(rects_for_child)
                step_override = rect_h2
                # 同上: 矩形列スキャンではアンカー範囲高さで件数クリップしない。
                if (
                    isinstance(parent_range_span_for_container, int)
                    and parent_range_span_for_container > 0
                    and element_count > 0
                ):
                    element_count = min(element_count, parent_range_span_for_container)

        return element_count, step_override
    except Exception:
        return None, None


def clip_element_count_by_parent_bottom_for_single_numeric_child(
    *,
    container_name: str,
    workbook,
    target_sheet: Optional[str],
    current_positions: dict[str, tuple[int, int]],
    eff_increment: int,
    step_override: Optional[int],
    element_count: int,
) -> int:
    """
    直近の親が非数値（例: json.表1.1 の親は json.表1）で、その親に範囲がある場合、
    親範囲の下端で件数をクリップ（親定義がテーブルの行数上限となる仕様）。
    対象は「.1 で終わり、数値トークンが1つだけ」の子コンテナ。
    例外は内部で握り、元の element_count を返す。
    """
    try:
        parts_cn = container_name.split(".")
        num_positions = [i for i, t in enumerate(parts_cn) if t.isdigit()]
        if container_name.endswith(".1") and len(num_positions) == 1:
            parent_non_numeric = ".".join(parts_cn[:-1])  # 例: json.表1
            parent_top = None
            parent_bottom = None
            for sn, coord in iter_defined_name_destinations_all(
                parent_non_numeric, workbook
            ):
                if sn != target_sheet or not coord:
                    continue
                cc = coord.replace("$", "")
                if ":" in cc:
                    (_sc, _sr), (_ec, _er) = parse_range(cc)
                    parent_top, parent_bottom = _sr, _er
                break
            if parent_bottom is not None and element_count > 0:
                anchor_row = min(
                    (row for (col, row) in current_positions.values()), default=None
                )
                if anchor_row is not None and anchor_row <= parent_bottom:
                    step = (
                        step_override if (isinstance(step_override, int) and step_override > 0) else eff_increment
                    ) or 1
                    cap = ((parent_bottom - anchor_row) // step) + 1
                    if cap < element_count:
                        element_count = cap
    except Exception:
        return element_count
    return element_count


def build_parent_and_ancestor_rects_for_nested_scan(
    *,
    workbook,
    target_sheet: Optional[str],
    parent_anchor: str,
    direction: str,
    generated_names: Dict[str, Any],
    anchor_names_chain: List[str],
    ends_with_numeric: bool,
) -> tuple[List[Tuple[int, int, int, int]], List[List[Tuple[int, int, int, int]]]]:
    """
    ネスト走査用に親矩形と祖先矩形チェーンを構築するヘルパー。
    - まず親矩形を _get_anchor_rects_naive で検出
    - 親矩形が1つのみで、親要素が複数生成されていると推定できる場合は定義範囲を等分割して合成
    - 罫線矩形が見つからない場合は、定義範囲と既存名から親件数を推定して等分割するフォールバック
    - 祖先矩形チェーン（直近の親を除く上位矩形列）を検出
    例外は安全側で握り、空を返す
    """
    if target_sheet is None:
        return [], []
    parent_rects = _get_anchor_rects_naive(workbook, parent_anchor, target_sheet, col_tolerance=0)
    parent_rects = _split_parent_rects_by_existing_indices(
        parent_rects=parent_rects,
        target_sheet=target_sheet,
        parent_anchor=parent_anchor,
        workbook=workbook,
        direction=direction,
        generated_names=generated_names,
    )
    if not parent_rects:
        parent_rects = _fallback_parent_rects_without_borders(
            target_sheet=target_sheet,
            parent_anchor=parent_anchor,
            workbook=workbook,
            direction=direction,
        )
    ancestors_rects_chain = _build_ancestors_rects_chain(
        anchor_names_chain=anchor_names_chain,
        ends_with_numeric=ends_with_numeric,
        target_sheet=target_sheet,
        workbook=workbook,
    )
    return parent_rects, ancestors_rects_chain


def _split_parent_rects_by_existing_indices(
    *,
    parent_rects: List[Tuple[int, int, int, int]],
    target_sheet: Optional[str],
    parent_anchor: str,
    workbook,
    direction: str,
    generated_names: Dict[str, Any],
) -> List[Tuple[int, int, int, int]]:
    if not (parent_rects and len(parent_rects) == 1 and target_sheet is not None):
        return parent_rects
    try:
        allowed_parent_indices: set[int] = set()
        parent_base_non_numeric = parent_anchor.rsplit(".", 1)[0] if parent_anchor.endswith(".1") else None
        if parent_base_non_numeric:
            pref = parent_base_non_numeric + "."
            for gk in list(generated_names.keys()):
                if not gk.startswith(pref):
                    continue
                tail = gk[len(pref) :]
                idx_str = tail.split(".", 1)[0]
                if idx_str.isdigit():
                    allowed_parent_indices.add(int(idx_str))
        needed = max(allowed_parent_indices) if allowed_parent_indices else 0
        if not needed or needed <= 1:
            return parent_rects
        pr_coords = None
        for sn, coord in iter_defined_name_destinations_all(parent_anchor, workbook):
            if sn != target_sheet or not coord:
                continue
            cc = str(coord).replace("$", "")
            if ":" in cc:
                (sc0, sr0), (ec0, er0) = parse_range(cc)
                pr_coords = (sc0, sr0, ec0, er0)
            break
        if pr_coords is None:
            return parent_rects
        sc0, sr0, ec0, er0 = pr_coords
        base_rect = RectChain(top=sr0, left=sc0, bottom=er0, right=ec0)
        out: List[Tuple[int, int, int, int]] = []
        if (direction or "row").lower() == "row":
            total = base_rect.height()
            base = total // needed
            rem = total % needed
            cur_top2 = base_rect.top
            for idx in range(1, needed + 1):
                height = base + (1 if idx <= rem else 0)
                bottom2 = min(base_rect.bottom, cur_top2 + height - 1)
                out.append((base_rect.left, cur_top2, base_rect.right, bottom2))
                cur_top2 = bottom2 + 1
        else:
            total = base_rect.width()
            base = total // needed
            rem = total % needed
            cur_left2 = base_rect.left
            for idx in range(1, needed + 1):
                width = base + (1 if idx <= rem else 0)
                right2 = min(base_rect.right, cur_left2 + width - 1)
                out.append((cur_left2, base_rect.top, right2, base_rect.bottom))
                cur_left2 = right2 + 1
        return out or parent_rects
    except Exception:
        return parent_rects


def _fallback_parent_rects_without_borders(
    *,
    target_sheet: Optional[str],
    parent_anchor: str,
    workbook,
    direction: str,
) -> List[Tuple[int, int, int, int]]:
    if target_sheet is None:
        return []
    try:
        pr_coords = None
        for sn, coord in iter_defined_name_destinations_all(parent_anchor, workbook):
            if sn != target_sheet or not coord:
                continue
            cc = str(coord).replace("$", "")
            if ":" in cc:
                (sc0, sr0), (ec0, er0) = parse_range(cc)
                pr_coords = (sc0, sr0, ec0, er0)
            else:
                pr_coords = None
            break
        if pr_coords is None:
            return []
        sc0, sr0, ec0, er0 = pr_coords
        base_rect = RectChain(top=sr0, left=sc0, bottom=er0, right=ec0)
        parts_pa = parent_anchor.split(".")
        base_name = parts_pa[1] if len(parts_pa) >= 2 else None
        parent_count = 0
        if base_name:
            try:
                parent_count = detect_card_count_from_existing_names(
                    base_name, workbook, sheet_name=target_sheet, prefix="json"
                )
            except Exception:
                parent_count = 0
        if parent_count <= 0:
            parent_count = 1
        out: List[Tuple[int, int, int, int]] = []
        if (direction or "row").lower() == "row":
            total = base_rect.height()
            base = total // parent_count
            rem = total % parent_count
            cur_top = base_rect.top
            for idx in range(1, parent_count + 1):
                height = base + (1 if idx <= rem else 0)
                bottom = min(base_rect.bottom, cur_top + height - 1)
                out.append((base_rect.left, cur_top, base_rect.right, bottom))
                cur_top = bottom + 1
        else:
            total = base_rect.width()
            base = total // parent_count
            rem = total % parent_count
            cur_left = base_rect.left
            for idx in range(1, parent_count + 1):
                width = base + (1 if idx <= rem else 0)
                right = min(base_rect.right, cur_left + width - 1)
                out.append((cur_left, base_rect.top, right, base_rect.bottom))
                cur_left = right + 1
        return out
    except Exception:
        return []


def _build_ancestors_rects_chain(
    *,
    anchor_names_chain: List[str],
    ends_with_numeric: bool,
    target_sheet: Optional[str],
    workbook,
) -> List[List[Tuple[int, int, int, int]]]:
    chain: List[List[Tuple[int, int, int, int]]] = []
    try:
        names = anchor_names_chain[:-2] if ends_with_numeric else anchor_names_chain[:-1]
        for anc_name in names:
            if target_sheet is None:
                continue
            rects = _get_anchor_rects_naive(workbook, anc_name, target_sheet, col_tolerance=0)
            # normalize to RectChain internally then back to tuples for compatibility
            normalized: List[Tuple[int, int, int, int]] = []
            for (sc, sr, ec, er) in rects:
                rc = RectChain(top=sr, left=sc, bottom=er, right=ec)
                normalized.append(rc.as_tuple())
            chain.append(normalized)
    except Exception:
        pass
    return chain


def find_first_defined_range_coords(name: str, target_sheet: Optional[str], workbook) -> Optional[RectTuple]:
    """定義名から最初に一致するシート範囲座標 (sc, sr, ec, er) を返す。なければ None。
    target_sheet が指定されている場合は同名シートの範囲のみ対象。
    """
    for (sh_name, coord) in iter_defined_name_destinations_all(name, workbook):
        eff_sheet = sh_name
        eff_coord = coord
        if (not eff_sheet) and isinstance(coord, str) and "!" in coord:
            try:
                eff_sheet, eff_coord = coord.split("!", 1)
            except Exception:
                eff_coord = coord
        if target_sheet is not None and eff_sheet != target_sheet:
            continue
        if isinstance(eff_coord, str):
            cc = eff_coord.replace("$", "")
            if ":" in cc:
                (sc0, sr0), (ec0, er0) = (parse_range(cc))
                return sc0, sr0, ec0, er0
    return None


def emit_field_value_with_optional_range(
    *,
    container_name: str,
    cont_key_parented: str,
    child_idx_emitted: int,
    fname: str,
    value_scalar: Any,
    used_positions: Dict[str, Tuple[int, int]],
    direction: str,
    eff_step_local: int,
    target_sheet: Optional[str],
    workbook,
    ws0,
    generated_names: Dict[str, Any],
) -> None:
    """フィールド値の出力（範囲優先）。範囲が無ければ単一値を出力。"""
    try:
        coords = _resolve_field_range_coords_for_field(container_name, fname, target_sheet, workbook)
        if coords is not None:
            sc, sr, ec, er = coords
            sub_vals = _read_sub_values_for_field(
                sc=sc,
                sr=sr,
                ec=ec,
                er=er,
                ws0=ws0,
                used_positions=used_positions,
                fname=fname,
                child_idx_emitted=child_idx_emitted,
                eff_step_local=eff_step_local,
                direction=direction,
            )
            _emit_field_as_range(
                generated_names=generated_names,
                cont_key_parented=cont_key_parented,
                child_idx_emitted=child_idx_emitted,
                fname=fname,
                sub_vals=sub_vals,
            )
            return
    except Exception:
        # フォールバックしてスカラ出力
        pass
    _emit_field_scalar(
        generated_names=generated_names,
        cont_key_parented=cont_key_parented,
        child_idx_emitted=child_idx_emitted,
        fname=fname,
        value=value_scalar,
    )


def _resolve_field_range_coords_for_field(
    container_name: str,
    fname: str,
    target_sheet: Optional[str],
    workbook,
) -> Optional[tuple[int, int, int, int]]:
    base_range_name = f"{container_name}.{fname}.1"
    coords = find_first_defined_range_coords(base_range_name, target_sheet, workbook)
    if coords is None and container_name.endswith(".1"):
        parent_anchor2 = container_name.rsplit(".", 1)[0]
        alt_range_name = f"{parent_anchor2}.{fname}.1"
        coords = find_first_defined_range_coords(alt_range_name, target_sheet, workbook)
    if coords is None:
        field_range_name = f"{container_name}.{fname}"
        coords = find_first_defined_range_coords(field_range_name, target_sheet, workbook)
        if coords is None and container_name.endswith(".1"):
            parent_anchor2 = container_name.rsplit(".", 1)[0]
            alt_field_range = f"{parent_anchor2}.{fname}"
            coords = find_first_defined_range_coords(alt_field_range, target_sheet, workbook)
    return coords


def _read_sub_values_for_field(
    *,
    sc: int,
    sr: int,
    ec: int,
    er: int,
    ws0,
    used_positions: Dict[str, Tuple[int, int]],
    fname: str,
    child_idx_emitted: int,
    eff_step_local: int,
    direction: str,
) -> Any:
    rows = abs(er - sr) + 1
    cols = abs(ec - sc) + 1
    ws_read = ws0
    tc_tr = used_positions.get(fname)
    if isinstance(tc_tr, tuple) and len(tc_tr) == 2 and all(isinstance(x, int) for x in tc_tr):
        tc, tr = tc_tr
        if rows == 1 and cols >= 1:
            r = tr
            return [read_cell_value((tc + dc, r), ws_read) for dc in range(0, cols)]
        if cols == 1 and rows >= 1:
            c = tc
            return [read_cell_value((c, tr + dr), ws_read) for dr in range(0, rows)]
        return [[read_cell_value((tc + dc, tr + dr), ws_read) for dc in range(0, cols)] for dr in range(0, rows)]
    offset_n = (child_idx_emitted - 1) * max(1, int(eff_step_local))
    if direction == "column":
        sc2, ec2 = sc + offset_n, ec + offset_n
        if rows == 1 and cols >= 1:
            r = sr
            return [read_cell_value((c, r), ws_read) for c in range(sc2, ec2 + 1)]
        if cols == 1 and rows >= 1:
            c = sc2
            return [read_cell_value((c, r), ws_read) for r in range(sr, er + 1)]
        return [[read_cell_value((c, r), ws_read) for c in range(sc2, ec2 + 1)] for r in range(sr, er + 1)]
    # row 方向
    sr2, er2 = sr + offset_n, er + offset_n
    if rows == 1 and cols >= 1:
        r = sr2
        return [read_cell_value((c, r), ws_read) for c in range(sc, ec + 1)]
    if cols == 1 and rows >= 1:
        c = sc
        return [read_cell_value((c, r), ws_read) for r in range(sr2, er2 + 1)]
    return [[read_cell_value((c, r), ws_read) for c in range(sc, ec + 1)] for r in range(sr2, er2 + 1)]


def _emit_field_as_range(
    *,
    generated_names: Dict[str, Any],
    cont_key_parented: str,
    child_idx_emitted: int,
    fname: str,
    sub_vals: Any,
) -> None:
    sub_vals = trim_trailing_empty(sub_vals)
    maybe_idx_key = f"{fname}.1"
    key_with_j = generate_cell_name_for_element(cont_key_parented, child_idx_emitted, maybe_idx_key)
    generated_names[key_with_j] = sub_vals
    try:
        alt_key = generate_cell_name_for_element(cont_key_parented, child_idx_emitted, fname)
        if alt_key not in generated_names:
            generated_names[alt_key] = sub_vals
    except Exception:
        logger.debug("failed to set alt generated key for %s", fname, exc_info=True)
    try:
        stats().cells_generated += 1
    except Exception:
        logger.debug("failed to increment cells_generated in nested emit", exc_info=True)
    logger.debug("NESTED-SET %s=%r", key_with_j, sub_vals)


def _emit_field_scalar(
    *,
    generated_names: Dict[str, Any],
    cont_key_parented: str,
    child_idx_emitted: int,
    fname: str,
    value: Any,
) -> None:
    key = generate_cell_name_for_element(cont_key_parented, child_idx_emitted, fname)
    generated_names[key] = value
    try:
        stats().cells_generated += 1
    except Exception:
        logger.debug("failed to increment cells_generated for scalar emit", exc_info=True)
    logger.debug("NESTED-SET %s=%r", key, value)

def scan_and_emit_nested_children(
    *,
    params: NestedScanParams,
) -> bool:
    """
    親矩形列に沿って子要素をネスト走査し、生成する。処理した場合は True を返す。
    親矩形が無い場合は False。
    """
    try:
        logger.debug(
            "NESTED-SCAN %s sheet=%s parent=%s rects=%s",
            params.container_name,
            params.target_sheet,
            params.parent_anchor,
            len(params.parent_rects) if isinstance(params.parent_rects, list) else None,
        )
        if not params.parent_rects:
            logger.debug("NESTED-NO-PARENT-RECTS %s", params.container_name)
            return False
        child_anchor_row = min((row for (col, row) in params.current_positions.values()))
        allowed_parent_indices = _get_allowed_parent_indices(params.parent_anchor, params.generated_names)
        base_numeric_token_fields = _get_base_numeric_token_fields(params.current_positions, params.ws0)
        emitted_total = _scan_and_emit_main(
            params.workbook,
            params.target_sheet,
            params.ws0,
            params.container_name,
            params.direction,
            params.current_positions,
            params.labels,
            params.policy,
            params.parent_rects,
            params.ancestors_rects_chain,
            params.generated_names,
            params.num_positions,
            params.ends_with_numeric,
            child_anchor_row,
            allowed_parent_indices,
            base_numeric_token_fields,
        )
        if params.ends_with_numeric and params.ancestors_rects_chain:
            emitted_total += _scan_and_emit_fallback(
                params.workbook,
                params.target_sheet,
                params.ws0,
                params.container_name,
                params.direction,
                params.current_positions,
                params.labels,
                params.policy,
                params.parent_rects,
                params.ancestors_rects_chain,
                params.generated_names,
                params.num_positions,
                params.ends_with_numeric,
                child_anchor_row,
                base_numeric_token_fields,
            )
        logger.debug("NESTED-EMIT %s sheet=%s total=%s", params.container_name, params.target_sheet, emitted_total)
        return True
    except Exception as e:
        logger.debug("NESTED-ERR %s: %s", params.container_name, e, exc_info=True)
        return False

# --- サブ関数群 ---
def _get_allowed_parent_indices(parent_anchor, generated_names):
    allowed_parent_indices: set[int] = set()
    parent_base_non_numeric = parent_anchor.rsplit(".", 1)[0] if parent_anchor.endswith(".1") else None
    if parent_base_non_numeric:
        prefix = parent_base_non_numeric + "."
        for gk in list(generated_names.keys()):
            if not gk.startswith(prefix):
                continue
            tail = gk[len(prefix) :]
            idx_str = tail.split(".", 1)[0]
            if idx_str.isdigit():
                allowed_parent_indices.add(int(idx_str))
    return allowed_parent_indices

def _get_base_numeric_token_fields(current_positions, ws0):
    base_numeric_token_fields: list[str] = []
    for fn, (c0, r0) in current_positions.items():
        v0 = read_cell_value((c0, r0), ws0)
        if is_numeric_token_string(v0):
            base_numeric_token_fields.append(fn)
    return base_numeric_token_fields

def should_suppress_element(
    *,
    values_by_field: Dict[str, Any],
    non_empty: bool,
    labels_in_scope: Optional[List[str]],
    policy: ExtractionPolicy,
) -> bool:
    """ネスト子要素の出力抑止判定。

    最小互換仕様:
    - ラベルが指定されている場合: そのラベルのセルが空ならスキップ
    - 非ラベルの場合: 行全体が空ならスキップ
    """
    try:
        # ラベル指定あり: ラベルいずれかが非空なら採用。全て空なら抑止。
        if labels_in_scope:
            for lf in labels_in_scope:
                if values_by_field.get(lf) not in (None, ""):
                    return False
            return True
        # ラベル指定なし: 行全体が空なら抑止
        return not non_empty
    except Exception:
        return False

def compute_group_indexes_and_bounds(
    *,
    pt: int,
    pb: int,
    ancestors_rects_chain: List[List[RectTuple]] | None,
    policy: ExtractionPolicy,
) -> Tuple[List[int], int, int]:
    """先祖矩形チェーンからグループインデックス列と有効上下端(eff_pt, eff_pb)を算出する。

    互換ポリシー:
    - `policy.nested_scan.ancestors_first_bounds` が真の場合、`pick_effective_bounds` に従う。
    - 偽の場合、親矩形の上下端を採用。
    - 先祖が無い場合は ([], pt, pb) を返す。
    """
    def _filter_ordered_bounds(
        rects_at_level: List[RectTuple], bounds: tuple[int, int] | None
    ) -> List[Tuple[int, int, int]]:
        ordered: List[Tuple[int, int, int]] = []
        bpt, bpb = (None, None) if bounds is None else bounds
        for j, (_al, at, _ar, ab) in enumerate(rects_at_level, start=1):
            if bpt is None and bpb is None:
                ordered.append((j, at, ab))
            else:
                # フィルタ: 先に決まっている境界内に完全に入るもののみ
                if at >= cast(int, bpt) and ab <= cast(int, bpb):
                    ordered.append((j, at, ab))
        return ordered

    def _find_group_index_for_top(
        ordered_rects: List[Tuple[int, int, int]], top: int
    ) -> tuple[int, tuple[int, int] | None]:
        gi_local = 1
        matched_bounds: tuple[int, int] | None = None
        for idx_local, (_jg, at, ab) in enumerate(ordered_rects, start=1):
            if at <= top <= ab:
                gi_local = idx_local
                matched_bounds = (at, ab)
                break
        return gi_local, matched_bounds

    def _compute_eff_bounds(
        pt0: int, pb0: int, chain: List[List[RectTuple]] | None, pol: ExtractionPolicy
    ) -> tuple[int, int]:
        if not chain:
            return pt0, pb0
        if pol.nested_scan.ancestors_first_bounds:
            return pick_effective_bounds(pt0, pb0, chain)
        return pt0, pb0

    eff_pt, eff_pb = pt, pb
    group_indexes: List[int] = []
    if ancestors_rects_chain:
        cur_top_for_mapping = pt
        bounds: tuple[int, int] | None = None
        for level_idx, rects_at_level in enumerate(ancestors_rects_chain):
            ordered = _filter_ordered_bounds(rects_at_level, bounds)
            gi_local, matched = _find_group_index_for_top(ordered, cur_top_for_mapping)
            group_indexes.append(gi_local)
            if matched is not None:
                bounds = matched
                if level_idx == 0:
                    eff_pt, eff_pb = matched

    eff_pt, eff_pb = _compute_eff_bounds(pt, pb, ancestors_rects_chain, policy)
    return group_indexes, eff_pt, eff_pb

def collect_row_values(
    *,
    ws,
    current_positions: Mapping[str, Tuple[int, int]],
    direction: str,
    local_index: int,
    step: int,
    eff_top: int,
    eff_bottom: int,
) -> tuple[Dict[str, Any], Dict[str, Tuple[int, int]], bool]:
    """一行分のセル値を収集し、使用座標と非空判定を返す。

    入力:
    - ws: ワークシートオブジェクト（`read_cell_value` が参照）
    - current_positions: フィールド -> (col,row) の基準座標
    - direction: 進行方向（"row"/"col"）
    - local_index: 相対インデックス（1始まり）
    - step: ステップ幅
    - eff_top/eff_bottom: 有効範囲の上下端（行方向の判定に利用）

    出力:
    - values_by_field: フィールド -> 値
    - used_positions: フィールド -> 実読取位置 (col,row)
    - non_empty: いずれかの値が非空であれば True
    """
    values_by_field: Dict[str, Any] = {}
    used_positions: Dict[str, Tuple[int, int]] = {}
    non_empty = False
    for fname, (c0, r0) in current_positions.items():
        tc, tr = calculate_target_position((c0, r0), direction, local_index, step)
        if tr < eff_top or tr > eff_bottom:
            val = ""
        else:
            val = read_cell_value((tc, tr), ws)
        values_by_field[fname] = val
        used_positions[fname] = (tc, tr)
        if val not in (None, ""):
            non_empty = True
    return values_by_field, used_positions, non_empty

def has_non_numeric_payload(
    *,
    values_by_field: Mapping[str, Any],
    numeric_token_fields: Sequence[str],
) -> bool:
    """行内に、数値トークン以外で非空の値が一つでもあるかを判定する。"""
    return any(
        (fn not in numeric_token_fields) and (vv not in (None, ""))
        for fn, vv in values_by_field.items()
    )

def should_skip_early_checks(
    *,
    values_by_field: Mapping[str, Any],
    non_empty: bool,
    resolved_labels: Sequence[str] | None,
    policy: ExtractionPolicy,
    expected_len: int,
    numeric_token_fields: Sequence[str],
    used_positions: Mapping[str, Tuple[int, int]],
    group_key: Tuple[int, ...],
    claims_by_group: Dict[Tuple[int, ...], set[int]],
) -> bool:
    """_scan_and_emit_main での早期スキップ判定をひとまとめにする。

    順序:
    1) ラベル／非空に基づく抑止判定
    2) 所有権ベースのスキップ
    どちらかが真なら True（スキップ）。
    """
    if should_suppress_element(
        values_by_field=dict(values_by_field),
        non_empty=non_empty,
        labels_in_scope=list(resolved_labels) if resolved_labels is not None else None,
        policy=policy,
    ):
        return True
    if should_skip_by_row_ownership(
        policy=policy.nested_scan,
        expected_len=expected_len,
        numeric_token_fields=list(numeric_token_fields),
        used_positions=dict(used_positions),
        non_empty=non_empty,
        group_key=group_key,
        claims_by_group=claims_by_group,
    ):
        return True
    return False

def _scan_and_emit_main(
    workbook,
    target_sheet,
    ws0,
    container_name,
    direction,
    current_positions,
    labels,
    policy,
    parent_rects,
    ancestors_rects_chain,
    generated_names,
    num_positions,
    ends_with_numeric,
    child_anchor_row,
    allowed_parent_indices,
    base_numeric_token_fields,
) -> int:
    """メイン走査: 親矩形ごとに行を走査し、子要素を出力する。

    入力:
    - ワークブック/シート、コンテナ名、走査方向、現在位置、ラベル、ポリシ、親矩形と祖先チェーン、生成名マップ。
    - 数値トークン構成（`num_positions`/`ends_with_numeric`）、アンカー行、許可親インデックス、数値トークン対象フィールド。

    出力:
    - 出力した子要素の合計数（int）。

    例外/副作用:
    - 原則ここでは例外を送出しない（下位が握りつぶす）。
    - `generated_names` 等のマップを破壊的に更新し、`stats()` を通じ処理統計を更新しうる。

    フェーズ構成（概要）:
    1) 親矩形ループ: 有効境界とグループインデックスを算出
    2) アンカー探索: 予測インデックス→ローカルアンカー→反復範囲 i0..i1
    3) 行ループ: 値収集→早期スキップ→数値トークン調整→受理判定
    4) 出力: EmitOps に委譲
    """
    # === 0) ループ外の採番・重複・所有権トラッカ初期化 ===
    grand_local_parent_counts: Dict[Tuple[int, ...], int] = {}
    parent_local_counts_by_group: Dict[Tuple[int, ...], int] = {}
    child_emitted_by_group: Dict[Tuple[int, ...], int] = {}
    seen_numeric_tokens_by_group: Dict[Tuple[int, ...], set[str]] = {}
    child_row_claims_by_group: Dict[Tuple[int, ...], set[int]] = {}
    emitted_total = 0
    for p_index, (pl, pt, pr, pb) in enumerate(parent_rects, start=1):
        # === 1) 親矩形ごとの前処理: 有効境界とグループインデックス ===
        group_indexes, eff_pt, eff_pb = compute_group_indexes_and_bounds(
            pt=pt,
            pb=pb,
            ancestors_rects_chain=ancestors_rects_chain,
            policy=policy,
        )
        group_key = tuple(group_indexes) if group_indexes else (1,)

        # === 2) アンカー探索: ステップとローカルアンカー決定 ===
        eff_step_local = derive_eff_step_local(
            labels_present=bool(labels),
            ends_with_numeric=ends_with_numeric,
            workbook=workbook,
            container_name=container_name,
            target_sheet=target_sheet,
            eff_pt=eff_pt,
            eff_pb=eff_pb,
            direction=direction,
            policy=policy.nested_scan,
        )

        if eff_pb < child_anchor_row:
            continue

        numeric_token_fields: list[str] = list(base_numeric_token_fields)

        # 従来どおり、プローブ選定は select_probe_fields で行う（挙動不変）
        probe_fields = select_probe_fields(
            current_positions=current_positions,
            labels=labels,
            numeric_token_fields=numeric_token_fields,
        )

        local_aligned_row = align_row_phase(eff_pt, child_anchor_row, eff_step_local)

        ws_probe = ws0
        predicted_parent_local_index = AnchorOps.predict_parent_local_index(
            group_key=group_key,
            ancestors_rects_chain=ancestors_rects_chain,
            parent_local_counts_by_group=parent_local_counts_by_group,
            grand_local_parent_counts=grand_local_parent_counts,
        )
        found_local_anchor = AnchorOps.locate_found_local_anchor(
            ws_probe=ws_probe,
            current_positions=current_positions,
            probe_fields=probe_fields,
            local_aligned_row=local_aligned_row,
            eff_pb=eff_pb,
            step=eff_step_local,
            num_positions=num_positions,
            group_indexes=group_indexes,
            predicted_parent_local_index=predicted_parent_local_index,
        )

        anchor_for_calc = found_local_anchor if found_local_anchor is not None else local_aligned_row
        i0, i1 = _compute_iteration_bounds(
            child_anchor_row=child_anchor_row,
            region_bottom=eff_pb,
            anchor_row=anchor_for_calc,
            step=eff_step_local,
        )
        logger.debug(
            "NESTED-DBG %s p=%s eff_pt=%s eff_pb=%s step=%s child_anchor_row=%s found_row=%s i0=%s i1=%s",
            container_name,
            p_index,
            eff_pt,
            eff_pb,
            eff_step_local,
            child_anchor_row,
            found_local_anchor,
            i0,
            i1,
        )
        if i1 < i0:
            continue

        parent_local_index = AnchorOps.increment_parent_local_index(
            group_key=group_key,
            ancestors_rects_chain=ancestors_rects_chain,
            parent_local_counts_by_group=parent_local_counts_by_group,
            grand_local_parent_counts=grand_local_parent_counts,
        )
        # 許可された親インデックス集合が十分に得られている場合のみフィルタを適用。
        # 1件だけのときは不完全な観測の可能性があるため、抑止しない。
        if (
            isinstance(allowed_parent_indices, set)
            and len(allowed_parent_indices) >= 2
            and parent_local_index not in allowed_parent_indices
        ):
            continue

    # === 3) 行ループ: 値収集→早期スキップ→数値調整→受理 ===
        for local_i in range(i0, i1 + 1):
            # ラベル解決はループ内で都度行う（従来の等価動作に戻す）
            resolved_labels2 = [lf for lf in labels if lf in current_positions]
            values_by_field, used_positions, non_empty = collect_row_values(
                ws=ws0,
                current_positions=current_positions,
                direction=direction,
                local_index=local_i,
                step=eff_step_local,
                eff_top=eff_pt,
                eff_bottom=eff_pb,
            )

            if not passes_numeric_row_requirements(
                values_by_field=values_by_field,
                num_positions=num_positions,
                numeric_token_fields=base_numeric_token_fields,
            ):
                continue

            if should_skip_early_checks(
                values_by_field=values_by_field,
                non_empty=non_empty,
                resolved_labels=resolved_labels2,
                policy=policy,
                expected_len=len(num_positions),
                numeric_token_fields=base_numeric_token_fields,
                used_positions=used_positions,
                group_key=group_key,
                claims_by_group=child_row_claims_by_group,
            ):
                continue

            _adjust_values_with_numeric_token(
                values_by_field=values_by_field,
                used_positions=used_positions,
                numeric_token_fields=base_numeric_token_fields,
                ws=ws0,
                eff_top=eff_pt,
                eff_bottom=eff_pb,
            )

            seq_like_val2 = extract_seq_like_value(values_by_field, base_numeric_token_fields)
            if not check_seq_accept_and_dedup(
                policy=policy.numeric_tokens,
                expected_len=len(num_positions),
                has_numeric_series_field=bool(base_numeric_token_fields),
                seq_like_val=seq_like_val2,
                group_indexes=group_indexes,
                parent_local_index=parent_local_index,
                group_key_for_dedup=group_key + (parent_local_index,),
                seen_tokens=seen_numeric_tokens_by_group,
            ):
                continue

            if not row_is_acceptable_for_emission(
                values_by_field=values_by_field,
                non_empty=non_empty,
                resolved_labels=resolved_labels2,
                policy=policy,
            ):
                continue

            # === 4) 出力: EmitOps に委譲 ===
            emitted_total += EmitOps.emit_row_payload(
                container_name=container_name,
                parent_local_index=parent_local_index,
                group_indexes=group_indexes,
                group_key=group_key,
                child_emitted_by_group=child_emitted_by_group,
                values_by_field=values_by_field,
                used_positions=used_positions,
                direction=direction,
                eff_step_local=eff_step_local,
                target_sheet=target_sheet,
                workbook=workbook,
                ws0=ws0,
                generated_names=generated_names,
            )
    return emitted_total

def row_is_acceptable_for_emission(
    *,
    values_by_field: Mapping[str, Any],
    non_empty: bool,
    resolved_labels: Sequence[str],
    policy,
) -> bool:
    """最終出力直前の行が受理可能かを判定する。

    - 抑止ポリシー（labels_in_scope=None）で弾かれる場合は False
    - ラベルがなく、かつ実質的に空行なら False
    - それ以外は True（従来ロジックと等価）
    """
    if should_suppress_element(
        values_by_field=dict(values_by_field),
        non_empty=non_empty,
        labels_in_scope=None,
        policy=policy,
    ):
        return False
    if not resolved_labels and not non_empty:
        return False
    return True

def _scan_and_emit_fallback(
    workbook,
    target_sheet,
    ws0,
    container_name,
    direction,
    current_positions,
    labels,
    policy,
    parent_rects,
    ancestors_rects_chain,
    generated_names,
    num_positions,
    ends_with_numeric,
    child_anchor_row,
    base_numeric_token_fields,
) -> int:
    """フォールバック走査: 非範囲の単一値として出力する安全網。

    役割:
    - アンカー行に揃えた上で、数値トークン行を主とした簡易走査を行い、`EmitOps.emit_row_payload_fallback` で出力する。

    出力: 出力した子要素の合計数（int）。
    例外/副作用: 原則ここで例外は上げず、`generated_names`/統計を更新。

    フェーズ:
    1) 親矩形から祖先グループとの交差を把握
    2) 祖先グループごとに、親が未配置のグループのみ走査範囲を確定
    3) アンカー近傍の行反復で値収集→受理判定→フォールバック出力
    """
    emitted_total = 0
    top_anc_rects = ancestors_rects_chain[0] if ancestors_rects_chain else []
    # フォールバック内の一時状態（重複トークン・子インデックス）
    seen_numeric_tokens_by_group: Dict[Tuple[int, int], set[str]] = {}
    child_emitted_by_group: Dict[Tuple[int, int], int] = {}
    groups_with_parent: set[int] = set()
    for _pidx, (_pl, _pt2, _pr2, _pb2) in enumerate(parent_rects or [], start=1):
        for g_i, (_al, _at2, _ar2, _ab2) in enumerate(top_anc_rects, start=1):
            if not (_pb2 < _at2 or _pt2 > _ab2):
                groups_with_parent.add(g_i)

    for g_i, (_al, _at2, _ar2, _ab2) in enumerate(top_anc_rects, start=1):
        if g_i in groups_with_parent:
            continue
        fb_pt, fb_pb = _at2, _ab2
        if fb_pb < child_anchor_row:
            continue
        fb_step = 1
        # 祖先 g_i と親が交差しないグループのアンカー候補を探索
        local_aligned_fb = align_row_phase(fb_pt, child_anchor_row, fb_step)
        predicted_parent_local_index = 1
        expected_len_fb = len(num_positions)
        expected_prefix_fb = [str(x) for x in ([g_i] + [predicted_parent_local_index])]
        anchor_row_fb = AnchorOps.find_numeric_token_anchor_row(
            current_positions=current_positions,
            ws0=ws0,
            start_row=local_aligned_fb,
            end_row=fb_pb,
            expected_prefix=expected_prefix_fb,
            expected_len=expected_len_fb,
        ) or local_aligned_fb
        i0_fb, i1_fb = _compute_iteration_bounds(
            child_anchor_row=child_anchor_row,
            region_bottom=fb_pb,
            anchor_row=anchor_row_fb,
            step=fb_step,
        )
        if i1_fb < i0_fb:
            continue
        # アンカー周辺の行を順次評価
        for li in range(i0_fb, i1_fb + 1):
            vals_fb, used_pos, non_empty_fb = collect_row_values(
                ws=ws0,
                current_positions=current_positions,
                direction=direction,
                local_index=li,
                step=fb_step,
                eff_top=fb_pt,
                eff_bottom=fb_pb,
            )
            seq_val_fb = extract_seq_like_value(vals_fb, None)
            if not seq_val_fb:
                continue
            if not has_non_numeric_payload(values_by_field=vals_fb, numeric_token_fields=base_numeric_token_fields):
                continue
            # dedup判定
            grp_key_fb2 = (g_i, predicted_parent_local_index)
            seen_fb = seen_numeric_tokens_by_group.setdefault(grp_key_fb2, set())
            if seq_val_fb in seen_fb:
                continue
            seen_fb.add(seq_val_fb)
            if should_suppress_element(
                values_by_field=vals_fb,
                non_empty=non_empty_fb,
                labels_in_scope=None,
                policy=policy,
            ):
                continue
            emitted_total += EmitOps.emit_row_payload_fallback(
                container_name=container_name,
                predicted_parent_local_index=predicted_parent_local_index,
                group_index=g_i,
                values_by_field=vals_fb,
                child_emitted_by_group=child_emitted_by_group,
                generated_names=generated_names,
            )
    return emitted_total

def passes_numeric_row_requirements(
    *,
    values_by_field: Mapping[str, Any],
    num_positions: Sequence[Any],
    numeric_token_fields: Sequence[str] | None,
) -> bool:
    """数値トークンを使用するケースで、行が最低限の要件を満たすか判定する。"""
    try:
        if not (len(num_positions) >= 2 and numeric_token_fields):
            return True
        numeric_on_row = any(
            (fn in numeric_token_fields) and is_numeric_token_string(vv)
            for fn, vv in values_by_field.items()
        )
        if not numeric_on_row:
            return False
        return has_non_numeric_payload(values_by_field=values_by_field, numeric_token_fields=numeric_token_fields)
    except Exception:
        return False

def extract_seq_like_value(values_by_field: Mapping[str, Any], numeric_token_fields: Optional[Sequence[str]]) -> Optional[str]:
    """値集合から数値トークン様の値を1つ拾って返す（なければ None）。

    - `numeric_token_fields` が与えられた場合: そのフィールド名群の中から最初に見つかったトークンを優先
    - `None` の場合: 行内のいずれのフィールドでも最初に見つかったトークンを返す（フォールバックと同等）
    返値は文字列化済み。
    """
    if numeric_token_fields:
        for fn in numeric_token_fields:
            vv = values_by_field.get(fn)
            if is_numeric_token_string(vv):
                return str(vv)
    else:
        for _fn, _vv in values_by_field.items():
            if is_numeric_token_string(_vv):
                return str(_vv)
    return None

def _adjust_values_with_numeric_token(
    *,
    values_by_field: Dict[str, Any],
    used_positions: Mapping[str, Tuple[int, int]],
    numeric_token_fields: Sequence[str],
    ws,
    eff_top: int,
    eff_bottom: int,
) -> None:
    """行内の数値トークンと同値な値を持つフィールドに対し、近傍セルから代替値を補完する。"""
    try:
        seq_like_val = None
        seq_like_col = None
        primary_numeric_field = None
        for fn, vc in values_by_field.items():
            if is_numeric_token_string(vc):
                seq_like_val = vc
                seq_like_col, _ = used_positions.get(fn, (None, None))
                primary_numeric_field = fn
                break
        if seq_like_val is None:
            return
        for fn in list(values_by_field.keys()):
            if primary_numeric_field is not None and fn == primary_numeric_field:
                continue
            vc = values_by_field.get(fn)
            fc, fr = used_positions.get(fn, (None, None))
            if vc == seq_like_val and fr is not None:
                for d in (1, 2, 3):
                    alt_candidates: list[tuple[int, int]] = []
                    if seq_like_col is not None:
                        alt_candidates.append((seq_like_col + d, fr))
                    if fc is not None:
                        alt_candidates.append((fc + d, fr))
                    for ac, ar in alt_candidates:
                        if ar < eff_top or ar > eff_bottom:
                            continue
                        alt_val = read_cell_value((ac, ar), ws)
                        if alt_val not in (None, "") and alt_val != seq_like_val:
                            values_by_field[fn] = alt_val
                            raise StopIteration  # 1つ見つけたら終了（元実装のbreakのネスト解消）
    except StopIteration:
        return
    except Exception:
        # 失敗しても無視（元実装と同等の寛容さ）
        return

class EmitOps:
    """出力系ユーティリティをまとめたクラス。

    役割:
    - 子要素の採番、親・祖先インデックスを付与したキー生成、フィールド値の出力（範囲対応含む）。
    - 既存の関数実装（薄いラッパー）と同等の振る舞いを維持する。
    """

    @classmethod
    def build_parented_container_key(
        cls,
        *,
        container_name: str,
        parent_local_index: int,
        group_indexes: Sequence[int] | None,
    ) -> str:
        """親インデックスと祖先グループインデックスを反映したコンテナキーを生成する。

        入力:
        - `container_name`: 基底のコンテナ名（例: `json.A.1.items.1`）。
        - `parent_local_index`: 親ローカルインデックス（1始まり）。
        - `group_indexes`: 祖先グループのインデックス列（外側→内側）。無ければ `None`。

        出力:
        - 親・祖先を反映したキー文字列。

        例外:
        - なし（入力は呼び出し側で整合がとれている前提）。
        副作用:
        - なし。
        """
        key = _set_parent_index_in_key(container_name, parent_local_index)
        if group_indexes:
            for offset, gi in enumerate(reversed(group_indexes), start=3):
                key = _replace_nth_from_end_numeric(key, offset, gi)
        return key

    @classmethod
    def emit_row_payload(
        cls,
        *,
        container_name: str,
        parent_local_index: int,
        group_indexes: Sequence[int],
        group_key: Tuple[int, ...],
        child_emitted_by_group: Dict[Tuple[int, ...], int],
        values_by_field: Mapping[str, Any],
        used_positions: Mapping[str, Tuple[int, int]],
        direction: str,
        eff_step_local: int,
        target_sheet: Optional[str],
        workbook,
        ws0,
        generated_names: Dict[str, Any],
    ) -> int:
        """1 行分の値を出力し、子要素インデックスを採番する（範囲対応）。

        入力:
        - `container_name`: 対象コンテナの `.1` アンカー名。
        - `parent_local_index`: 親のローカルインデックス（1始まり）。
        - `group_indexes`: 祖先グループのインデックス列（外側→内側）。
        - `group_key`: グループ境界のキー（採番スコープ）。
        - `child_emitted_by_group`: 子インデックスの採番管理マップ（更新される）。
        - `values_by_field`: フィールド名→スカラ値の辞書。
        - `used_positions`: フィールド名→セル位置（範囲読み取り時の起点）。
        - `direction`: 走査方向（"row"/"column"）。
        - `eff_step_local`: 走査ステップ数（1 以上）。
        - `target_sheet`/`workbook`/`ws0`: 値の読み取り元。
        - `generated_names`: 出力先辞書（更新される）。

        出力:
        - 当該呼び出しで出力した子要素数（通常 1）。

        例外:
        - 下位の値読み取りで例外は握りつぶし、フォールバックで単一値出力に切り替えるため、原則ここでは送出しない。
        副作用:
        - `generated_names` にキー/値を追記。
        - `child_emitted_by_group` を更新。
        - `processing_stats.cells_generated` を可能な範囲でインクリメント。
        """
        grp_key = group_key + (parent_local_index,)
        prev = child_emitted_by_group.get(grp_key, 0)
        child_idx_emitted = prev + 1
        child_emitted_by_group[grp_key] = child_idx_emitted
        emitted_here = 1

        cont_key_parented = cls.build_parented_container_key(
            container_name=container_name,
            parent_local_index=parent_local_index,
            group_indexes=group_indexes,
        )

        for fname, val in list(values_by_field.items()):
            emit_field_value_with_optional_range(
                container_name=container_name,
                cont_key_parented=cont_key_parented,
                child_idx_emitted=child_idx_emitted,
                fname=fname,
                value_scalar=val,
                used_positions=dict(used_positions),
                direction=direction,
                eff_step_local=eff_step_local,
                target_sheet=target_sheet,
                workbook=workbook,
                ws0=ws0,
                generated_names=generated_names,
            )
        return emitted_here

    @classmethod
    def emit_row_payload_fallback(
        cls,
        *,
        container_name: str,
        predicted_parent_local_index: int,
        group_index: int,
        values_by_field: Mapping[str, Any],
        child_emitted_by_group: Dict[Tuple[int, int], int],
        generated_names: Dict[str, Any],
    ) -> int:
        """範囲解釈を行わないフォールバック出力（単一値のみ）。

        入力:
        - `container_name`: 対象コンテナの `.1` アンカー名。
        - `predicted_parent_local_index`: 親ローカルインデックスの予測値（1始まり）。
        - `group_index`: 現在のグループインデックス。
        - `values_by_field`: フィールド名→スカラ値の辞書。
        - `child_emitted_by_group`: 子インデックスの採番管理マップ（更新される）。
        - `generated_names`: 出力先辞書（更新される）。

        出力:
        - 当該呼び出しで出力した子要素数（1 固定）。

        例外/副作用:
        - 例外は握りつぶし、`generated_names` への追加と統計を継続。
        - `processing_stats.cells_generated` のインクリメントを試行。
        """
        cont_key_parented_fb = cls.build_parented_container_key(
            container_name=container_name,
            parent_local_index=predicted_parent_local_index,
            group_indexes=[group_index],
        )
        grp_key_fb = (group_index, predicted_parent_local_index)
        prev_fb = child_emitted_by_group.get(grp_key_fb, 0)
        child_idx_fb = prev_fb + 1
        child_emitted_by_group[grp_key_fb] = child_idx_fb
        for fn, vv in values_by_field.items():
            k = generate_cell_name_for_element(cont_key_parented_fb, child_idx_fb, fn)
            generated_names[k] = vv
            try:
                stats().cells_generated += 1
            except Exception:
                logger.debug("failed to increment cells_generated in AnchorOps.emit_fallback_child_values", exc_info=True)
        return 1

class AnchorOps:
    """アンカー探索と親ローカルインデックス管理をまとめたクラス。

    役割:
    - 親ローカルインデックスの予測/増分、子要素ローカルアンカー行の探索、数値トークン行のアンカー推定。
    - 既存の関数実装と同等の振る舞いを維持する。
    """

    @classmethod
    def predict_parent_local_index(
        cls,
        *,
        group_key: Tuple[int, ...],
        ancestors_rects_chain: Sequence[Sequence[RectTuple]] | None,
        parent_local_counts_by_group: Mapping[Tuple[int, ...], int],
        grand_local_parent_counts: Mapping[Tuple[int, ...], int],
    ) -> int:
        """親ローカルインデックスを予測して返す（1始まり）。

        入力:
        - `group_key`: グループ境界のキー。
        - `ancestors_rects_chain`: 祖先矩形のチェーン（あれば親ローカル採番を優先）。
        - `parent_local_counts_by_group`: 直近親に対する採番カウンタ（読み取り専用）。
        - `grand_local_parent_counts`: 祖父系の採番カウンタ（読み取り専用）。

        出力: 予測される親ローカルインデックス。
        例外/副作用: なし。
        """
        if ancestors_rects_chain:
            return parent_local_counts_by_group.get(group_key, 0) + 1
        return grand_local_parent_counts.get(group_key, 0) + 1

    @classmethod
    def locate_found_local_anchor(
        cls,
        *,
        ws_probe,
        current_positions: Mapping[str, Tuple[int, int]],
        probe_fields: Sequence[str],
        local_aligned_row: int,
        eff_pb: int,
        step: int,
        num_positions: Sequence[Any],
        group_indexes: Sequence[int],
        predicted_parent_local_index: int,
    ) -> Optional[int]:
        """子要素のローカルアンカー行を探索して返す。

        入力:
        - `ws_probe`: 参照ワークシート。
        - `current_positions`: フィールド→セル位置。
        - `probe_fields`: 探索対象フィールド。
        - `local_aligned_row`: 探索の基準行（1始まり）。
        - `eff_pb`: 有効領域の下端行。
        - `step`: 走査ステップ。
        - `num_positions`: 数値トークンの位置配列。
        - `group_indexes`: 祖先グループインデックス列。
        - `predicted_parent_local_index`: 予測される親ローカルインデックス。

        出力: 見つかったローカルアンカー行（見つからなければ `None`）。
        例外/副作用: なし（内部で失敗時はフォールバック探索へ）。
        """
        found_local_anchor: Optional[int] = None
        numeric_probe_cols: list[int] = []
        try:
            for fn in probe_fields:
                pos = current_positions.get(fn)
                if pos is None:
                    continue
                c0, _r0 = pos
                if c0 is not None:
                    numeric_probe_cols.append(c0)
        except Exception:
            logger.debug("AnchorOps.locate_found_local_anchor: collecting numeric_probe_cols failed", exc_info=True)
        if numeric_probe_cols:
            expected_len = len(num_positions)
            expected_prefix = [str(x) for x in (list(group_indexes) + [predicted_parent_local_index])]
            found_local_anchor = find_local_anchor_row(
                ws=ws_probe,
                current_positions=dict(current_positions),
                probe_fields=list(probe_fields),
                numeric_probe_cols=numeric_probe_cols,
                local_aligned_row=local_aligned_row,
                eff_pb=eff_pb,
                step=step,
                expected_len=expected_len,
                expected_prefix=expected_prefix,
            )
        if found_local_anchor is None:
            found_local_anchor = find_local_anchor_row(
                ws=ws_probe,
                current_positions=dict(current_positions),
                probe_fields=list(probe_fields),
                numeric_probe_cols=[],
                local_aligned_row=local_aligned_row,
                eff_pb=eff_pb,
                step=step,
                expected_len=len(num_positions),
                expected_prefix=[str(x) for x in list(group_indexes)],
            )
        return found_local_anchor

    @classmethod
    def increment_parent_local_index(
        cls,
        *,
        group_key: Tuple[int, ...],
        ancestors_rects_chain: Sequence[Sequence[RectTuple]] | None,
        parent_local_counts_by_group: Dict[Tuple[int, ...], int],
        grand_local_parent_counts: Dict[Tuple[int, ...], int],
    ) -> int:
        """親ローカルインデックスのカウンタを 1 増やし、その新値を返す。

        入力:
        - `group_key`: グループ境界のキー。
        - `ancestors_rects_chain`: 祖先矩形のチェーン（あれば親ローカル採番を使用）。
        - `parent_local_counts_by_group`/`grand_local_parent_counts`: カウンタ（更新される）。

        出力: 更新後の親ローカルインデックス。
        例外/副作用: 該当カウンタマップを破壊的に更新。
        """
        if ancestors_rects_chain:
            parent_local_counts_by_group.setdefault(group_key, 0)
            parent_local_counts_by_group[group_key] += 1
            return parent_local_counts_by_group[group_key]
        grand_local_parent_counts.setdefault(group_key, 0)
        grand_local_parent_counts[group_key] += 1
        return grand_local_parent_counts[group_key]

    @classmethod
    def find_numeric_token_anchor_row(
        cls,
        *,
        current_positions: Mapping[str, Tuple[int, int]],
        ws0,
        start_row: int,
        end_row: int,
        expected_prefix: Sequence[str],
        expected_len: int,
        max_scans: int = 5000,
    ) -> Optional[int]:
        """数値トークン（ハイフン区切り）を用いてアンカー行を粗く探索する。

        入力:
        - `current_positions`: フィールド→セル位置。
        - `ws0`: 参照ワークシート。
        - `start_row`/`end_row`: 探索範囲（含む）。
        - `expected_prefix`: トークンの先頭一致配列（例: 祖先グループインデックス群）。
        - `expected_len`: 期待トークン長。
        - `max_scans`: 安全のための最大走査行数。

        出力: 見つかった行番号、なければ `None`。
        例外/副作用: なし。
        """
        rfb = start_row
        scans = 0
        while rfb <= end_row and scans < max_scans:
            sval = ""
            for _fn, (_c0, _r0) in current_positions.items():
                sval = read_cell_value((_c0, rfb), ws0)
                if is_numeric_token_string(sval):
                    break
            if is_numeric_token_string(sval):
                toks = [t for t in str(sval).split("-") if t]
                prefix_ok = len(toks) >= len(expected_prefix) and all(
                    (i < len(toks) and toks[i] == expected_prefix[i]) for i in range(len(expected_prefix))
                )
                if prefix_ok and len(toks) == expected_len:
                    return rfb
            rfb += 1
            scans += 1
        return None

 

def _compute_iteration_bounds(
    *,
    child_anchor_row: int,
    region_bottom: int,
    anchor_row: int,
    step: int,
) -> Tuple[int, int]:
    """子のアンカー行と領域下端から、反復の開始・終了インデックスを算出する（1始まり）。"""
    i0_fb = 1 if anchor_row <= child_anchor_row else ((anchor_row - child_anchor_row) // step) + 1
    i1_fb = ((region_bottom - child_anchor_row) // step) + 1 if region_bottom >= child_anchor_row else 0
    return i0_fb, i1_fb

def process_container(
    container_name,
    container_def,
    workbook,
    generated_names,
    global_max_elements: Optional[int] = None,
    *,
    extraction_policy: Optional[ExtractionPolicy] = None,
):
    """コンテナ処理（range非依存・マルチシート集約対応）"""
    # デフォルト値
    policy = extraction_policy or _DEFAULT_EXTRACTION_POLICY
    direction = container_def.get("direction", "row")
    increment = container_def.get("increment", 0)
    logger.debug(
        f"コンテナ処理: {container_name}, direction={direction}, increment={increment}"
    )
    # 全シートを対象として、定義名と実セル値から推定する。
    sheeted_ranges: list[tuple[Optional[str], Optional[str]]] = enumerate_sheeted_ranges_sorted(workbook)

    next_index = 1  # シート横断のグローバル連番

    # 1) まずコンテナ直下の全フィールド名（定義名）をワークブック全体から収集
    try:
        global_field_names = get_cell_names_in_container_range(
            container_name, workbook, None
        )
    except Exception:
        global_field_names = {}

    # 2) 各シートで、そのフィールド定義名が座標解決できるかで nameful を判定
    nameful_sheets, per_sheet_positions = build_nameful_sheets_and_positions(
        container_name, workbook, sheeted_ranges
    )

    # 3) テンプレート座標は、nameful な最初のシートから確保
    template_positions, template_sheet = select_template_positions(
        nameful_sheets, per_sheet_positions, sheeted_ranges
    )

    logger.debug("SHEETS ORDER %s: %s", container_name, [s for s, _ in sheeted_ranges])
    logger.debug("NAMEFUL %s: %s", container_name, sorted(list(nameful_sheets)))
    logger.debug("TEMPLATE_SHEET %s: %s", container_name, template_sheet)
    for target_sheet, _ignored in sheeted_ranges:
        next_index = _process_sheet_for_container(
            container_name=container_name,
            container_def=container_def,
            workbook=workbook,
            generated_names=generated_names,
            target_sheet=target_sheet,
            base_cell_names=global_field_names,
            nameful_sheets=nameful_sheets,
            per_sheet_positions=per_sheet_positions,
            template_positions=template_positions,
            template_sheet=template_sheet,
            global_max_elements=global_max_elements,
            next_index=next_index,
            extraction_policy=extraction_policy,
        )


def iter_defined_name_destinations_all(cell_name: str, workbook):
    """同名のDefinedNameが複数存在する場合でも、全エントリのdestinationsを列挙する安全なイテレータ"""
    # openpyxlでは workbook.defined_names.definedName に生のリストがある
    dn_list = getattr(workbook.defined_names, "definedName", None)
    if dn_list is not None:
        for dn in dn_list:
            if getattr(dn, "name", None) == cell_name:
                for sn, coord in dn.destinations:
                    yield sn, coord
    else:
        # フォールバック: マッピングAPI（同名が上書きされている可能性あり）
        if cell_name in workbook.defined_names:
            dn = workbook.defined_names[cell_name]
            for sn, coord in dn.destinations:
                yield sn, coord


def get_cell_names_in_container_range(
    container_key, workbook, sheet_name: Optional[str] = None
):
    logger.debug("GET_NAMES prefix for %s sheet=%s", container_key, sheet_name)

    def _make_search_prefix(key: str) -> str:
        parts = key.split(".")
        return (".".join(parts[:-1]) + ".") if (parts and parts[-1].isdigit()) else (key + ".")

    search_prefix = _make_search_prefix(container_key)

    def _collect_all_defined_names() -> List[str]:
        dn_list = getattr(workbook.defined_names, "definedName", None)
        if dn_list is not None:
            return [getattr(dn, "name", "") for dn in dn_list if getattr(dn, "name", "")]
        return list(set(workbook.defined_names))

    all_names = _collect_all_defined_names()

    # 採用するセル名のマップと、フィールドごとの最小インデックスを管理
    cell_names: Dict[str, str] = {}
    best_index_by_field: Dict[str, int] = {}

    def _sheet_matches(name: str) -> bool:
        if sheet_name is None:
            return True
        for eff_sheet, coord in iter_defined_name_destinations_all(name, workbook):
            if not eff_sheet and "!" in coord:
                eff_sheet = coord.split("!", 1)[0]
            if eff_sheet == sheet_name:
                return True
        return False

    def _analyze_tail(name: str) -> tuple[str, int | None] | None:
        tail = name[len(search_prefix) :]
        tail_parts = tail.split(".") if tail else []
        if len(tail_parts) == 2:
            if tail_parts[0].isdigit() and not tail_parts[1].isdigit():
                return tail_parts[1], int(tail_parts[0])
            if tail_parts[1].isdigit() and not tail_parts[0].isdigit():
                child_anchor_prefix = f"{search_prefix}{tail_parts[0]}.{tail_parts[1]}."
                for n2 in all_names:
                    if not n2.startswith(child_anchor_prefix):
                        continue
                    suffix = n2[len(child_anchor_prefix) :]
                    first = suffix.split(".")[0] if suffix else ""
                    if first and not first.isdigit():
                        return None
                return tail_parts[0], int(tail_parts[1])
            return None
        if (
            len(tail_parts) >= 3
            and (not tail_parts[-1].isdigit())
            and tail_parts[-2].isdigit()
            and all(seg.isdigit() for seg in tail_parts[:-2])
        ):
            return tail_parts[-1], int(tail_parts[-2])
        return None

    # 決定論的にするためにソートして走査
    for name in sorted(all_names):
        if not name or not name.startswith(search_prefix):
            continue
        if not _sheet_matches(name):
            continue
        analyzed = _analyze_tail(name)
        if not analyzed:
            continue
        field_name, idx_val = analyzed

        # 同一フィールドに対しては、より小さい index を優先して採用（例: .1 を優先）
        if idx_val is None:
            # index 不明の場合は、未設定のときのみ採用
            if field_name not in cell_names:
                cell_names[field_name] = name
                logger.debug("GET_NAMES add %s -> %s (sheet=%s)", name, field_name, sheet_name)
                logger.debug("セル名発見: %s -> フィールド名: %s, idx=%s", name, field_name, idx_val)
        else:
            prev = best_index_by_field.get(field_name)
            if prev is None or idx_val < prev:
                best_index_by_field[field_name] = idx_val
                cell_names[field_name] = name
                logger.debug("GET_NAMES add %s -> %s (sheet=%s)", name, field_name, sheet_name)
                logger.debug("セル名発見: %s -> フィールド名: %s, idx=%s", name, field_name, idx_val)

    return cell_names


def generate_cell_name_for_element(container_key, element_index, field_name):
    """要素インデックスとフィールド名から動的セル名を生成"""
    # "json.orders.1" + element_index=2 + "date" → "json.orders.2.date"
    base_parts = container_key.split(".")
    if base_parts[-1].isdigit():
        base_parts[-1] = str(element_index)  # 末尾の数値を置換
    else:
        base_parts.append(str(element_index))

    if field_name:
        return ".".join(base_parts + [field_name])
    else:
        return ".".join(base_parts)


def generate_dynamic_cell_names_from_positions(
    container_key: str,
    base_positions: dict,
    element_count: int,
    direction: str,
    increment: int,
    generated_names: dict,
    workbook,
    start_index: int = 1,
    sheet_name: str | None = None,
    step_override: Optional[int] = None,
    *,
    force_emit_all: bool = False,
):
    """基準座標を直接受け取り、任意シートから値を読み取ってセル名を生成。
    1要素内の全フィールドが空値（None/""）の場合、その要素はスキップし、
    実際に生成した要素数を返す。
    """
    if not base_positions or element_count <= 0:
        return 0
    ws = (
        workbook[sheet_name]
        if (sheet_name and sheet_name in getattr(workbook, "sheetnames", []))
        else workbook.active
    )
    eff_step = _effective_step_for(step_override, increment)
    emitted = 0
    for local_i in range(1, element_count + 1):
        _values, _positions, non_empty = _read_fields_for_local_index_from_positions(
            local_idx=local_i,
            base_positions=base_positions,
            direction=direction,
            eff_step=eff_step,
            ws=ws,
        )
        _adjust_seq_like_values_in_row(_values, _positions, ws)

        if not non_empty and not force_emit_all:
            continue

        emitted += 1
        global_idx_emitted = start_index + emitted - 1
        _emit_generated_values(
            container_key=container_key,
            generated_names=generated_names,
            emitted_idx=global_idx_emitted,
            values=_values,
        )

    return emitted


def _effective_step_for(step_override: Optional[int], increment: int) -> int:
    return step_override if (isinstance(step_override, int) and step_override > 0) else increment


def _read_fields_for_local_index_from_positions(
    *,
    local_idx: int,
    base_positions: Mapping[str, Tuple[int, int]],
    direction: str,
    eff_step: int,
    ws,
) -> tuple[dict[str, Any], dict[str, tuple[int, int]], bool]:
    values: Dict[str, Any] = {}
    positions: Dict[str, Tuple[int, int]] = {}
    non_empty = False
    for fn, base_pos in base_positions.items():
        target_pos = calculate_target_position(base_pos, direction, local_idx, eff_step)
        cell_value = read_cell_value(target_pos, ws)
        values[fn] = cell_value
        try:
            positions[fn] = (target_pos[0], target_pos[1])
        except Exception:
            logger.debug("failed to record used_positions for %s", fn, exc_info=True)
        if cell_value not in (None, ""):
            non_empty = True
    return values, positions, non_empty


def _adjust_seq_like_values_in_row(values: dict[str, Any], positions: dict[str, tuple[int, int]], ws) -> None:
    try:
        seq_like_val = None
        seq_like_col = None
        primary_numeric_field = None
        for fn, vc in values.items():
            if is_numeric_token_string(vc):
                seq_like_val = vc
                seq_like_col, _ = positions.get(fn, (None, None))
                primary_numeric_field = fn
                break
        if seq_like_val is None:
            return
        for fn in list(values.keys()):
            if primary_numeric_field is not None and fn == primary_numeric_field:
                continue
            vc = values.get(fn)
            fc, fr = positions.get(fn, (None, None))
            if vc == seq_like_val and fr is not None:
                for d in (1, 2, 3):
                    alt_candidates: list[tuple[int, int]] = []
                    if seq_like_col is not None:
                        alt_candidates.append((seq_like_col + d, fr))
                    if fc is not None:
                        alt_candidates.append((fc + d, fr))
                    for ac, ar in alt_candidates:
                        alt_val = read_cell_value((ac, ar), ws)
                        if alt_val not in (None, "") and alt_val != seq_like_val:
                            values[fn] = alt_val
                            raise StopIteration
    except StopIteration:
        return
    except Exception:
        logger.debug("failed to adjust seq-like alternative value", exc_info=True)


def _emit_generated_values(
    *, container_key: str, generated_names: dict, emitted_idx: int, values: dict[str, Any]
) -> None:
    try:
        stats().cells_generated += len(values)
    except Exception:
        logger.debug("failed to add cells_generated count for emitted row", exc_info=True)
    for fn, val in values.items():
        generated_key = generate_cell_name_for_element(container_key, emitted_idx, fn)
        generated_names[generated_key] = val
        logger.debug("GEN-SET %s=%r", generated_key, val)


def sort_containers_by_hierarchy(containers, prefix: str = "json"):
    """コンテナを階層の深さでソート（浅い順）"""
    container_items = list(containers.items())
    return sorted(
        container_items, key=lambda x: calculate_hierarchy_depth(x[0], prefix=prefix)
    )


def detect_card_count_from_existing_names(
    base_container_name,
    workbook,
    sheet_name: Optional[str] = None,
    *,
    prefix: str = "json",
):
    """既存のセル名からカード数を検出"""
    card_indices = set()
    prefix_str = f"{prefix}.{base_container_name}."

    for name in workbook.defined_names.keys():
        if name.startswith(prefix_str):
            # json.card.1.customer_name -> ['1', 'customer_name']
            parts = name[len(prefix_str) :].split(".")
            if parts and parts[0].isdigit():
                # シート指定がある場合は定義名の宛先シートでフィルタ
                if sheet_name is not None:
                    try:
                        if not any(
                            sn == sheet_name
                            for sn, _ in iter_defined_name_destinations_all(
                                name, workbook
                            )
                        ):
                            continue
                    except Exception:
                        continue
                card_indices.add(int(parts[0]))

    return max(card_indices) if card_indices else 0


def get_cell_position_from_name(cell_name, workbook, sheet_name: Optional[str] = None):
    """セル名から座標位置を取得（範囲の場合は左上セル）。
    robust: dn.destinations の sheet 名が None でも coord に含まれる場合を考慮。
    """
    # 同名の複数定義を横断して、シート条件に一致する最初の座標を返す
    for sn, coord in iter_defined_name_destinations_all(cell_name, workbook):
        eff_sheet = sn
        eff_coord = coord
        # coord にシート名が含まれる形式（"Sheet1!$B$2"）に対応
        if (not eff_sheet) and isinstance(coord, str) and "!" in coord:
            try:
                sheet_part, cell_part = coord.split("!", 1)
                eff_sheet = sheet_part
                eff_coord = cell_part
            except Exception:
                eff_coord = coord
        # Worksheet 等をタイトルへ正規化
        if eff_sheet is not None and not isinstance(eff_sheet, str):
            try:
                eff_sheet = getattr(eff_sheet, "title", str(eff_sheet))
            except Exception:
                eff_sheet = str(eff_sheet)
        if sheet_name is not None and eff_sheet != sheet_name:
            continue
        # 単一セル
        m1 = re.match(r"^\$?([A-Z]+)\$?(\d+)$", eff_coord)
        if m1:
            col_letter, row_num = m1.groups()
            col_num = column_index_from_string(col_letter)
            return (col_num, int(row_num))
        # 範囲 → 先頭セル
        coord_clean = eff_coord.replace("$", "")
        m2 = re.match(r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$", coord_clean)
        if m2:
            start_col, start_row, _ec, _er = m2.groups()
            return (column_index_from_string(start_col), int(start_row))
    return None


def _process_sheet_for_container(
    *,
    container_name: str,
    container_def: dict,
    workbook,
    generated_names: Dict[str, Any],
    target_sheet: Optional[str],
    base_cell_names: Dict[str, str],
    nameful_sheets: set,
    per_sheet_positions: Dict[str, dict],
    template_positions: Dict[str, tuple],
    template_sheet: Optional[str],
    global_max_elements: Optional[int],
    next_index: int,
    extraction_policy: Optional[ExtractionPolicy],
) -> int:
    """Process a single sheet for the given container.

    This helper is an extraction of the per-sheet body of `process_container`.
    It intentionally performs no behavioral changes; arguments are passed through
    and the returned `next_index` will be applied by the caller.
    """
    # シートでフィルタして基準座標を取得（無ければテンプレを使用）
    using_template = False
    try:
        current_positions, using_template = gather_current_positions_and_template(
            container_name=container_name,
            container_def=container_def,
            workbook=workbook,
            target_sheet=target_sheet,
            nameful_sheets=nameful_sheets,
            template_positions=template_positions,
            template_sheet=template_sheet,
            global_field_names=base_cell_names,
        )

        if not current_positions:
            logger.debug(
                "コンテナ %s に既存のセル名が見つからずテンプレも無いためスキップ（sheet=%s）",
                container_name,
                target_sheet,
            )
            return next_index
    except Exception:
        # propagate to caller (process_container) as earlier behavior
        raise

    # Step 3: 実効 increment を推定（必要時）
    eff_increment, anchor_range_span = compute_effective_increment_and_anchor_span(
        container_name=container_name,
        container_def=container_def,
        workbook=workbook,
        target_sheet=target_sheet,
        global_field_names=base_cell_names,
    )

    # Step 3.5: 要素数の推定（labels による決定論的停止条件を適用）
    labels = _normalize_labels(container_def)

    # 件数・ステップ見積り（ラベル停止・範囲・スライス・矩形列・親境界を包含）
    (
        element_count,
        step_override,
        parent_range_span_for_container,
        range_count,
        internal_slice_count,
    ) = estimate_element_count_and_step(
        container_name=container_name,
        container_def=container_def,
        workbook=workbook,
        target_sheet=target_sheet,
        current_positions=current_positions,
        base_cell_names=base_cell_names,
        global_max_elements=global_max_elements,
        template_sheet=template_sheet,
        using_template=using_template,
        direction=container_def.get("direction", "row"),
        eff_increment=eff_increment,
        anchor_range_span=anchor_range_span,
        labels=labels,
    )

    element_count = _apply_global_max_cap(element_count, global_max_elements, next_index)

    if element_count <= 0:
        return next_index

    # ネストスキャン優先
    if _try_nested_scan_and_emit(
        workbook=workbook,
        target_sheet=target_sheet,
        container_name=container_name,
        direction=container_def.get("direction", "row"),
        current_positions=current_positions,
        labels=labels,
        policy=extraction_policy or _DEFAULT_EXTRACTION_POLICY,
        generated_names=generated_names,
    ):
        return next_index

    # フォールバック: 多数数値トークンのときは一括生成を抑止
    try:
        _parts_cn_chk = container_name.split(".")
        _num_positions_chk = [i for i, t in enumerate(_parts_cn_chk) if t.isdigit()]
        if len(_num_positions_chk) >= 2:
            logger.debug("NESTED-SKIP-FALLBACK %s (multi-numeric)", container_name)
            return next_index
    except Exception:
        return next_index

    # 動的セル名生成
    logger.debug(
        "GEN %s sheet=%s start=%s count=%s template=%s",
        container_name,
        target_sheet,
        next_index,
        element_count,
        using_template,
    )
    emitted = generate_dynamic_cell_names_from_positions(
        container_name,
        current_positions,
        element_count,
        container_def.get("direction", "row"),
        eff_increment,
        generated_names,
        workbook,
        start_index=next_index,
        sheet_name=target_sheet,
        step_override=step_override,
        force_emit_all=bool(range_count),
    )

    next_index += int(emitted or 0)
    return next_index


def get_sheet_from_defined_name(cell_name, workbook):
    """定義名から最初のシート名を取得（存在しない場合はNone）"""
    for sheet_name, _coord in iter_defined_name_destinations_all(cell_name, workbook):
        return sheet_name
    return None


def calculate_target_position(base_position, direction, instance_idx, increment):
    """基準位置からターゲット位置を計算"""
    base_col, base_row = base_position

    if direction == "row":
        return (base_col, base_row + (instance_idx - 1) * increment)
    else:  # column
        return (base_col + (instance_idx - 1) * increment, base_row)


def read_cell_value(position, worksheet):
    """指定位置からセル値を読み取り"""
    try:
        col, row = position
        cell = worksheet.cell(row=row, column=col)
        val = cell.value if cell.value is not None else ""
        stats().cells_read += 1
        if val == "":
            stats().empty_cells_skipped += 1
        return val
    except Exception as e:
        logger.warning(f"セル値読み取りエラー: {e}")
        return ""


def detect_card_count(base_positions, direction, increment, labels, worksheet):
    """カード数を検出"""
    # 簡易実装：最初のアイテムの位置からincrement間隔でラベル確認
    max_cards = 10  # 最大検索数
    card_count = 0

    if not base_positions:
        return 0

    first_item = list(base_positions.keys())[0]
    base_position = base_positions[first_item]

    for card_idx in range(1, max_cards + 1):
        target_position = calculate_target_position(
            base_position, direction, card_idx, increment
        )
        cell_value = read_cell_value(target_position, worksheet)

        if cell_value:  # 値がある場合はカードが存在
            card_count = card_idx
        else:
            break  # 空の場合は終了

    return card_count


# =============================================================================
# Main Function and CLI
# =============================================================================
def main():
    """メインエントリーポイント"""
    try:
        parser = create_argument_parser()
        args = parser.parse_args()
    except SystemExit:
        logger.error("引数の解析に失敗しました")
        return 1

    try:
        config = create_config_from_args(args)
        converter = Xlsx2JsonConverter(config)
        return converter.process_files(config.input_files)
    except (ConfigurationError, FileProcessingError) as e:
        logger.error(f"エラー: {e}")
        return 1
    except Exception as e:
        logger.error(f"予期しないエラー: {e}")
        return 1


def create_argument_parser() -> argparse.ArgumentParser:
    """コマンドライン引数パーサーを作成"""
    parser = argparse.ArgumentParser(
        description="Excel の名前付き範囲を JSON/YAML に変換（厳格な罫線・アンカーに基づく決定的抽出）"
    )
    parser.add_argument(
        "input_files",
        nargs="*",
        help="入力 Excel ファイル（.xlsx）を複数指定可。ディレクトリ指定時は再帰的に走査",
    )
    parser.add_argument("--config", type=Path, help="設定ファイル")
    parser.add_argument("--output-dir", "-o", type=Path, help="出力ディレクトリ")
    parser.add_argument(
        "--prefix",
        "-p",
        default="json",
        help="プレフィックス（定義名の接頭辞。例: json.表1.1.項目）",
    )
    parser.add_argument("--schema", "-s", type=Path, help="JSONスキーマファイル")
    parser.add_argument(
        "--output-format",
        "-f",
        choices=["json", "yaml"],
        default=None,
        help="出力フォーマット (json/yaml)。未指定時は json",
    )
    parser.add_argument(
        "--log-level",
        choices=["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"],
        help="ログレベル（デフォルト: INFO）",
    )
    parser.add_argument(
        "--trim", action="store_true", help="文字列の前後の空白を削除（既定はそのまま）"
    )
    parser.add_argument("--container", action="append", help="コンテナ定義")
    parser.add_argument("--transform", action="append", help="変換ルール")
    parser.add_argument(
        "--max-elements",
        type=int,
        help="全コンテナに共通で適用する要素数の上限（1以上の整数）。未指定時は無制限",
    )
    parser.add_argument(
        "--log-format",
        help="ログフォーマット（例: '%(asctime)s.%(msecs)03d %(levelname)s: %(message)s'。未指定時は日時付き標準フォーマット）",
    )
    parser.add_argument(
        "--log-datefmt",
        help="ログ日時フォーマット（例: '%Y/%m/%d %H:%M:%S'）。未指定時は '%Y/%m/%d %H:%M:%S'",
    )
    return parser


def create_config_from_args(args) -> ProcessingConfig:
    """コマンドライン引数から設定を作成"""
    raw = _load_config_file_from_args(args)
    merged = _apply_cli_overrides_to_config(args, raw)
    cli_cfg = CLIConfig(args=args, raw_config=raw, merged=merged)

    _configure_logging_from_config(cli_cfg.merged)

    if not cli_cfg.merged.get("input-files"):
        raise ConfigurationError("入力ファイルが指定されていません")

    schema_obj = _load_schema_from_config(cli_cfg.merged)
    return _build_processing_config_from_config(cli_cfg.merged, schema_obj)


def _load_config_file_from_args(args) -> Dict[str, Any]:
    cfg: Dict[str, Any] = {}
    if not args.config:
        return cfg
    try:
        with args.config.open("r", encoding="utf-8") as f:
            loaded = yaml.safe_load(f)
        if loaded is None:
            return {}
        if isinstance(loaded, dict):
            return loaded
        raise ConfigurationError("設定ファイルの形式が不正です（マップ型が必要です）")
    except yaml.YAMLError as e:
        raise ConfigurationError(f"設定ファイルの読み込みに失敗（YAML解析エラー）: {e}")
    except (FileNotFoundError, OSError, UnicodeDecodeError) as e:
        raise ConfigurationError(f"設定ファイルの読み込みに失敗: {e}")


def _apply_cli_overrides_to_config(args, cfg: Dict[str, Any]) -> Dict[str, Any]:
    if args.input_files:
        cfg["input-files"] = args.input_files
    if args.output_dir:
        cfg["output-dir"] = args.output_dir
    if args.prefix and "prefix" not in cfg:
        cfg["prefix"] = args.prefix
    if args.schema:
        cfg["schema"] = args.schema
    if args.trim:
        cfg["trim"] = True
    if args.log_level:
        cfg["log-level"] = args.log_level
    if args.container:
        validate_cli_containers(args.container, prefix=cfg.get("prefix", args.prefix or "json"))
        cli_containers = parse_container_args(args.container)
        cfg["containers"] = {**cfg.get("containers", {}), **cli_containers}
    if args.transform:
        cfg["transform"] = cfg.get("transform", []) + args.transform
    if args.output_format:
        cfg["output-format"] = args.output_format
    if args.max_elements is not None:
        cfg["max-elements"] = args.max_elements
    if args.log_format:
        cfg["log-format"] = args.log_format
    if args.log_datefmt:
        cfg["log-datefmt"] = args.log_datefmt
    return cfg


def _configure_logging_from_config(cfg: Dict[str, Any]) -> None:
    raw_log_level = cfg.get("log-level", "INFO")
    if isinstance(raw_log_level, str):
        log_level = getattr(logging, raw_log_level.upper(), logging.INFO)
    else:
        try:
            log_level = int(raw_log_level)
        except Exception:
            log_level = logging.INFO
    default_format = "%(asctime)s %(levelname)s: %(message)s"
    default_datefmt = "%Y/%m/%d %H:%M:%S"
    log_format = cfg.get("log-format", default_format)
    datefmt = str(cfg.get("log-datefmt")) if cfg.get("log-datefmt") not in (None, "") else default_datefmt
    logging.basicConfig(level=log_level, format=log_format, datefmt=datefmt)


def _load_schema_from_config(cfg: Dict[str, Any]):
    if not cfg.get("schema"):
        return None
    schema_path = Path(cfg["schema"])
    return SchemaLoader.load_schema(schema_path)


def _build_processing_config_from_config(cfg: Dict[str, Any], schema_obj) -> ProcessingConfig:
    output_dir_val = cfg.get("output-dir")
    if output_dir_val is not None:
        output_dir_val = Path(output_dir_val)
    return ProcessingConfig(
        input_files=cfg.get("input-files", []),
        prefix=cfg.get("prefix", "json"),
        trim=cfg.get("trim", False),
        output_dir=output_dir_val,
        output_format=cfg.get("output-format", "json"),
        schema=schema_obj,
        containers=cfg.get("containers", {}),
        transform_rules=cfg.get("transform", []),
        max_elements=(int(str(cfg.get("max-elements"))) if cfg.get("max-elements") not in (None, "") else None),
        log_format=(str(cfg.get("log-format")) if cfg.get("log-format") not in (None, "") else None),
    )


# =============================================================================
# Wildcard and 2D Array Transform Extensions
# =============================================================================


def apply_wildcard_transforms(
    data: dict, transform_rules: Dict[str, List[ArrayTransformRule]], prefix: str
) -> dict:
    """
    パターン（ワイルドカード含む／含まない）変換ルールを記載順で適用する。

    Args:
        data: 変換対象のデータ
        transform_rules: 変換ルール（キーはパスパターン。'*' を含まない場合は完全一致として扱う）
        prefix: プレフィックス

    Returns:
        変換後のデータ
    """
    if not transform_rules:
        return data

    # ヘルパー関数はモジュールレベル版を利用

    # 各変換ルールを適用
    for pattern, rule_list in transform_rules.items():
        # 非ワイルドカードかつ split のみのルールは既に挿入時点で適用済みのため二重適用をスキップ
        if "*" not in pattern:
            if all(r.transform_type == "split" for r in rule_list):
                continue
            # split と他種が混在する場合: split は先に適用済みとみなし除外し、残りのみ適用
            if any(r.transform_type == "split" for r in rule_list):
                filtered = [r for r in rule_list if r.transform_type != "split"]
            else:
                filtered = rule_list
            effective_rules = filtered
        else:
            effective_rules = rule_list

        if not effective_rules:
            continue

        # ワイルドカード有無に関わらず find_matching_paths で探索（'*' 無しは完全一致）
        matching_paths = find_matching_paths(data, pattern)
        for path in matching_paths:
            original_value = get_nested_value(data, path)
            if original_value is None:
                continue
            try:
                new_value = original_value
                for rule in effective_rules:
                    if is_json_list(new_value) and all(is_json_dict(e) for e in new_value):
                        transformed_elements = [rule.transform(elem) for elem in new_value]
                        new_value = transformed_elements
                    else:
                        new_value = rule.transform(new_value)
                if isinstance(new_value, dict):
                    dynamic_keys = [k for k in new_value.keys() if k.startswith(prefix) or '.' in k]
                    if dynamic_keys:
                        expand_and_insert_dict(
                            result_dict=new_value,
                            base_path=path,
                            prefix=prefix,
                            root_result=data,
                            transform_rules_map=None,
                            workbook=None,
                            apply_dynamic_rules=False,
                        )
                    else:
                        set_nested_value(data, path, new_value)
                elif is_json_list(new_value):
                    set_nested_value(data, path, new_value)
                else:
                    set_nested_value(data, path, new_value)
            except Exception as e:
                logger.error(
                    "ワイルドカード変換エラー: パス=%s, ルール=%s, エラー=%s", path, rule_list, e
                )

    return data

# 一般化名称（後方互換のためエイリアス）: 非ワイルドカードも含めたパターン変換適用
apply_pattern_transforms = apply_wildcard_transforms




# テスト互換性のための関数群（後方互換ラッパー）
# データクリーニング関数群（テスト互換性のため）
def is_empty_value(value: Any) -> bool:
    """空の値かどうかを判定"""
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    if isinstance(value, (list, dict)) and len(value) == 0:
        return True
    return False


def is_completely_empty(data: Any) -> bool:
    """データが完全に空かどうかを判定"""
    if is_empty_value(data):
        return True

    if isinstance(data, dict):
        return all(is_completely_empty(v) for v in data.values())
    elif isinstance(data, list):
        return all(is_completely_empty(item) for item in data)

    return False


def clean_empty_values(data: Any, suppress_empty: bool = True, *, schema: Optional[Dict[str, Any]] = None, _path: tuple[str, ...] = ()):  # noqa: E501, PLR0915
    """空の値をクリーニング。

    ポリシー:
    - 元が空配列([])は削除対象（従来互換）
    - 元が配列で要素は全て空(null/"")の場合は、必要に応じて [] を保持
      - 同階層に非空の兄弟がある場合
      - スキーマで該当プロパティが array/object として定義されている場合
    - オブジェクト内の更にネストした配列についても上記を再帰的に適用
    """

    def _all_empty_list(lst: list[Any]) -> bool:
        return all(is_completely_empty(it) for it in lst)

    def _all_empty_scalars_list(lst: list[Any]) -> bool:
        """全要素がスカラー空(None/空文字)のみか判定（入れ子のlist/dictは不可）。"""
        def _is_scalar_empty(x: Any) -> bool:
            if x is None:
                return True
            if isinstance(x, str) and x.strip() == "":
                return True
            return False
        return all(_is_scalar_empty(it) for it in lst)

    def _contains_empty_array(x: Any) -> bool:
        """辞書ツリー内に少なくとも1つの空配列([])が含まれるか。"""
        if isinstance(x, list):
            return len(x) == 0
        if isinstance(x, dict):
            for vv in x.values():
                if _contains_empty_array(vv):
                    return True
        return False

    def _is_object_schema(s: Optional[Dict[str, Any]]) -> bool:
        return isinstance(s, dict) and (
            s.get("type") == "object" or ("properties" in s)
        )

    def _is_array_schema(s: Optional[Dict[str, Any]]) -> bool:
        return isinstance(s, dict) and (
            s.get("type") == "array" or ("items" in s)
        )

    def _normalized_preserved_from_original(orig: Any, sub_schema: Optional[Dict[str, Any]] = None) -> Any:
        """元データから空形状を再構成して返す（再帰）。

        - 配列: 要素が全て空のとき [] を返す。元が空配列([])は保持しない（ここでは None）。
        - 辞書: 子を再帰処理し、何かしら保持対象（[] 等）が生成されれば辞書で返す。全て空なら {} を返すが、
                呼び出し側で必要に応じてドロップ判定を行う。
        - スカラー/None: スキーマが array/object を示す場合は [] を返す。それ以外は None。
        """
        # Schema hint
        sub_type = None
        if isinstance(sub_schema, dict):
            t = sub_schema.get("type")
            if isinstance(t, str):
                sub_type = t
        if isinstance(orig, list):
            if len(orig) == 0:
                return None  # 元が空配列は保持しない
            if _all_empty_scalars_list(orig):
                return []
            # 非空要素がある場合、通常のクリーン処理に任せる（ここでは None）
            return None
        if isinstance(orig, dict):
            props2 = sub_schema.get("properties", {}) if isinstance(sub_schema, dict) else {}
            out_d: Dict[str, Any] = {}
            for kk, vv in orig.items():
                ss = props2.get(kk) if isinstance(props2, dict) else None
                # 通常クリーン
                cleaned_sub = clean_empty_values(vv, suppress_empty, schema=ss, _path=(*_path, kk))
                if not (suppress_empty and is_completely_empty(cleaned_sub)):
                    out_d[kk] = cleaned_sub
                    continue
                # 空になったが、元が配列で全てスカラー空のみの場合は [] を保持
                if isinstance(vv, list) and len(vv) > 0 and _all_empty_scalars_list(vv):
                    out_d[kk] = []
                    continue
                # 更に内側で保持できるものがあるか再帰的に探索
                if isinstance(vv, dict):
                    preserved = _normalized_preserved_from_original(vv, ss)
                    if preserved is not None and (_contains_empty_array(preserved) or not is_completely_empty(preserved)):
                        out_d[kk] = preserved
            # 何も保持できない場合は None
            return out_d if out_d else None
        # スカラー/None の場合、スキーマに応じてプレースホルダーを返す
        if _is_object_schema(sub_schema):
            return {}
        if _is_array_schema(sub_schema):
            return []
        return None

    def _preserve_by_schema_and_data(d: Dict[str, Any], sch: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
        """スキーマと元データから空形状を再構成（配列は []、オブジェクトは {} を保持）。"""
        if not isinstance(d, dict) or not isinstance(sch, dict):
            return None
        props = sch.get("properties", {}) if isinstance(sch, dict) else {}
        out: Dict[str, Any] = {}
        for k, v in d.items():
            ssub = props.get(k) if isinstance(props, dict) else None
            if isinstance(v, dict):
                sub = _preserve_by_schema_and_data(v, ssub)
                if sub:
                    out[k] = sub
                elif _is_object_schema(ssub):
                    out[k] = {}
                elif _is_array_schema(ssub):
                    out[k] = []
            elif isinstance(v, list):
                if len(v) > 0 and _all_empty_scalars_list(v):
                    out[k] = []
                elif _is_object_schema(ssub):
                    out[k] = {}
                elif _is_array_schema(ssub):
                    out[k] = []
            else:
                if _is_object_schema(ssub):
                    out[k] = {}
                elif _is_array_schema(ssub):
                    out[k] = []
        return out if out else None

    if suppress_empty and is_completely_empty(data):
        # スキーマが無い場合だけ従来通りの早期リターン
        if not isinstance(schema, dict):
            if isinstance(data, list):
                return []
            if isinstance(data, dict):
                return {}
            return None

    if isinstance(data, dict):
        # まず子をクリーン
        props = schema.get("properties", {}) if isinstance(schema, dict) else {}
        cleaned_children: Dict[str, Any] = {}
        orig_children: Dict[str, Any] = {}
        for k, v in data.items():
            sub_schema = props.get(k) if isinstance(props, dict) else None
            orig_children[k] = v
            cleaned_children[k] = clean_empty_values(v, suppress_empty, schema=sub_schema, _path=(*_path, k))

        # 同階層に非空の兄弟が存在するか
        has_non_empty_sibling = any(
            not is_completely_empty(cv) for cv in cleaned_children.values()
        )

        result: Dict[str, Any] = {}
        for k, cleaned in cleaned_children.items():
            v = orig_children[k]
            sub_schema = props.get(k) if isinstance(props, dict) else None
            if not (suppress_empty and is_completely_empty(cleaned)):
                result[k] = cleaned
                continue

            # ここから空になった子の保持判定
            kept: Any = None
            # 1) 元が配列で全要素が空（かつ元の長さ>0）の場合は [] を保持
            if isinstance(v, list) and len(v) > 0 and _all_empty_scalars_list(v):
                kept = []
            # 2) 元が辞書の場合、内側に保持対象が作れるか探索（list-of-nulls を [] にする等）
            elif isinstance(v, dict):
                preserved = _normalized_preserved_from_original(v, sub_schema)
                if preserved is not None and (_contains_empty_array(preserved) or not is_completely_empty(preserved)):
                    kept = preserved
            # 3) スキーマが array/object を示す場合は型に応じてプレースホルダーを保持（兄弟が無い場合にも適用）
            elif isinstance(sub_schema, dict) and (
                _is_object_schema(sub_schema) or _is_array_schema(sub_schema)
            ):
                kept = {} if _is_object_schema(sub_schema) else []

            # 4) 兄弟が非空で、保持候補がある場合に採用
            if has_non_empty_sibling and kept is not None:
                result[k] = kept
            # 5) 兄弟が非空でない（=この階層が全空）場合でも、スキーマがあるなら採用
            elif (not has_non_empty_sibling) and kept is not None and isinstance(sub_schema, dict):
                result[k] = kept

        # 全て落ちて空になったが、スキーマがある場合は復元を試みる
        if not result and isinstance(schema, dict):
            preserved_top = _normalized_preserved_from_original(data, schema) or _preserve_by_schema_and_data(data, schema)
            # {} プレースホルダーも許可するため、辞書で1キー以上ある場合は採用
            if preserved_top is not None and (
                (isinstance(preserved_top, dict) and len(preserved_top) > 0)
                or _contains_empty_array(preserved_top)
                or not is_completely_empty(preserved_top)
            ):
                return preserved_top
        return result
    elif isinstance(data, list):
        cleaned_list: List[Any] = []
        item_schema = None
        if isinstance(schema, dict):
            item_schema = schema.get("items") if isinstance(schema.get("items"), dict) else None
        for item in data:
            cleaned = clean_empty_values(item, suppress_empty, schema=item_schema, _path=_path)
            if not (suppress_empty and is_completely_empty(cleaned)):
                cleaned_list.append(cleaned)
        # スキーマ参照なしでも、完全空リストは [] を返す（呼び出し側で扱う）
        if suppress_empty and not cleaned_list:
            return []
        return cleaned_list

    return data


if __name__ == "__main__":
    exit(main())
